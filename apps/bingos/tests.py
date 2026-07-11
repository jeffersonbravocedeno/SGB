import asyncio
import json
from contextlib import nullcontext
from datetime import timedelta
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from unittest.mock import AsyncMock, MagicMock, Mock, call, patch

from channels.layers import get_channel_layer
from channels.testing import WebsocketCommunicator
from openpyxl import load_workbook
from django import forms
from django.conf import settings
from django.contrib.auth.models import AnonymousUser, User
from django.contrib.messages import get_messages
from django.contrib.messages.storage.fallback import FallbackStorage
from django.core.exceptions import ImproperlyConfigured, PermissionDenied
from django.db import (
    DatabaseError,
    IntegrityError,
    connection,
    models as django_models,
)
from django.http import Http404, HttpResponse
from django.template.loader import render_to_string
from django.test import RequestFactory, SimpleTestCase, TestCase, override_settings
from django.urls import resolve, reverse
from django.utils import timezone

from apps.jugadores.models import Jugador

from .forms import (
    AccesoCartonPublicoForm,
    CompraCartonJugadorForm,
    CierreFinancieroBingoForm,
    CostoPremioMaterialBingoForm,
    GenerarAsignarCartonForm,
    GenerarCartonBingoForm,
    GastoOperativoBingoForm,
    MotivoAnulacionForm,
    PartidaBingoForm,
)
from .models import (
    Bingo,
    BingoCierreFinanciero,
    BingoGastoOperativo,
    BingoPremioMaterialCosto,
    Carton,
    CartonPartidaBingo,
    Partidabingo,
)
from .consumers import PartidaPublicaConsumer
from .realtime import (
    EVENTOS_REQUIEREN_RECARGA,
    construir_payload_publico_partida,
    nombre_grupo_partida,
    programar_publicacion_partida,
)
from .reportes import (
    PDF_CONTENT_TYPE,
    ReporteHibridoError,
    TIPO_HIBRIDO,
    TIPO_HISTORICO,
    XLSX_CONTENT_TYPE,
    calcular_recaudacion_registrada,
    construir_datos_reporte_partida,
    construir_filas_reporte_partida,
    construir_resumenes_cartones_bingo,
    generar_excel_cartones_partida,
    generar_excel_resumen_bingo,
    generar_pdf_reporte_partida,
)
from .services import (
    CASILLA_LIBRE,
    ESTADO_PARTIDA_CANCELADA,
    ESTADO_PARTIDA_DESEMPATE,
    ESTADO_PARTIDA_EN_CURSO,
    ESTADO_PARTIDA_EN_ESPERA,
    ESTADO_PARTIDA_FINALIZADA,
    ESTADO_PARTIDA_PAUSADA,
    ESTADO_PARTIDA_PROGRAMADA,
    BolaBingoError,
    BolilleroAgotadoError,
    CartonNoCompletoError,
    CartonAsignacionError,
    CartonPublicoError,
    CierreFinancieroError,
    DatosDesempateInvalidosError,
    DesempateError,
    DesempateIncompletoError,
    EstadoPartidaError,
    MatrizCartonInvalidaError,
    RondaCreacionError,
    ValidacionCartonError,
    acciones_disponibles_consola,
    aplicar_accion_consola,
    carton_tiene_bingo_completo,
    contar_numeros_marcados_carton,
    confirmar_y_finalizar_desempate,
    confirmar_y_finalizar_desempate_participaciones,
    construir_candidato_desempate_participacion,
    construir_preliquidacion_financiera_bingo,
    crear_carton_maestro_para_bingo,
    crear_ronda_con_participaciones,
    crear_y_asignar_carton,
    deserializar_matriz_carton_bingo,
    estado_partida_mostrar,
    evaluar_carton_en_partida,
    evaluar_participacion_en_partida,
    extraer_siguiente_bola,
    construir_matriz_marcada_carton,
    formatear_bola_bingo,
    generar_codigo_carton,
    generar_codigo_carton_bingo,
    generar_matriz_carton_bingo,
    letra_bingo,
    normalizar_estado_partida,
    normalizar_patron_ganador,
    obtener_etiqueta_patron_ganador,
    normalizar_candidatos_desempate,
    normalizar_candidatos_desempate_participaciones,
    construir_resumen_financiero_real_bingo,
    obtener_partidas_elegibles_venta_carton,
    obtener_participaciones_hibridas_partida,
    obtener_participacion_carton_en_partida,
    obtener_balotas_disponibles_desempate,
    obtener_numeros_carton,
    obtener_numeros_faltantes_carton,
    obtener_resultado_desempate,
    obtener_tiros_desempate,
    obtener_bolas_disponibles,
    parsear_bolas_cantadas,
    parsear_candidatos_desempate,
    carton_cumple_patron_ganador,
    obtener_numeros_faltantes_patron_ganador,
    preparar_cartones_para_validacion,
    preparar_datos_carton_jugador,
    preparar_datos_desempate,
    preparar_datos_tablero_publico,
    preparar_participaciones_hibridas_para_consola,
    preparar_resumen_patron_ganador,
    preparar_resumen_partida_publica,
    puede_asignar_cartones,
    registrar_costo_premio_material_bingo,
    registrar_gasto_operativo_bingo,
    total_casillas_patron_ganador,
    serializar_bolas_cantadas,
    serializar_candidatos_desempate,
    serializar_matriz_carton_bingo,
    sortear_balota_desempate,
    validar_carton_ganador,
    validar_participacion_ganadora,
    validar_venta_carton_para_bingo,
)
from .views import (
    _participaciones_hibridas_carton,
    _procesar_accion_consola,
    _preparar_datos_carton_hibrido,
    _resumen_carton_privado,
    _seleccionar_participacion_hibrida,
    acceder_carton_publico,
    bingo_carton_nuevo,
    bingo_detalle,
    bingo_finanzas,
    bingo_finanzas_cerrar,
    bingo_finanzas_costo_anular,
    bingo_finanzas_costo_registrar,
    bingo_finanzas_gasto_anular,
    bingo_finanzas_gasto_registrar,
    carton_publico,
    carton_editar,
    carton_nuevo,
    cartones_lista,
    comprar_carton_bingo,
    consola_operador,
    confirmar_desempate,
    desempate_operador,
    bingo_resumen_excel,
    partida_cartones_excel,
    partida_carton_nuevo,
    partida_carton_editar,
    partida_detalle,
    partida_reporte_pdf,
    partidas_lista,
    mi_carton_detalle,
    mis_cartones,
    partida_nueva,
    preliquidacion_financiera,
    sacar_bola,
    sala_juego_publica,
    sortear_desempate,
    tablero_publico,
    validar_carton,
)


def datos_partida_form(**overrides):
    data = {
        "idjugadorganador": "",
        "nombreronda": "1",
        "valorefectivo": "22",
        "premiomaterial": "cafsf",
        "estadopartida": "Programada",
        "patronganador": "carton_lleno",
        "bolascantadas": "[]",
        "ultimabola": "0",
        "haydesempate": "",
        "idbingadores": "",
        "bolamayordesempate": "",
        "horainicio": "2026-06-27T17:00",
        "horafin": "",
    }
    data.update(overrides)
    return data


def matriz_carton_prueba():
    return [
        [1, 16, 31, 46, 61],
        [2, 17, 32, 47, 62],
        [3, 18, CASILLA_LIBRE, 48, 63],
        [4, 19, 33, 49, 64],
        [5, 20, 34, 50, 65],
    ]


def candidatos_desempate_prueba(tiro_uno=None, tiro_dos=None):
    return [
        {
            "idjugador": 41,
            "jugador": "juan123",
            "cartones": [
                {"idcarton": 31, "codigocarton": "P20-C-31"},
            ],
            "tiro_desempate": tiro_uno,
        },
        {
            "idjugador": 42,
            "jugador": "maria456",
            "cartones": [
                {"idcarton": 32, "codigocarton": "P20-C-32"},
            ],
            "tiro_desempate": tiro_dos,
        },
    ]


def partida_publica_prueba(
    estado=ESTADO_PARTIDA_EN_CURSO,
    bolas=None,
    ultima=None,
    ganador=None,
    pk=20,
):
    bolas = [] if bolas is None else list(bolas)
    bingo = Bingo(
        idbingo=7,
        titulobingo="Bingo público",
        premiomayor=Decimal("500.00"),
    )
    return Partidabingo(
        idpartidabingo=pk,
        idbingo=bingo,
        idjugadorganador=ganador,
        nombreronda="Ronda pública",
        valorefectivo=Decimal("100.00"),
        premiomaterial="Canasta",
        estadopartida=estado,
        bolascantadas=serializar_bolas_cantadas(bolas),
        ultimabola=ultima if ultima is not None else (bolas[-1] if bolas else 0),
        haydesempate=False,
    )


def carton_publico_prueba(partida=None, codigo="PUBLICO-001", jugador=None):
    partida = partida or partida_publica_prueba()
    jugador = jugador or Jugador(idjugador=41, aliasjugador="jugador_publico")
    return Carton(
        idcarton=31,
        idjugador=jugador,
        idpartida=partida,
        codigocarton=codigo,
        matriznumeros=serializar_matriz_carton_bingo(matriz_carton_prueba()),
        estadocarton="Vendido",
        preciopagado=Decimal("9876.54"),
    )


class PartidaBingoFormTests(SimpleTestCase):
    def test_programada_es_valida_al_crear_partida(self):
        form = PartidaBingoForm(data=datos_partida_form(estadopartida="Programada"))

        self.assertTrue(form.is_valid(), form.errors)
        self.assertEqual(form.cleaned_data["estadopartida"], "Programada")

    def test_formulario_nuevo_incluye_patron_ganador_por_defecto(self):
        form = PartidaBingoForm()

        self.assertIn("patronganador", form.fields)
        self.assertEqual(form.fields["patronganador"].initial, "carton_lleno")
        self.assertEqual(form.fields["patronganador"].label, "Patrón ganador")
        self.assertEqual(form.fields["patronganador"].choices[0][0], "carton_lleno")

    def test_en_espera_es_valida_al_crear_partida(self):
        form = PartidaBingoForm(data=datos_partida_form(estadopartida="En espera"))

        self.assertTrue(form.is_valid(), form.errors)
        self.assertEqual(form.cleaned_data["estadopartida"], "En espera")

    def test_estado_inventado_es_rechazado(self):
        form = PartidaBingoForm(data=datos_partida_form(estadopartida="Estado inventado"))

        self.assertFalse(form.is_valid())
        self.assertIn("estadopartida", form.errors)

    def test_en_juego_legacy_se_normaliza_al_editar(self):
        partida = Partidabingo(idpartidabingo=1, estadopartida="En Juego")
        form = PartidaBingoForm(
            data=datos_partida_form(estadopartida="En Juego"),
            instance=partida,
        )

        self.assertTrue(form.is_valid(), form.errors)
        self.assertEqual(form.cleaned_data["estadopartida"], "En curso")

    def test_estados_nuevos_se_mantienen_reales_al_guardar_form(self):
        form = PartidaBingoForm(data=datos_partida_form(estadopartida="Programada"))

        self.assertTrue(form.is_valid(), form.errors)
        partida = form.save(commit=False)
        self.assertEqual(partida.estadopartida, "Programada")

    def test_programada_permite_cambiar_patron_ganador(self):
        partida = Partidabingo(
            idpartidabingo=1,
            estadopartida=ESTADO_PARTIDA_PROGRAMADA,
            patronganador="carton_lleno",
        )
        form = PartidaBingoForm(
            data=datos_partida_form(patronganador="cuatro_esquinas"),
            instance=partida,
        )

        self.assertTrue(form.is_valid(), form.errors)
        self.assertEqual(form.cleaned_data["patronganador"], "cuatro_esquinas")

    def test_en_curso_y_finalizada_conservan_patron_ganador_guardado(self):
        for estado in (ESTADO_PARTIDA_EN_CURSO, ESTADO_PARTIDA_FINALIZADA):
            with self.subTest(estado=estado):
                partida = Partidabingo(
                    idpartidabingo=2,
                    estadopartida=estado,
                    patronganador="cuatro_esquinas",
                )
                form = PartidaBingoForm(
                    data=datos_partida_form(patronganador="x"),
                    instance=partida,
                )

                self.assertTrue(form.is_valid(), form.errors)
                self.assertEqual(
                    form.cleaned_data["patronganador"],
                    "cuatro_esquinas",
                )

    def test_no_se_crearon_migraciones_de_bingos(self):
        migration_files = sorted(
            path.name
            for path in Path("apps/bingos/migrations").glob("*.py")
            if path.name != "__init__.py"
        )

        self.assertEqual(migration_files, [])


@override_settings(
    STORAGES={
        "default": {
            "BACKEND": "django.core.files.storage.FileSystemStorage",
        },
        "staticfiles": {
            "BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage",
        },
    }
)
class PlantillaPartidaFormularioPatronTests(SimpleTestCase):
    def setUp(self):
        self.factory = RequestFactory()
        self.bingo = Bingo(idbingo=7, titulobingo="Bingo de prueba")

    def _request(self, path):
        request = self.factory.get(path)
        request.user = User(username="admin", is_staff=True)
        return request

    def test_nueva_ronda_muestra_selector_de_patron_ganador(self):
        with patch(
            "apps.bingos.forms.Jugador.objects.order_by",
            return_value=Jugador.objects.none(),
        ):
            html = render_to_string(
                "bingos/partida_formulario.html",
                {
                    "form": PartidaBingoForm(),
                    "bingo": self.bingo,
                    "titulo": "Nueva ronda",
                },
                request=self._request("/partidas/nueva/"),
            )

        self.assertIn("Patrón ganador", html)
        self.assertIn("Cartón lleno", html)
        self.assertNotIn("Solo se puede cambiar cuando la ronda está Programada.", html)

    def test_edicion_no_programada_muestra_patron_solo_lectura(self):
        partida = Partidabingo(
            idpartidabingo=21,
            idbingo=self.bingo,
            nombreronda="Ronda 21",
            valorefectivo=Decimal("25.00"),
            premiomaterial="",
            estadopartida=ESTADO_PARTIDA_EN_CURSO,
            patronganador="cuatro_esquinas",
            bolascantadas="[]",
            ultimabola=0,
            haydesempate=False,
            horainicio=timezone.now(),
        )

        with patch(
            "apps.bingos.forms.Jugador.objects.order_by",
            return_value=Jugador.objects.none(),
        ):
            html = render_to_string(
                "bingos/partida_formulario.html",
                {
                    "form": PartidaBingoForm(instance=partida),
                    "partida": partida,
                    "bingo": self.bingo,
                    "titulo": "Editar partida",
                },
                request=self._request("/partidas/21/editar/"),
            )

        self.assertIn("Patrón ganador", html)
        self.assertIn("Cuatro esquinas", html)
        self.assertIn(
            "El patrón ganador no se puede cambiar porque la ronda ya no está Programada.",
            html,
        )


class PatronGanadorPartidaTests(SimpleTestCase):
    def test_partidabingo_reconoce_patronganador(self):
        field = Partidabingo._meta.get_field("patronganador")
        self.assertEqual(field.db_column, "patronganador")
        self.assertEqual(field.max_length, 20)
        self.assertEqual(field.default, "carton_lleno")
        self.assertEqual(
            dict(field.flatchoices),
            {
                "carton_lleno": "Cartón lleno",
                "linea_horizontal": "Línea horizontal",
                "linea_vertical": "Línea vertical",
                "diagonal": "Diagonal",
                "cuatro_esquinas": "Cuatro esquinas",
                "cruz": "Cruz",
                "x": "Letra X",
            },
        )

    def test_partidabingo_nueva_usa_carton_lleno_por_defecto(self):
        partida = Partidabingo()

        self.assertEqual(partida.patronganador, "carton_lleno")

    def test_ronda_existente_puede_leerse_con_carton_lleno(self):
        partida = Partidabingo(
            idpartidabingo=99,
            patronganador="carton_lleno",
        )

        self.assertEqual(partida.patronganador, "carton_lleno")


class CierreFinancieroModelMetadataTests(SimpleTestCase):
    def test_modelos_financieros_existen_y_apuntan_a_tablas_fisicas(self):
        casos = (
            (
                BingoGastoOperativo,
                "BingoGastoOperativo",
                "bingo_gasto_operativo",
                "idbingogastooperativo",
            ),
            (
                BingoPremioMaterialCosto,
                "BingoPremioMaterialCosto",
                "bingo_premio_material_costo",
                "idbingopremiomaterialcosto",
            ),
            (
                BingoCierreFinanciero,
                "BingoCierreFinanciero",
                "bingo_cierre_financiero",
                "idbingocierrefinanciero",
            ),
        )

        for model, nombre, db_table, pk_name in casos:
            with self.subTest(model=nombre):
                self.assertEqual(model.__name__, nombre)
                self.assertFalse(model._meta.managed)
                self.assertEqual(model._meta.db_table, db_table)
                self.assertEqual(model._meta.pk.name, pk_name)
                self.assertIsInstance(model._meta.pk, django_models.AutoField)

    def test_columnas_y_relaciones_relevantes_respetan_nombres_fisicos(self):
        campos = (
            (BingoGastoOperativo, "idbingo", "idbingo"),
            (BingoGastoOperativo, "idusuarioregistro", "idusuarioregistro"),
            (BingoGastoOperativo, "idusuarioanulacion", "idusuarioanulacion"),
            (BingoPremioMaterialCosto, "idbingo", "idbingo"),
            (BingoPremioMaterialCosto, "idpartidabingo", "idpartidabingo"),
            (
                BingoPremioMaterialCosto,
                "idusuarioregistro",
                "idusuarioregistro",
            ),
            (
                BingoPremioMaterialCosto,
                "idusuarioanulacion",
                "idusuarioanulacion",
            ),
            (BingoCierreFinanciero, "idbingo", "idbingo"),
            (BingoCierreFinanciero, "idusuariocierre", "idusuariocierre"),
        )

        for model, field_name, db_column in campos:
            with self.subTest(model=model.__name__, field=field_name):
                field = model._meta.get_field(field_name)
                self.assertEqual(field.db_column, db_column)

        self.assertIs(
            BingoGastoOperativo._meta.get_field("idbingo").remote_field.model,
            Bingo,
        )
        self.assertIs(
            BingoPremioMaterialCosto._meta.get_field("idbingo").remote_field.model,
            Bingo,
        )
        self.assertIs(
            BingoCierreFinanciero._meta.get_field("idbingo").remote_field.model,
            Bingo,
        )
        self.assertIs(
            BingoGastoOperativo._meta.get_field(
                "idusuarioregistro"
            ).remote_field.model,
            User,
        )
        self.assertIs(
            BingoCierreFinanciero._meta.get_field(
                "idusuariocierre"
            ).remote_field.model,
            User,
        )

    def test_idbingo_del_cierre_representa_unicidad_fisica(self):
        field = BingoCierreFinanciero._meta.get_field("idbingo")

        self.assertIsInstance(field, django_models.OneToOneField)
        self.assertTrue(field.unique)
        self.assertEqual(field.db_column, "idbingo")

    def test_idpartidabingo_no_inventa_fk_compuesta_falsa(self):
        field = BingoPremioMaterialCosto._meta.get_field("idpartidabingo")

        self.assertIsInstance(field, django_models.IntegerField)
        self.assertNotIsInstance(field, django_models.ForeignKey)
        self.assertEqual(field.db_column, "idpartidabingo")

    def test_estados_tienen_choices_y_defaults_esperados(self):
        casos = (
            (
                BingoGastoOperativo,
                {
                    "Registrado": "Registrado",
                    "Anulado": "Anulado",
                },
                "Registrado",
            ),
            (
                BingoPremioMaterialCosto,
                {
                    "Registrado": "Registrado",
                    "Anulado": "Anulado",
                },
                "Registrado",
            ),
            (
                BingoCierreFinanciero,
                {
                    "Abierto": "Abierto",
                    "Cerrado": "Cerrado",
                },
                "Abierto",
            ),
        )

        for model, choices, default in casos:
            with self.subTest(model=model.__name__):
                field = model._meta.get_field("estado")
                self.assertEqual(dict(field.flatchoices), choices)
                self.assertEqual(field.default, default)
                self.assertEqual(model().estado, default)


class GeneracionMatrizCartonTests(SimpleTestCase):
    def setUp(self):
        self.matriz = generar_matriz_carton_bingo()

    def test_matriz_tiene_cinco_filas_y_cinco_columnas(self):
        self.assertEqual(len(self.matriz), 5)
        self.assertTrue(all(len(fila) == 5 for fila in self.matriz))

    def test_columna_b_contiene_numeros_del_1_al_15(self):
        self.assertTrue(all(1 <= fila[0] <= 15 for fila in self.matriz))

    def test_columna_i_contiene_numeros_del_16_al_30(self):
        self.assertTrue(all(16 <= fila[1] <= 30 for fila in self.matriz))

    def test_columna_n_tiene_cuatro_numeros_y_centro_libre(self):
        columna_n = [fila[2] for fila in self.matriz]
        self.assertEqual(columna_n[2], CASILLA_LIBRE)
        numeros = [valor for valor in columna_n if valor != CASILLA_LIBRE]
        self.assertEqual(len(numeros), 4)
        self.assertTrue(all(31 <= numero <= 45 for numero in numeros))

    def test_columna_g_contiene_numeros_del_46_al_60(self):
        self.assertTrue(all(46 <= fila[3] <= 60 for fila in self.matriz))

    def test_columna_o_contiene_numeros_del_61_al_75(self):
        self.assertTrue(all(61 <= fila[4] <= 75 for fila in self.matriz))

    def test_no_hay_numeros_repetidos(self):
        numeros = [
            valor
            for fila in self.matriz
            for valor in fila
            if valor != CASILLA_LIBRE
        ]
        self.assertEqual(len(numeros), len(set(numeros)))

    def test_matriz_se_serializa_como_json_en_lista_de_filas(self):
        serializada = serializar_matriz_carton_bingo(self.matriz)

        self.assertEqual(deserializar_matriz_carton_bingo(serializada), self.matriz)
        self.assertIn('"LIBRE"', serializada)


class AsignacionCartonTests(SimpleTestCase):
    def _crear_carton_sin_base_de_datos(self, estado):
        partida = Partidabingo(
            idpartidabingo=12,
            estadopartida=estado,
        )
        partida_bloqueada = Partidabingo(
            idpartidabingo=12,
            estadopartida=estado,
        )
        jugador = Jugador(idjugador=4)

        def asignar_pk(carton):
            carton.idcarton = 30
            return carton

        with (
            patch(
                "apps.bingos.services.transaction.atomic",
                return_value=nullcontext(),
            ) as atomic_mock,
            patch(
                "apps.bingos.services.Partidabingo.objects.select_for_update"
            ) as select_for_update,
            patch(
                "apps.bingos.services.assign_next_integer_pk",
                side_effect=asignar_pk,
            ),
            patch(
                "apps.bingos.services.generar_codigo_carton",
                return_value="P12-C-ABC1234567",
            ),
            patch.object(Carton, "full_clean") as full_clean_mock,
            patch.object(Carton, "save") as save_mock,
        ):
            select_for_update.return_value.get.return_value = partida_bloqueada
            carton = crear_y_asignar_carton(
                partida=partida,
                jugador=jugador,
                precio_pagado=Decimal("5.00"),
            )

        atomic_mock.assert_called_once_with()
        full_clean_mock.assert_called_once_with(
            validate_unique=False,
            validate_constraints=False,
        )
        save_mock.assert_called_once_with(force_insert=True)
        return carton

    def test_no_se_asigna_carton_a_partida_en_curso(self):
        partida = Partidabingo(
            idpartidabingo=1,
            estadopartida=ESTADO_PARTIDA_EN_CURSO,
        )

        with patch.object(Carton, "save") as save_mock:
            with self.assertRaises(CartonAsignacionError):
                crear_y_asignar_carton(partida, Jugador(idjugador=1), "5.00")

        save_mock.assert_not_called()

    def test_no_se_asigna_carton_a_partida_pausada(self):
        partida = Partidabingo(
            idpartidabingo=1,
            estadopartida=ESTADO_PARTIDA_PAUSADA,
        )

        with patch.object(Carton, "save") as save_mock:
            with self.assertRaises(CartonAsignacionError):
                crear_y_asignar_carton(partida, Jugador(idjugador=1), "5.00")

        save_mock.assert_not_called()

    def test_desempate_finalizada_y_cancelada_bloquean_asignacion(self):
        for estado in (
            ESTADO_PARTIDA_DESEMPATE,
            ESTADO_PARTIDA_FINALIZADA,
            ESTADO_PARTIDA_CANCELADA,
        ):
            with self.subTest(estado=estado):
                partida = Partidabingo(estadopartida=estado)
                self.assertFalse(puede_asignar_cartones(partida))

    def test_se_genera_carton_para_partida_programada(self):
        carton = self._crear_carton_sin_base_de_datos(
            ESTADO_PARTIDA_PROGRAMADA
        )

        self.assertEqual(carton.estadocarton, "Vendido")
        self.assertEqual(carton.indicevictoria, 0)
        self.assertEqual(len(deserializar_matriz_carton_bingo(carton.matriznumeros)), 5)

    def test_se_genera_carton_para_partida_en_espera(self):
        carton = self._crear_carton_sin_base_de_datos(
            ESTADO_PARTIDA_EN_ESPERA
        )

        self.assertEqual(carton.idpartida.estadopartida, ESTADO_PARTIDA_EN_ESPERA)
        self.assertEqual(carton.preciopagado, Decimal("5.00"))

    def test_codigo_del_carton_se_genera_de_forma_unica(self):
        codigos_existentes = set()
        tokens = iter(("A" * 32, "B" * 32))

        primero = generar_codigo_carton(
            12,
            existe_codigo=codigos_existentes.__contains__,
            generador_token=lambda: next(tokens),
        )
        codigos_existentes.add(primero)
        segundo = generar_codigo_carton(
            12,
            existe_codigo=codigos_existentes.__contains__,
            generador_token=lambda: next(tokens),
        )

        self.assertNotEqual(primero, segundo)
        self.assertEqual(len({primero, segundo}), 2)
        self.assertTrue(primero.startswith("P12-C-"))

    def test_no_se_crea_carton_si_falla_validacion_del_precio(self):
        partida = Partidabingo(
            idpartidabingo=1,
            estadopartida=ESTADO_PARTIDA_PROGRAMADA,
        )

        with (
            patch("apps.bingos.services.transaction.atomic") as atomic_mock,
            patch.object(Carton, "save") as save_mock,
        ):
            with self.assertRaises(CartonAsignacionError):
                crear_y_asignar_carton(partida, Jugador(idjugador=1), "0")

        atomic_mock.assert_not_called()
        save_mock.assert_not_called()


class CartonMaestroBingoTests(SimpleTestCase):
    def setUp(self):
        self.bingo = Bingo(idbingo=7, titulobingo="Bingo híbrido")
        self.jugador = Jugador(idjugador=4)
        self.fecha_compra = timezone.now()
        self.partidas = [
            Partidabingo(
                idpartidabingo=21,
                idbingo=self.bingo,
                nombreronda="Ronda 1",
                estadopartida=ESTADO_PARTIDA_PROGRAMADA,
            ),
            Partidabingo(
                idpartidabingo=22,
                idbingo=self.bingo,
                nombreronda="Ronda 2",
                estadopartida=ESTADO_PARTIDA_EN_ESPERA,
            ),
        ]

    def _crear_sin_base(
        self,
        partidas=None,
        error_participacion=None,
        capturar_error=False,
    ):
        partidas = self.partidas if partidas is None else partidas

        def asignar_pk(carton):
            carton.idcarton = 40
            return carton

        carton = None
        error = None
        with (
            patch(
                "apps.bingos.services.transaction.atomic",
                return_value=nullcontext(),
            ) as atomic_mock,
            patch(
                "apps.bingos.services.Bingo.objects.select_for_update"
            ) as lock_bingo,
            patch(
                "apps.bingos.services.Partidabingo.objects.select_for_update"
            ) as lock_partidas,
            patch(
                "apps.bingos.services.assign_next_integer_pk",
                side_effect=asignar_pk,
            ) as assign_pk_mock,
            patch(
                "apps.bingos.services.generar_codigo_carton_bingo",
                return_value="B7-C-ABC1234567",
            ) as codigo_mock,
            patch(
                "apps.bingos.services.generar_matriz_carton_bingo",
                return_value=matriz_carton_prueba(),
            ) as matriz_mock,
            patch.object(Carton, "full_clean") as full_clean_mock,
            patch.object(Carton, "save") as save_mock,
            patch(
                "apps.bingos.services.CartonPartidaBingo.objects.create",
                side_effect=error_participacion,
            ) as crear_participacion_mock,
        ):
            lock_bingo.return_value.get.return_value = self.bingo
            consulta = lock_partidas.return_value.filter.return_value
            consulta.order_by.return_value = partidas
            try:
                carton = crear_carton_maestro_para_bingo(
                    bingo=self.bingo,
                    jugador=self.jugador,
                    precio_pagado=Decimal("5.00"),
                    fecha_compra=self.fecha_compra,
                )
            except Exception as exc:
                if not capturar_error:
                    raise
                error = exc

        return {
            "carton": carton,
            "error": error,
            "atomic": atomic_mock,
            "lock_bingo": lock_bingo,
            "lock_partidas": lock_partidas,
            "assign_pk": assign_pk_mock,
            "codigo": codigo_mock,
            "matriz": matriz_mock,
            "full_clean": full_clean_mock,
            "save": save_mock,
            "crear_participacion": crear_participacion_mock,
        }

    def test_crea_un_maestro_y_una_participacion_por_partida(self):
        resultado = self._crear_sin_base()
        carton = resultado["carton"]

        self.assertEqual(carton.idbingo, self.bingo)
        self.assertIsNone(carton.idpartida)
        self.assertIsNone(carton.indicevictoria)
        self.assertEqual(carton.idjugador, self.jugador)
        self.assertEqual(carton.codigocarton, "B7-C-ABC1234567")
        self.assertEqual(carton.preciopagado, Decimal("5.00"))
        self.assertEqual(carton.fechacompra, self.fecha_compra)
        self.assertEqual(carton.estadocarton, "Vendido")
        self.assertEqual(
            deserializar_matriz_carton_bingo(carton.matriznumeros),
            matriz_carton_prueba(),
        )
        resultado["atomic"].assert_called_once_with()
        resultado["lock_bingo"].return_value.get.assert_called_once_with(pk=7)
        resultado["lock_partidas"].return_value.filter.assert_called_once_with(
            idbingo=self.bingo
        )
        resultado["lock_partidas"].return_value.filter.return_value.order_by.assert_called_once_with(
            "idpartidabingo"
        )
        resultado["assign_pk"].assert_called_once_with(carton)
        resultado["codigo"].assert_called_once_with(7)
        resultado["matriz"].assert_called_once_with()
        resultado["full_clean"].assert_called_once_with(
            validate_unique=False,
            validate_constraints=False,
        )
        resultado["save"].assert_called_once_with(force_insert=True)
        self.assertEqual(resultado["crear_participacion"].call_count, 2)

        partidas_creadas = []
        for llamada in resultado["crear_participacion"].call_args_list:
            valores = llamada.kwargs
            partidas_creadas.append(valores["idpartida"])
            self.assertEqual(valores["idcarton"], carton)
            self.assertEqual(valores["idbingo"], self.bingo)
            self.assertEqual(
                valores["estado_participacion"],
                CartonPartidaBingo.ESTADO_PENDIENTE,
            )
            self.assertIsNone(valores["indicevictoria"])
            self.assertFalse(valores["es_asignacion_original"])
            self.assertEqual(
                valores["origen_asignacion"],
                CartonPartidaBingo.ORIGEN_APLICACION,
            )
            self.assertEqual(valores["fechacreacion"], self.fecha_compra)
            self.assertIsNone(valores["fechavalidacion"])
            self.assertNotIn("idcartonpartidabingo", valores)
            self.assertNotIn("preciopagado", valores)

        self.assertEqual(partidas_creadas, self.partidas)

    def test_codigo_nuevo_usa_prefijo_de_bingo(self):
        codigo = generar_codigo_carton_bingo(
            17,
            existe_codigo=lambda _codigo: False,
            generador_token=lambda: "abc-def-1234567890",
        )

        self.assertTrue(codigo.startswith("B17-C-"))
        self.assertLessEqual(len(codigo), 30)

    def test_rechaza_bingo_sin_partidas(self):
        with self.assertRaisesMessage(
            CartonAsignacionError,
            "ya no quedan rondas futuras disponibles",
        ):
            validar_venta_carton_para_bingo(self.bingo, [])

    def test_venta_tardia_crea_solo_participaciones_elegibles_y_precio_unico(self):
        partidas_no_elegibles = [
            Partidabingo(
                idpartidabingo=18,
                idbingo=self.bingo,
                nombreronda="Ronda finalizada",
                estadopartida=ESTADO_PARTIDA_FINALIZADA,
            ),
            Partidabingo(
                idpartidabingo=19,
                idbingo=self.bingo,
                nombreronda="Ronda en curso",
                estadopartida=ESTADO_PARTIDA_EN_CURSO,
            ),
            Partidabingo(
                idpartidabingo=20,
                idbingo=self.bingo,
                nombreronda="Ronda cancelada",
                estadopartida=ESTADO_PARTIDA_CANCELADA,
            ),
        ]
        resultado = self._crear_sin_base(
            [*partidas_no_elegibles, *self.partidas]
        )

        carton = resultado["carton"]
        partidas_creadas = [
            llamada.kwargs["idpartida"]
            for llamada in resultado["crear_participacion"].call_args_list
        ]
        self.assertEqual(partidas_creadas, self.partidas)
        self.assertEqual(carton.preciopagado, Decimal("5.00"))
        for llamada in resultado["crear_participacion"].call_args_list:
            self.assertNotIn("preciopagado", llamada.kwargs)

    def test_venta_tardia_excluye_pausada_y_desempate(self):
        pausada = Partidabingo(
            idpartidabingo=25,
            idbingo=self.bingo,
            nombreronda="Ronda pausada",
            estadopartida=ESTADO_PARTIDA_PAUSADA,
        )
        desempate = Partidabingo(
            idpartidabingo=26,
            idbingo=self.bingo,
            nombreronda="Ronda desempate",
            estadopartida=ESTADO_PARTIDA_DESEMPATE,
        )

        self.assertEqual(
            obtener_partidas_elegibles_venta_carton(
                [pausada, desempate, *self.partidas]
            ),
            self.partidas,
        )

    def test_ronda_en_espera_es_elegible_aunque_haya_una_finalizada(self):
        finalizada = Partidabingo(
            idpartidabingo=20,
            idbingo=self.bingo,
            estadopartida=ESTADO_PARTIDA_FINALIZADA,
        )

        self.assertEqual(
            obtener_partidas_elegibles_venta_carton(
                [finalizada, self.partidas[1]]
            ),
            [self.partidas[1]],
        )

    def test_sin_rondas_elegibles_no_crea_maestro_ni_participaciones(self):
        partidas = [
            Partidabingo(
                idpartidabingo=pk,
                idbingo=self.bingo,
                nombreronda=f"Ronda {pk}",
                estadopartida=estado,
            )
            for pk, estado in (
                (18, ESTADO_PARTIDA_FINALIZADA),
                (19, ESTADO_PARTIDA_EN_CURSO),
                (20, ESTADO_PARTIDA_CANCELADA),
            )
        ]

        resultado = self._crear_sin_base(
            partidas,
            capturar_error=True,
        )

        self.assertIsInstance(resultado["error"], CartonAsignacionError)
        self.assertIn(
            "ya no quedan rondas futuras disponibles",
            str(resultado["error"]),
        )
        self.assertIsNone(resultado["carton"])
        resultado["assign_pk"].assert_not_called()
        resultado["codigo"].assert_not_called()
        resultado["matriz"].assert_not_called()
        resultado["full_clean"].assert_not_called()
        resultado["save"].assert_not_called()
        resultado["crear_participacion"].assert_not_called()

    def test_acepta_programada_y_en_espera(self):
        self.assertEqual(
            validar_venta_carton_para_bingo(self.bingo, self.partidas),
            self.partidas,
        )

    def test_ronda_nueva_incorpora_carton_vendido_tarde(self):
        finalizada = Partidabingo(
            idpartidabingo=20,
            idbingo=self.bingo,
            estadopartida=ESTADO_PARTIDA_FINALIZADA,
        )
        carton = self._crear_sin_base(
            [finalizada, self.partidas[0]]
        )["carton"]
        nueva_ronda = Partidabingo(
            idbingo=self.bingo,
            nombreronda="Ronda creada después",
            estadopartida=ESTADO_PARTIDA_PROGRAMADA,
        )

        def asignar_pk(partida):
            partida.idpartidabingo = 30
            return partida

        with (
            patch(
                "apps.bingos.services.transaction.atomic",
                return_value=nullcontext(),
            ),
            patch(
                "apps.bingos.services.Bingo.objects.select_for_update"
            ) as lock_bingo,
            patch(
                "apps.bingos.services.Carton.objects.select_for_update"
            ) as lock_cartones,
            patch(
                "apps.bingos.services.CartonPartidaBingo.objects.select_for_update"
            ) as lock_participaciones,
            patch(
                "apps.bingos.services.assign_next_integer_pk",
                side_effect=asignar_pk,
            ),
            patch.object(Partidabingo, "save"),
            patch(
                "apps.bingos.services.CartonPartidaBingo.objects.create"
            ) as crear_participacion,
        ):
            lock_bingo.return_value.get.return_value = self.bingo
            consulta_cartones = lock_cartones.return_value.filter.return_value
            consulta_cartones.order_by.return_value = [carton]
            consulta_participaciones = (
                lock_participaciones.return_value.filter.return_value
            )
            consulta_participaciones.values_list.return_value = ()

            crear_ronda_con_participaciones(
                self.bingo,
                nueva_ronda,
                fecha_creacion=self.fecha_compra,
            )

        crear_participacion.assert_called_once()
        self.assertIs(
            crear_participacion.call_args.kwargs["idcarton"],
            carton,
        )
        self.assertIs(
            crear_participacion.call_args.kwargs["idpartida"],
            nueva_ronda,
        )

    def test_error_de_participacion_se_propaga_para_rollback_atomico(self):
        with self.assertRaisesMessage(RuntimeError, "fallo de participacion"):
            self._crear_sin_base(
                error_participacion=RuntimeError("fallo de participacion")
            )


class SincronizacionRondaParticipacionesTests(SimpleTestCase):
    def setUp(self):
        self.bingo = Bingo(idbingo=7, titulobingo="Bingo híbrido")
        self.otro_bingo = Bingo(idbingo=8, titulobingo="Otro Bingo")
        self.jugador = Jugador(idjugador=4)
        self.fecha_creacion = timezone.now()
        self.partida = Partidabingo(
            idbingo=self.bingo,
            nombreronda="Ronda 4",
            estadopartida=ESTADO_PARTIDA_PROGRAMADA,
            valorefectivo=Decimal("25.00"),
        )

    def _carton(
        self,
        pk,
        *,
        bingo=None,
        partida=None,
        estado="Vendido",
        jugador=None,
        precio="5.00",
    ):
        return Carton(
            idcarton=pk,
            idbingo=bingo or self.bingo,
            idpartida=partida,
            idjugador=self.jugador if jugador is None else jugador,
            codigocarton=f"B7-C-{pk}",
            estadocarton=estado,
            preciopagado=Decimal(precio),
        )

    def _crear_sin_base(
        self,
        cartones,
        *,
        participaciones_existentes=(),
        error_participacion=None,
        capturar_error=False,
    ):
        def asignar_pk(partida):
            partida.idpartidabingo = 30
            return partida

        atomic_context = MagicMock()
        resultado = None
        error = None
        with (
            patch(
                "apps.bingos.services.transaction.atomic",
                return_value=atomic_context,
            ) as atomic_mock,
            patch(
                "apps.bingos.services.Bingo.objects.select_for_update"
            ) as lock_bingo,
            patch(
                "apps.bingos.services.Carton.objects.select_for_update"
            ) as lock_cartones,
            patch(
                "apps.bingos.services.CartonPartidaBingo.objects.select_for_update"
            ) as lock_participaciones,
            patch(
                "apps.bingos.services.assign_next_integer_pk",
                side_effect=asignar_pk,
            ) as assign_pk_mock,
            patch.object(Partidabingo, "save") as guardar_partida_mock,
            patch.object(Carton, "save") as guardar_carton_mock,
            patch(
                "apps.bingos.services.CartonPartidaBingo.objects.create",
                side_effect=error_participacion,
            ) as crear_participacion_mock,
        ):
            lock_bingo.return_value.get.return_value = self.bingo
            consulta_cartones = lock_cartones.return_value.filter.return_value
            consulta_cartones.order_by.return_value = cartones
            consulta_participaciones = (
                lock_participaciones.return_value.filter.return_value
            )
            consulta_participaciones.values_list.return_value = (
                participaciones_existentes
            )
            try:
                resultado = crear_ronda_con_participaciones(
                    bingo=self.bingo,
                    partida=self.partida,
                    fecha_creacion=self.fecha_creacion,
                )
            except Exception as exc:
                if not capturar_error:
                    raise
                error = exc

        return {
            "resultado": resultado,
            "error": error,
            "atomic": atomic_mock,
            "atomic_context": atomic_context,
            "lock_bingo": lock_bingo,
            "lock_cartones": lock_cartones,
            "lock_participaciones": lock_participaciones,
            "assign_pk": assign_pk_mock,
            "guardar_partida": guardar_partida_mock,
            "guardar_carton": guardar_carton_mock,
            "crear_participacion": crear_participacion_mock,
        }

    def test_tres_cartones_maestros_validos_crean_tres_participaciones(self):
        cartones = [self._carton(pk) for pk in (40, 41, 42)]

        prueba = self._crear_sin_base(cartones)

        self.assertEqual(self.partida.idbingo, self.bingo)
        self.assertEqual(self.partida.pk, 30)
        self.assertEqual(
            len(prueba["resultado"]["participaciones_creadas"]),
            3,
        )
        prueba["atomic"].assert_called_once_with()
        prueba["lock_bingo"].return_value.get.assert_called_once_with(pk=7)
        prueba["lock_cartones"].assert_called_once_with(of=("self",))
        prueba["lock_cartones"].return_value.filter.assert_called_once_with(
            idbingo=self.bingo,
            idpartida__isnull=True,
        )
        prueba["assign_pk"].assert_called_once_with(self.partida)
        prueba["guardar_partida"].assert_called_once_with(force_insert=True)
        prueba["guardar_carton"].assert_not_called()
        self.assertEqual(prueba["crear_participacion"].call_count, 3)
        self.assertEqual(
            {
                llamada.kwargs["idcarton"].pk
                for llamada in prueba["crear_participacion"].call_args_list
            },
            {40, 41, 42},
        )
        for llamada in prueba["crear_participacion"].call_args_list:
            self.assertEqual(llamada.kwargs["idpartida"], self.partida)
            self.assertEqual(llamada.kwargs["idbingo"], self.bingo)
            self.assertEqual(
                llamada.kwargs["estado_participacion"],
                CartonPartidaBingo.ESTADO_PENDIENTE,
            )

    def test_excluye_otro_bingo_historicos_y_cartones_no_validos(self):
        partida_historica = Partidabingo(
            idpartidabingo=11,
            idbingo=self.bingo,
        )
        cartones = [
            self._carton(40),
            self._carton(41, bingo=self.otro_bingo),
            self._carton(42, partida=partida_historica),
            self._carton(43, estado="Disponible"),
            self._carton(44, estado="Cerrado"),
            self._carton(45, jugador=Jugador()),
        ]

        prueba = self._crear_sin_base(cartones)

        prueba["crear_participacion"].assert_called_once()
        self.assertEqual(
            prueba["crear_participacion"].call_args.kwargs["idcarton"].pk,
            40,
        )

    def test_no_duplica_una_participacion_ya_existente(self):
        cartones = [self._carton(pk) for pk in (40, 41, 42)]

        prueba = self._crear_sin_base(
            cartones,
            participaciones_existentes=(41,),
        )

        self.assertEqual(prueba["crear_participacion"].call_count, 2)
        ids_creados = {
            llamada.kwargs["idcarton"].pk
            for llamada in prueba["crear_participacion"].call_args_list
        }
        self.assertEqual(ids_creados, {40, 42})

    def test_fallo_de_participacion_sale_de_atomic_y_propaga_el_error(self):
        prueba = self._crear_sin_base(
            [self._carton(40)],
            error_participacion=IntegrityError("fallo de participación"),
            capturar_error=True,
        )

        self.assertIsInstance(prueba["error"], IntegrityError)
        self.assertEqual(str(prueba["error"]), "fallo de participación")
        self.assertIs(
            prueba["atomic_context"].__exit__.call_args.args[0],
            IntegrityError,
        )
        prueba["guardar_partida"].assert_called_once_with(force_insert=True)

    def test_crear_ronda_no_modifica_precio_pago_premio_ni_cartones(self):
        cartones = [
            self._carton(40, precio="4.50"),
            self._carton(41, precio="6.25"),
        ]
        precios_antes = [carton.preciopagado for carton in cartones]
        premio_antes = self.partida.valorefectivo

        prueba = self._crear_sin_base(cartones)

        self.assertEqual(
            [carton.preciopagado for carton in cartones],
            precios_antes,
        )
        self.assertEqual(self.partida.valorefectivo, premio_antes)
        prueba["guardar_carton"].assert_not_called()

    def test_rechaza_una_ronda_ya_registrada_antes_de_abrir_transaccion(self):
        self.partida.idpartidabingo = 30

        with (
            patch("apps.bingos.services.transaction.atomic") as atomic_mock,
            self.assertRaisesMessage(
                RondaCreacionError,
                "ya fue registrada",
            ),
        ):
            crear_ronda_con_participaciones(self.bingo, self.partida)

        atomic_mock.assert_not_called()


class CreacionRondaVistaTests(SimpleTestCase):
    def setUp(self):
        self.factory = RequestFactory()
        self.bingo = Bingo(idbingo=7, titulobingo="Bingo híbrido")

    def _request(self, data=None):
        request = self.factory.post(
            reverse(
                "bingos:partida_nueva",
                kwargs={"idbingo": self.bingo.pk},
            ),
            data or datos_partida_form(),
        )
        request.user = User(username="admin", is_staff=True)
        request.session = {}
        request._messages = FallbackStorage(request)
        return request

    def test_ruta_de_creacion_usa_servicio_atomico_y_muestra_resultado(self):
        request = self._request()

        def resultado_servicio(*, bingo, partida):
            self.assertEqual(bingo, self.bingo)
            self.assertIsInstance(partida, Partidabingo)
            partida.idpartidabingo = 30
            return {
                "partida": partida,
                "participaciones_creadas": [Mock(), Mock(), Mock()],
            }

        with (
            patch(
                "apps.bingos.views.get_object_or_404",
                return_value=self.bingo,
            ),
            patch(
                "apps.bingos.views.crear_ronda_con_participaciones",
                side_effect=resultado_servicio,
            ) as crear_ronda_mock,
        ):
            response = partida_nueva(request, self.bingo.pk)

        crear_ronda_mock.assert_called_once()
        self.assertEqual(response.status_code, 302)
        self.assertEqual(
            response["Location"],
            reverse(
                "bingos:partida_detalle",
                kwargs={"idpartidabingo": 30},
            ),
        )
        self.assertIn(
            "La ronda fue creada correctamente y se registraron las "
            "participaciones de los cartones activos del Bingo.",
            [message.message for message in get_messages(request)],
        )


class GenerarAsignarCartonFormTests(SimpleTestCase):
    def test_formulario_solo_expone_jugador_y_precio(self):
        form = GenerarAsignarCartonForm()

        self.assertEqual(list(form.fields), ["idjugador", "preciopagado"])

    def test_jugador_y_precio_mayor_que_cero_son_obligatorios(self):
        form = GenerarAsignarCartonForm(
            data={"idjugador": "", "preciopagado": "0"}
        )

        self.assertFalse(form.is_valid())
        self.assertIn("idjugador", form.errors)
        self.assertIn("preciopagado", form.errors)


class GastoOperativoBingoFormTests(SimpleTestCase):
    def test_formulario_gasto_solo_expone_campos_permitidos(self):
        form = GastoOperativoBingoForm()

        self.assertEqual(
            list(form.fields),
            ["concepto", "monto", "fechagasto", "observacion"],
        )
        for campo_no_permitido in (
            "estado",
            "idbingo",
            "idusuarioregistro",
            "motivoanulacion",
        ):
            self.assertNotIn(campo_no_permitido, form.fields)

    def test_formulario_anulacion_solo_expone_motivo_obligatorio(self):
        form = MotivoAnulacionForm(data={"motivo": ""})

        self.assertEqual(list(form.fields), ["motivo"])
        self.assertFalse(form.is_valid())
        self.assertIn("motivo", form.errors)

    def test_formulario_costo_solo_expone_campos_permitidos(self):
        form = CostoPremioMaterialBingoForm()

        self.assertEqual(
            list(form.fields),
            ["partida", "descripcion_premio", "monto", "observacion"],
        )
        for campo_no_permitido in (
            "idbingo",
            "idusuarioregistro",
            "estado",
            "cierre",
            "motivoanulacion",
        ):
            self.assertNotIn(campo_no_permitido, form.fields)

    def test_formulario_cierre_solo_expone_observacion_opcional(self):
        form = CierreFinancieroBingoForm(data={"observacion_cierre": ""})

        self.assertEqual(list(form.fields), ["observacion_cierre"])
        self.assertTrue(form.is_valid())
        for campo_no_permitido in (
            "estado",
            "idbingo",
            "idusuariocierre",
            "fechacierre",
            "recaudacionregistrada",
            "utilidadneta",
            "cartonesvendidosunicos",
        ):
            self.assertNotIn(campo_no_permitido, form.fields)


class RutaHeredadaCartonTests(SimpleTestCase):
    def setUp(self):
        self.factory = RequestFactory()
        self.bingo = Bingo(
            idbingo=2,
            titulobingo="Bingo de prueba",
            preciocarton=Decimal("5.00"),
        )
        self.partida = Partidabingo(
            idpartidabingo=1,
            idbingo=self.bingo,
            estadopartida=ESTADO_PARTIDA_PROGRAMADA,
            nombreronda="Ronda 1",
        )
        self.url = reverse(
            "bingos:partida_carton_nuevo",
            kwargs={"idpartidabingo": self.partida.pk},
        )

    def _request(self, method="get", data=None, usuario=None):
        request_method = getattr(self.factory, method)
        request = request_method(self.url, data or {})
        request.user = usuario or User(username="admin", is_staff=True)
        request.session = {}
        request._messages = FallbackStorage(request)
        return request

    def test_get_administrador_redirige_al_detalle_del_bingo_asociado(self):
        request = self._request()

        with patch(
            "apps.bingos.views.get_object_or_404",
            return_value=self.partida,
        ):
            response = partida_carton_nuevo(
                request,
                self.partida.idpartidabingo,
            )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(
            response["Location"],
            reverse("bingos:detalle", kwargs={"idbingo": self.bingo.pk}),
        )
        mensajes = list(get_messages(request))
        self.assertEqual(len(mensajes), 1)
        self.assertEqual(mensajes[0].tags, "info")
        self.assertIn("se crean para todo el Bingo", mensajes[0].message)

    def test_post_no_crea_ni_modifica_cartones_o_participaciones(self):
        request = self._request(
            method="post",
            data={"idjugador": "4", "preciopagado": "5.00"},
        )

        with (
            patch(
                "apps.bingos.views.get_object_or_404",
                return_value=self.partida,
            ),
            patch(
                "apps.bingos.views.GenerarAsignarCartonForm"
            ) as formulario_legado,
            patch("apps.bingos.views.crear_y_asignar_carton") as crear_legado,
            patch(
                "apps.bingos.views.crear_carton_maestro_para_bingo"
            ) as crear_maestro,
            patch("apps.bingos.views.Carton.objects.create") as crear_carton,
            patch("apps.bingos.views.Carton.objects.update") as modificar_carton,
            patch(
                "apps.bingos.views.CartonPartidaBingo.objects.create"
            ) as crear_participacion,
            patch(
                "apps.bingos.views.CartonPartidaBingo.objects.update"
            ) as modificar_participacion,
        ):
            response = partida_carton_nuevo(
                request,
                self.partida.idpartidabingo,
            )

        self.assertRedirects(
            response,
            reverse("bingos:detalle", kwargs={"idbingo": self.bingo.pk}),
            fetch_redirect_response=False,
        )
        formulario_legado.assert_not_called()
        crear_legado.assert_not_called()
        crear_maestro.assert_not_called()
        crear_carton.assert_not_called()
        modificar_carton.assert_not_called()
        crear_participacion.assert_not_called()
        modificar_participacion.assert_not_called()

    def test_usuarios_sin_permiso_conservan_la_proteccion(self):
        get_object_mock = Mock()

        request_anonimo = self._request(usuario=AnonymousUser())
        with patch(
            "apps.bingos.views.get_object_or_404",
            get_object_mock,
        ):
            response = partida_carton_nuevo(
                request_anonimo,
                self.partida.idpartidabingo,
            )

        self.assertEqual(response.status_code, 302)
        self.assertIn("/login/", response["Location"])
        get_object_mock.assert_not_called()

        request_no_admin = self._request(
            usuario=User(
                username="jugador",
                is_staff=False,
                is_superuser=False,
            )
        )
        with (
            patch(
                "apps.bingos.views.get_object_or_404",
                get_object_mock,
            ),
            self.assertRaises(PermissionDenied),
        ):
            partida_carton_nuevo(
                request_no_admin,
                self.partida.idpartidabingo,
            )

        get_object_mock.assert_not_called()


@override_settings(
    STORAGES={
        "default": {
            "BACKEND": "django.core.files.storage.FileSystemStorage",
        },
        "staticfiles": {
            "BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage",
        },
    }
)
class RutasGenericasCartonBloqueadasTests(SimpleTestCase):
    def setUp(self):
        self.factory = RequestFactory()
        self.usuario_admin = User(username="admin", is_staff=True)
        self.bingo = Bingo(
            idbingo=7,
            titulobingo="Bingo seguro",
            preciocarton=Decimal("5.00"),
        )
        self.partida = Partidabingo(
            idpartidabingo=21,
            idbingo=self.bingo,
            nombreronda="Ronda segura",
            valorefectivo=Decimal("25.00"),
            premiomaterial="Canasta",
            estadopartida=ESTADO_PARTIDA_PROGRAMADA,
            patronganador="carton_lleno",
            bolascantadas="[]",
            ultimabola=0,
            haydesempate=False,
            horainicio=timezone.now(),
        )
        self.carton = Carton(
            idcarton=31,
            idbingo=self.bingo,
            idjugador=Jugador(idjugador=4, aliasjugador="jugador4"),
            idpartida=self.partida,
            codigocarton="CARTON-BLOQUEADO",
            matriznumeros=serializar_matriz_carton_bingo(matriz_carton_prueba()),
            indicevictoria=0,
            preciopagado=Decimal("5.00"),
            fechacompra=timezone.now(),
            estadocarton="Vendido",
        )

    def _request(self, path, method="get", data=None):
        request = getattr(self.factory, method)(path, data or {})
        request.user = self.usuario_admin
        request.session = {}
        request._messages = FallbackStorage(request)
        return request

    def test_get_carton_nuevo_redirige_sin_mostrar_formulario(self):
        request = self._request(reverse("bingos:carton_nuevo"))

        with (
            patch("apps.bingos.views.CartonForm") as form_mock,
            patch("apps.bingos.views.render") as render_mock,
        ):
            response = carton_nuevo(request)

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response["Location"], reverse("bingos:cartones_lista"))
        form_mock.assert_not_called()
        render_mock.assert_not_called()
        self.assertIn(
            "venta por Bingo",
            " ".join(mensaje.message for mensaje in get_messages(request)),
        )

    def test_post_carton_nuevo_no_crea_carton(self):
        request = self._request(
            reverse("bingos:carton_nuevo"),
            method="post",
            data={"codigocarton": "MANUAL", "preciopagado": "1.00"},
        )

        with (
            patch("apps.bingos.views.CartonForm") as form_mock,
            patch("apps.bingos.views.save_new_model_form") as save_mock,
            patch("apps.bingos.views.Carton.objects.create") as create_mock,
        ):
            response = carton_nuevo(request)

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response["Location"], reverse("bingos:cartones_lista"))
        form_mock.assert_not_called()
        save_mock.assert_not_called()
        create_mock.assert_not_called()

    def test_get_carton_editar_redirige_sin_mostrar_formulario(self):
        request = self._request(
            reverse("bingos:carton_editar", kwargs={"idcarton": self.carton.pk})
        )

        with (
            patch("apps.bingos.views.get_object_or_404") as get_mock,
            patch("apps.bingos.views.CartonForm") as form_mock,
            patch("apps.bingos.views.render") as render_mock,
        ):
            response = carton_editar(request, self.carton.pk)

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response["Location"], reverse("bingos:cartones_lista"))
        get_mock.assert_not_called()
        form_mock.assert_not_called()
        render_mock.assert_not_called()

    def test_post_carton_editar_no_modifica_carton(self):
        request = self._request(
            reverse("bingos:carton_editar", kwargs={"idcarton": self.carton.pk}),
            method="post",
            data={"preciopagado": "999.00", "estadocarton": "Disponible"},
        )

        with (
            patch("apps.bingos.views.get_object_or_404") as get_mock,
            patch("apps.bingos.views.CartonForm") as form_mock,
            patch("apps.bingos.views.transaction.atomic") as atomic_mock,
            patch.object(Carton, "save") as save_mock,
        ):
            response = carton_editar(request, self.carton.pk)

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response["Location"], reverse("bingos:cartones_lista"))
        get_mock.assert_not_called()
        form_mock.assert_not_called()
        atomic_mock.assert_not_called()
        save_mock.assert_not_called()

    def test_get_partida_carton_editar_redirige_sin_mostrar_formulario(self):
        request = self._request(
            reverse(
                "bingos:partida_carton_editar",
                kwargs={
                    "idpartidabingo": self.partida.pk,
                    "idcarton": self.carton.pk,
                },
            )
        )

        with (
            patch(
                "apps.bingos.views.get_object_or_404",
                return_value=self.partida,
            ),
            patch("apps.bingos.views.CartonPartidaForm") as form_mock,
            patch("apps.bingos.views.render") as render_mock,
        ):
            response = partida_carton_editar(
                request,
                self.partida.pk,
                self.carton.pk,
            )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(
            response["Location"],
            reverse(
                "bingos:partida_detalle",
                kwargs={"idpartidabingo": self.partida.pk},
            ),
        )
        form_mock.assert_not_called()
        render_mock.assert_not_called()

    def test_post_partida_carton_editar_no_modifica_carton(self):
        request = self._request(
            reverse(
                "bingos:partida_carton_editar",
                kwargs={
                    "idpartidabingo": self.partida.pk,
                    "idcarton": self.carton.pk,
                },
            ),
            method="post",
            data={"preciopagado": "999.00", "estadocarton": "Disponible"},
        )

        with (
            patch(
                "apps.bingos.views.get_object_or_404",
                return_value=self.partida,
            ),
            patch("apps.bingos.views.CartonPartidaForm") as form_mock,
            patch("apps.bingos.views.transaction.atomic") as atomic_mock,
            patch.object(Carton, "save") as save_mock,
        ):
            response = partida_carton_editar(
                request,
                self.partida.pk,
                self.carton.pk,
            )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(
            response["Location"],
            reverse(
                "bingos:partida_detalle",
                kwargs={"idpartidabingo": self.partida.pk},
            ),
        )
        form_mock.assert_not_called()
        atomic_mock.assert_not_called()
        save_mock.assert_not_called()

    def test_cartones_lista_no_enlaza_creacion_ni_edicion_generica(self):
        html = render_to_string(
            "bingos/cartones_lista.html",
            {"page_obj": [self.carton], "busqueda": "", "total": 1},
            request=self._request(reverse("bingos:cartones_lista")),
        )

        self.assertIn("deben venderse desde el detalle del Bingo", html)
        self.assertNotIn(reverse("bingos:carton_nuevo"), html)
        self.assertNotIn(
            reverse("bingos:carton_editar", kwargs={"idcarton": self.carton.pk}),
            html,
        )
        self.assertNotIn("+ Nuevo cartón", html)

    def test_partida_detalle_no_enlaza_edicion_directa_por_partida(self):
        html = render_to_string(
            "bingos/partida_detalle.html",
            {
                "partida": self.partida,
                "cartones": [self.carton],
                "participaciones_hibridas": [],
                "bolas_cantadas": [],
            },
            request=self._request(
                reverse(
                    "bingos:partida_detalle",
                    kwargs={"idpartidabingo": self.partida.pk},
                )
            ),
        )

        self.assertNotIn(
            reverse(
                "bingos:partida_carton_editar",
                kwargs={
                    "idpartidabingo": self.partida.pk,
                    "idcarton": self.carton.pk,
                },
            ),
            html,
        )

    def test_consola_operador_no_enlaza_edicion_directa_por_partida(self):
        html = render_to_string(
            "bingos/consola_operador.html",
            {
                "partida": self.partida,
                "acciones_consola": [],
                "actualizacion_estados_pendiente": False,
                "cartones": [self.carton],
                "participaciones_hibridas": [],
                "participaciones_hibridas_validacion": [],
                "error_participaciones_hibridas": None,
                "puede_asignar_cartones": False,
                "puede_validar_cartones": False,
                "cartones_validacion": [],
                "candidatos_desempate": [],
                "ultima_bola_codigo": None,
                "total_bolas_extraidas": 0,
                "total_bolas_faltantes": 75,
                "hay_bolas_disponibles": True,
                "puede_sacar_bola": False,
                "historial_bolas": [],
                "tablero_bingo": [],
            },
            request=self._request(
                reverse(
                    "bingos:consola_operador",
                    kwargs={"idpartidabingo": self.partida.pk},
                )
            ),
        )

        self.assertNotIn(
            reverse(
                "bingos:partida_carton_editar",
                kwargs={
                    "idpartidabingo": self.partida.pk,
                    "idcarton": self.carton.pk,
                },
            ),
            html,
        )


@override_settings(
    STORAGES={
        "default": {
            "BACKEND": "django.core.files.storage.FileSystemStorage",
        },
        "staticfiles": {
            "BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage",
        },
    }
)
class VentaCartonBingoInterfazTests(SimpleTestCase):
    def setUp(self):
        self.factory = RequestFactory()
        self.usuario_admin = User(username="admin", is_staff=True)
        self.jugador = Jugador(idjugador=4, aliasjugador="jugador4")
        self.bingo = Bingo(
            idbingo=7,
            titulobingo="Bingo híbrido administrativo",
            preciocarton=Decimal("5.00"),
        )

    def _request(self, method="get", data=None, usuario=None, path=None):
        path = path or reverse(
            "bingos:bingo_carton_nuevo",
            kwargs={"idbingo": self.bingo.pk},
        )
        request = getattr(self.factory, method)(path, data or {})
        request.user = usuario or self.usuario_admin
        request.session = {}
        request._messages = FallbackStorage(request)
        return request

    def _form_valido(self, data=None):
        class FormPrueba(forms.Form):
            idjugador = forms.IntegerField()
            preciopagado = forms.DecimalField()

        form = FormPrueba(
            data=data
            or {"idjugador": self.jugador.pk, "preciopagado": "5.00"}
        )
        form.is_valid()
        form.cleaned_data = {
            "idjugador": self.jugador,
            "preciopagado": Decimal("5.00"),
        }
        form.is_valid = Mock(return_value=True)
        return form

    def test_formulario_hibrido_solo_expone_jugador_y_precio(self):
        form = GenerarCartonBingoForm()

        self.assertEqual(list(form.fields), ["idjugador", "preciopagado"])
        for campo_historico in (
            "idpartida",
            "indicevictoria",
            "codigocarton",
            "matriznumeros",
            "estadocarton",
            "idbingo",
        ):
            self.assertNotIn(campo_historico, form.fields)

    def test_formulario_rechaza_precio_cero_y_negativo(self):
        for precio in ("0", "-1"):
            with self.subTest(precio=precio):
                form = GenerarCartonBingoForm(
                    data={"idjugador": "", "preciopagado": precio}
                )
                self.assertFalse(form.is_valid())
                self.assertIn("preciopagado", form.errors)
                self.assertIn(
                    "mayor que cero",
                    " ".join(form.errors["preciopagado"]),
                )

    def test_ruta_nueva_exige_login_y_permiso_administrativo(self):
        request_anonimo = self._request(usuario=AnonymousUser())
        response = bingo_carton_nuevo(request_anonimo, self.bingo.pk)

        self.assertEqual(response.status_code, 302)
        self.assertIn("login", response["Location"])

        request_no_admin = self._request(
            usuario=User(username="jugador", is_staff=False)
        )
        with self.assertRaises(PermissionDenied):
            bingo_carton_nuevo(request_no_admin, self.bingo.pk)

    def test_get_informa_rondas_incluidas_y_excluidas(self):
        request = self._request()
        partidas = [
            Partidabingo(
                idpartidabingo=21,
                idbingo=self.bingo,
                estadopartida=ESTADO_PARTIDA_FINALIZADA,
            ),
            Partidabingo(
                idpartidabingo=22,
                idbingo=self.bingo,
                estadopartida=ESTADO_PARTIDA_EN_CURSO,
            ),
            Partidabingo(
                idpartidabingo=23,
                idbingo=self.bingo,
                estadopartida=ESTADO_PARTIDA_PROGRAMADA,
            ),
            Partidabingo(
                idpartidabingo=24,
                idbingo=self.bingo,
                estadopartida=ESTADO_PARTIDA_EN_ESPERA,
            ),
        ]
        with (
            patch(
                "apps.bingos.views.get_object_or_404",
                return_value=self.bingo,
            ),
            patch(
                "apps.bingos.views.Partidabingo.objects.filter"
            ) as partidas_filter,
            patch(
                "apps.bingos.forms.Jugador.objects.order_by",
                return_value=Jugador.objects.none(),
            ),
        ):
            partidas_filter.return_value = partidas
            response = bingo_carton_nuevo(request, self.bingo.pk)

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Vender cartón para todo el Bingo")
        self.assertContains(response, self.bingo.titulobingo)
        self.assertContains(response, "participará en 2 rondas futuras")
        self.assertContains(response, "Las 2 rondas que ya están en curso")
        self.assertContains(response, "Rondas futuras incluidas")
        self.assertContains(response, "Rondas no incluidas")

    def test_get_sin_elegibles_informa_que_no_quedan_rondas_futuras(self):
        request = self._request()
        partidas = [
            Partidabingo(
                idpartidabingo=21,
                idbingo=self.bingo,
                estadopartida=ESTADO_PARTIDA_FINALIZADA,
            )
        ]
        with (
            patch(
                "apps.bingos.views.get_object_or_404",
                return_value=self.bingo,
            ),
            patch(
                "apps.bingos.views.Partidabingo.objects.filter",
                return_value=partidas,
            ),
            patch(
                "apps.bingos.forms.Jugador.objects.order_by",
                return_value=Jugador.objects.none(),
            ),
        ):
            response = bingo_carton_nuevo(request, self.bingo.pk)

        self.assertContains(response, "No quedan rondas futuras disponibles")
        self.assertContains(response, "disabled")

    def test_post_valido_llama_servicio_redirige_y_muestra_resumen(self):
        request = self._request(
            method="post",
            data={"idjugador": "4", "preciopagado": "5.00"},
        )
        form = self._form_valido(request.POST)
        carton = Carton(
            idcarton=40,
            idbingo=self.bingo,
            idjugador=self.jugador,
            codigocarton="B7-C-PRUEBA",
        )
        with (
            patch(
                "apps.bingos.views.get_object_or_404",
                return_value=self.bingo,
            ),
            patch(
                "apps.bingos.views.GenerarCartonBingoForm",
                return_value=form,
            ),
            patch(
                "apps.bingos.views.crear_carton_maestro_para_bingo",
                return_value=carton,
            ) as crear_mock,
            patch(
                "apps.bingos.views.CartonPartidaBingo.objects.filter"
            ) as participaciones_filter,
        ):
            participaciones_filter.return_value.count.return_value = 3
            response = bingo_carton_nuevo(request, self.bingo.pk)

        crear_mock.assert_called_once_with(
            bingo=self.bingo,
            jugador=self.jugador,
            precio_pagado=Decimal("5.00"),
            fecha_compra=None,
        )
        self.assertEqual(response.status_code, 302)
        self.assertEqual(
            response["Location"],
            reverse("bingos:detalle", kwargs={"idbingo": self.bingo.pk}),
        )
        mensajes = [mensaje.message for mensaje in get_messages(request)]
        self.assertTrue(
            any(
                "cartón maestro" in mensaje
                and "3 participaciones" in mensaje
                and "rondas futuras elegibles" in mensaje
                for mensaje in mensajes
            )
        )

    def test_error_del_servicio_vuelve_al_formulario_sin_crear_directamente(self):
        request = self._request(
            method="post",
            data={"idjugador": "4", "preciopagado": "5.00"},
        )
        form = self._form_valido(request.POST)
        error = "No quedan rondas futuras disponibles."
        with (
            patch(
                "apps.bingos.views.get_object_or_404",
                return_value=self.bingo,
            ),
            patch(
                "apps.bingos.views.GenerarCartonBingoForm",
                return_value=form,
            ),
            patch(
                "apps.bingos.views.crear_carton_maestro_para_bingo",
                side_effect=CartonAsignacionError(error),
            ),
            patch(
                "apps.bingos.views.Partidabingo.objects.filter"
            ) as partidas_filter,
            patch(
                "apps.bingos.views.Carton.objects.create"
            ) as carton_create,
            patch(
                "apps.bingos.views.CartonPartidaBingo.objects.create"
            ) as participacion_create,
            patch(
                "apps.bingos.views.render",
                return_value=HttpResponse(status=200),
            ),
        ):
            partidas_filter.return_value = []
            response = bingo_carton_nuevo(request, self.bingo.pk)

        self.assertEqual(response.status_code, 200)
        self.assertIn(error, form.non_field_errors())
        self.assertEqual(form.data["idjugador"], "4")
        self.assertEqual(form.data["preciopagado"], "5.00")
        carton_create.assert_not_called()
        participacion_create.assert_not_called()

    def test_cierre_financiero_en_venta_admin_redirige_con_mensaje(self):
        request = self._request(
            method="post",
            data={"idjugador": "4", "preciopagado": "5.00"},
        )
        form = self._form_valido(request.POST)
        mensaje = (
            "No se pueden vender cartones porque el cierre financiero está cerrado."
        )
        with (
            patch(
                "apps.bingos.views.get_object_or_404",
                return_value=self.bingo,
            ),
            patch(
                "apps.bingos.views.GenerarCartonBingoForm",
                return_value=form,
            ),
            patch(
                "apps.bingos.views.crear_carton_maestro_para_bingo",
                side_effect=CierreFinancieroError(mensaje),
            ) as crear_mock,
            patch("apps.bingos.views.Carton.objects.create") as carton_create,
            patch(
                "apps.bingos.views.CartonPartidaBingo.objects.create"
            ) as participacion_create,
        ):
            response = bingo_carton_nuevo(request, self.bingo.pk)

        crear_mock.assert_called_once_with(
            bingo=self.bingo,
            jugador=self.jugador,
            precio_pagado=Decimal("5.00"),
            fecha_compra=None,
        )
        self.assertEqual(response.status_code, 302)
        self.assertEqual(
            response["Location"],
            reverse("bingos:detalle", kwargs={"idbingo": self.bingo.pk}),
        )
        self.assertIn(mensaje, [mensaje.message for mensaje in get_messages(request)])
        carton_create.assert_not_called()
        participacion_create.assert_not_called()

    def test_ruta_heredada_sigue_resolviendo(self):
        partida = Partidabingo(
            idpartidabingo=21,
            idbingo=self.bingo,
            nombreronda="Ronda heredada",
            estadopartida=ESTADO_PARTIDA_PROGRAMADA,
        )
        url = reverse(
            "bingos:partida_carton_nuevo",
            kwargs={"idpartidabingo": partida.pk},
        )
        self.assertEqual(resolve(url).url_name, "partida_carton_nuevo")

    def test_detalle_administrativo_enlaza_venta_por_bingo(self):
        request = self._request()
        html = render_to_string(
            "bingos/detalle.html",
            {"bingo": self.bingo, "partidas": [], "cartones": []},
            request=request,
        )
        url = reverse(
            "bingos:bingo_carton_nuevo",
            kwargs={"idbingo": self.bingo.pk},
        )

        self.assertIn("Vender cartón para todo el Bingo", html)
        self.assertIn(url, html)
        self.assertIn("se cobra", html)


@override_settings(
    STORAGES={
        "default": {
            "BACKEND": "django.core.files.storage.FileSystemStorage",
        },
        "staticfiles": {
            "BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage",
        },
    }
)
class CompraDirectaCartonJugadorTests(SimpleTestCase):
    def setUp(self):
        self.factory = RequestFactory()
        self.jugador = Jugador(
            idjugador=4,
            aliasjugador="jugador4",
            estadocuentajugador="Activo",
        )
        self.otro_jugador = Jugador(
            idjugador=99,
            aliasjugador="otro",
            estadocuentajugador="Activo",
        )
        self.bingo = Bingo(
            idbingo=7,
            titulobingo="Bingo compra directa",
            preciocarton=Decimal("8.50"),
            estadobingo="Programado",
        )
        self.partidas = [
            self._partida(21, ESTADO_PARTIDA_PROGRAMADA),
            self._partida(22, ESTADO_PARTIDA_EN_ESPERA),
            self._partida(23, ESTADO_PARTIDA_EN_CURSO),
        ]

    def _partida(self, pk, estado):
        return Partidabingo(
            idpartidabingo=pk,
            idbingo=self.bingo,
            nombreronda=f"Ronda {pk}",
            estadopartida=estado,
            horainicio=timezone.now() + timedelta(minutes=pk),
        )

    def _request(self, method="get", data=None, usuario=None):
        path = reverse(
            "bingos:comprar_carton_bingo",
            kwargs={"idbingo": self.bingo.pk},
        )
        request = getattr(self.factory, method)(path, data or {})
        request.user = usuario or User(
            username=self.jugador.aliasjugador,
            is_staff=False,
            is_superuser=False,
        )
        request.session = {}
        request._messages = FallbackStorage(request)
        return request

    def _autorizar_jugador(self):
        return patch(
            "apps.jugadores.services.obtener_jugador_autenticado",
            return_value=self.jugador,
        )

    def test_jugador_puede_abrir_la_compra(self):
        request = self._request()

        with (
            self._autorizar_jugador(),
            patch(
                "apps.bingos.views.get_object_or_404",
                return_value=self.bingo,
            ),
            patch(
                "apps.bingos.views.Partidabingo.objects.filter",
                return_value=self.partidas,
            ),
        ):
            response = comprar_carton_bingo(request, self.bingo.pk)

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Comprar cartón")
        self.assertContains(response, self.bingo.titulobingo)
        self.assertContains(response, "$8,50")
        self.assertContains(response, "Rondas donde participará")
        self.assertContains(
            response,
            "El cartón participa únicamente en las rondas elegibles",
        )

    def test_anonimo_es_redirigido_al_login(self):
        request = self._request(usuario=AnonymousUser())

        with patch("apps.bingos.views.get_object_or_404") as obtener_mock:
            response = comprar_carton_bingo(request, self.bingo.pk)

        self.assertEqual(response.status_code, 302)
        self.assertIn("/login/", response["Location"])
        obtener_mock.assert_not_called()

    def test_usuario_sin_perfil_jugador_no_puede_comprar(self):
        request = self._request(usuario=User(username="sin-perfil"))

        with self.assertRaises(PermissionDenied):
            comprar_carton_bingo(request, self.bingo.pk)

    def test_formulario_no_expone_jugador_precio_ni_pago(self):
        form = CompraCartonJugadorForm()

        self.assertEqual(list(form.fields), ["confirmar_compra"])
        self.assertNotIn("idjugador", form.fields)
        self.assertNotIn("preciopagado", form.fields)
        self.assertNotIn("idmetodopago", form.fields)
        self.assertNotIn("comprobantepago", form.fields)
        self.assertNotIn("numeroreferencia", form.fields)

    def test_post_usa_jugador_autenticado_y_precio_oficial_aunque_manipulen_datos(self):
        request = self._request(
            method="post",
            data={
                "confirmar_compra": "on",
                "idjugador": str(self.otro_jugador.pk),
                "preciopagado": "0.01",
            },
        )
        carton = Carton(
            idcarton=40,
            idbingo=self.bingo,
            idjugador=self.jugador,
            idpartida=None,
            codigocarton="B7-C-DIRECTA",
            preciopagado=self.bingo.preciocarton,
            estadocarton="Vendido",
        )

        with (
            self._autorizar_jugador(),
            patch(
                "apps.bingos.views.get_object_or_404",
                return_value=self.bingo,
            ),
            patch(
                "apps.bingos.views.Partidabingo.objects.filter",
                return_value=self.partidas,
            ),
            patch(
                "apps.bingos.views.crear_carton_maestro_para_bingo",
                return_value=carton,
            ) as crear_mock,
        ):
            response = comprar_carton_bingo(request, self.bingo.pk)

        crear_mock.assert_called_once_with(
            bingo=self.bingo,
            jugador=self.jugador,
            precio_pagado=self.bingo.preciocarton,
            fecha_compra=None,
        )
        self.assertEqual(response.status_code, 302)
        self.assertEqual(response["Location"], reverse("bingos:mis_cartones"))
        mensajes = [mensaje.message for mensaje in get_messages(request)]
        self.assertIn("Compra completada. Tu cartón es B7-C-DIRECTA.", mensajes)

    def test_compra_despues_de_cierre_redirige_con_mensaje_sin_crear(self):
        request = self._request(
            method="post",
            data={"confirmar_compra": "on"},
        )
        mensaje = (
            "No se pueden vender cartones porque el cierre financiero está cerrado."
        )

        with (
            self._autorizar_jugador(),
            patch(
                "apps.bingos.views.get_object_or_404",
                return_value=self.bingo,
            ),
            patch(
                "apps.bingos.views.Partidabingo.objects.filter",
                return_value=self.partidas,
            ),
            patch(
                "apps.bingos.views.crear_carton_maestro_para_bingo",
                side_effect=CierreFinancieroError(mensaje),
            ) as crear_mock,
            patch("apps.bingos.views.Carton.objects.create") as carton_create,
            patch(
                "apps.bingos.views.CartonPartidaBingo.objects.create"
            ) as participacion_create,
        ):
            response = comprar_carton_bingo(request, self.bingo.pk)

        crear_mock.assert_called_once_with(
            bingo=self.bingo,
            jugador=self.jugador,
            precio_pagado=self.bingo.preciocarton,
            fecha_compra=None,
        )
        self.assertEqual(response.status_code, 302)
        self.assertEqual(
            response["Location"],
            reverse("bingos:comprar_carton_bingo", kwargs={"idbingo": self.bingo.pk}),
        )
        self.assertIn(mensaje, [mensaje.message for mensaje in get_messages(request)])
        carton_create.assert_not_called()
        participacion_create.assert_not_called()

    def test_no_crea_carton_sin_rondas_elegibles(self):
        request = self._request(
            method="post",
            data={"confirmar_compra": "on"},
        )
        partidas = [
            self._partida(31, ESTADO_PARTIDA_FINALIZADA),
            self._partida(32, ESTADO_PARTIDA_EN_CURSO),
        ]

        with (
            self._autorizar_jugador(),
            patch(
                "apps.bingos.views.get_object_or_404",
                return_value=self.bingo,
            ),
            patch(
                "apps.bingos.views.Partidabingo.objects.filter",
                return_value=partidas,
            ),
            patch(
                "apps.bingos.views.crear_carton_maestro_para_bingo"
            ) as crear_mock,
        ):
            response = comprar_carton_bingo(request, self.bingo.pk)

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "ya no quedan rondas futuras disponibles")
        crear_mock.assert_not_called()

    def test_no_permite_comprar_en_bingo_finalizado_o_cancelado(self):
        for estado in ("Finalizado", "Cancelado"):
            with self.subTest(estado=estado):
                bingo = Bingo(
                    idbingo=7,
                    titulobingo="Bingo cerrado",
                    preciocarton=Decimal("8.50"),
                    estadobingo=estado,
                )
                request = self._request(
                    method="post",
                    data={"confirmar_compra": "on"},
                )

                with (
                    self._autorizar_jugador(),
                    patch(
                        "apps.bingos.views.get_object_or_404",
                        return_value=bingo,
                    ),
                    patch(
                        "apps.bingos.views.Partidabingo.objects.filter",
                        return_value=self.partidas,
                    ),
                    patch(
                        "apps.bingos.views.crear_carton_maestro_para_bingo"
                    ) as crear_mock,
                ):
                    response = comprar_carton_bingo(request, bingo.pk)

                self.assertEqual(response.status_code, 200)
                self.assertContains(response, "Bingo no admite compras")
                crear_mock.assert_not_called()

    def test_jugador_puede_comprar_varios_cartones_del_mismo_bingo(self):
        cartones = [
            Carton(
                idcarton=40,
                idbingo=self.bingo,
                idjugador=self.jugador,
                codigocarton="B7-C-UNO",
            ),
            Carton(
                idcarton=41,
                idbingo=self.bingo,
                idjugador=self.jugador,
                codigocarton="B7-C-DOS",
            ),
        ]

        with (
            self._autorizar_jugador(),
            patch(
                "apps.bingos.views.get_object_or_404",
                return_value=self.bingo,
            ),
            patch(
                "apps.bingos.views.Partidabingo.objects.filter",
                return_value=self.partidas,
            ),
            patch(
                "apps.bingos.views.crear_carton_maestro_para_bingo",
                side_effect=cartones,
            ) as crear_mock,
        ):
            primera = comprar_carton_bingo(
                self._request(method="post", data={"confirmar_compra": "on"}),
                self.bingo.pk,
            )
            segunda = comprar_carton_bingo(
                self._request(method="post", data={"confirmar_compra": "on"}),
                self.bingo.pk,
            )

        self.assertEqual(primera.status_code, 302)
        self.assertEqual(segunda.status_code, 302)
        self.assertEqual(crear_mock.call_count, 2)

    def test_carton_comprado_aparece_en_mis_cartones(self):
        carton = Carton(
            idcarton=40,
            idbingo=self.bingo,
            idjugador=self.jugador,
            idpartida=None,
            codigocarton="B7-C-DIRECTA",
            matriznumeros=serializar_matriz_carton_bingo(
                matriz_carton_prueba()
            ),
            estadocarton="Vendido",
        )
        carton.participaciones_hibridas_cargadas = [
            CartonPartidaBingo(
                idcartonpartidabingo=50,
                idcarton=carton,
                idpartida=self.partidas[0],
                idbingo=self.bingo,
                estado_participacion=CartonPartidaBingo.ESTADO_PENDIENTE,
            )
        ]
        capturado = {}

        def fake_render(_request, _template, contexto):
            capturado.update(contexto)
            return HttpResponse(status=200)

        with (
            self._autorizar_jugador(),
            patch(
                "apps.bingos.views.Carton.objects.filter",
                return_value=_CartonesConsultaQuerySet([carton]),
            ),
            patch("apps.bingos.views.render", side_effect=fake_render),
        ):
            response = mis_cartones(self._request())

        self.assertEqual(response.status_code, 200)
        resumenes = capturado["cartones_resumen"]
        self.assertEqual(len(resumenes), 1)
        self.assertEqual(resumenes[0]["carton"].codigocarton, "B7-C-DIRECTA")

    def test_sala_publica_marca_compra_solo_una_vez_por_bingo(self):
        bingo_otro = Bingo(
            idbingo=8,
            titulobingo="Bingo alterno",
            premiomayor=Decimal("750.00"),
        )
        partidas = [
            partida_publica_prueba(pk=20),
            partida_publica_prueba(pk=21),
            partida_publica_prueba(pk=22),
        ]
        partidas[1].idbingo = partidas[0].idbingo
        partidas[2].idbingo = bingo_otro

        class ConsultaGrupos:
            def filter(self, **_kwargs):
                resultado = Mock()
                resultado.exists.return_value = True
                return resultado

        usuario_jugador = Mock()
        usuario_jugador.is_authenticated = True
        usuario_jugador.pk = 5
        usuario_jugador.is_staff = False
        usuario_jugador.is_superuser = False
        usuario_jugador.groups = ConsultaGrupos()
        request = self.factory.get("/juego/")
        request.user = usuario_jugador

        capturado = {}

        def fake_render(_request, _template, contexto):
            capturado.update(contexto)
            return HttpResponse("Sala pública")

        consulta = Mock()
        consulta.order_by.return_value = partidas

        with (
            patch(
                "apps.bingos.views.Partidabingo.objects.select_related",
                return_value=consulta,
            ),
            patch("apps.bingos.views.render", side_effect=fake_render),
        ):
            response = sala_juego_publica(request)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            [item["mostrar_compra_carton"] for item in capturado["partidas_publicas"]],
            [True, False, True],
        )
        self.assertEqual(
            capturado["partidas_publicas"][0]["partida"].idbingo_id,
            self.bingo.pk,
        )
        self.assertEqual(
            capturado["partidas_publicas"][2]["partida"].idbingo_id,
            bingo_otro.pk,
        )

    def test_sala_publica_muestra_texto_y_enlace_solo_para_jugador(self):
        class ConsultaGrupos:
            def filter(self, **_kwargs):
                resultado = Mock()
                resultado.exists.return_value = True
                return resultado

        usuario_jugador = Mock()
        usuario_jugador.is_authenticated = True
        usuario_jugador.pk = 5
        usuario_jugador.is_staff = False
        usuario_jugador.is_superuser = False
        usuario_jugador.groups = ConsultaGrupos()
        request_jugador = self.factory.get("/juego/")
        request_jugador.user = usuario_jugador

        partidas_publicas = [
            {
                **preparar_resumen_partida_publica(self.partidas[0]),
                "mostrar_compra_carton": True,
            },
            {
                **preparar_resumen_partida_publica(self.partidas[1]),
                "mostrar_compra_carton": False,
            },
        ]

        html_jugador = render_to_string(
            "bingos/sala_juego_publica.html",
            {"partidas_publicas": partidas_publicas},
            request=request_jugador,
        )
        request_visitante = self.factory.get("/juego/")
        request_visitante.user = AnonymousUser()
        html_visitante = render_to_string(
            "bingos/sala_juego_publica.html",
            {"partidas_publicas": partidas_publicas},
            request=request_visitante,
        )
        usuario_staff = Mock(
            is_authenticated=True,
            pk=6,
            is_staff=True,
            is_superuser=False,
            groups=ConsultaGrupos(),
        )
        request_staff = self.factory.get("/juego/")
        request_staff.user = usuario_staff
        html_staff = render_to_string(
            "bingos/sala_juego_publica.html",
            {"partidas_publicas": partidas_publicas},
            request=request_staff,
        )
        usuario_superuser = Mock(
            is_authenticated=True,
            pk=7,
            is_staff=False,
            is_superuser=True,
            groups=ConsultaGrupos(),
        )
        request_superuser = self.factory.get("/juego/")
        request_superuser.user = usuario_superuser
        html_superuser = render_to_string(
            "bingos/sala_juego_publica.html",
            {"partidas_publicas": partidas_publicas},
            request=request_superuser,
        )

        self.assertIn("Comprar cartón para este Bingo", html_jugador)
        self.assertIn(
            reverse("bingos:comprar_carton_bingo", kwargs={"idbingo": self.bingo.pk}),
            html_jugador,
        )
        self.assertEqual(
            html_jugador.count(
                reverse(
                    "bingos:comprar_carton_bingo",
                    kwargs={"idbingo": self.bingo.pk},
                )
            ),
            1,
        )
        self.assertEqual(
            html_jugador.count("Comprar cartón para este Bingo"),
            1,
        )
        self.assertNotIn("Comprar cartón para este Bingo", html_visitante)
        self.assertNotIn("Comprar cartón para este Bingo", html_staff)
        self.assertNotIn("Comprar cartón para este Bingo", html_superuser)

    def test_venta_administrativa_actual_sigue_disponible(self):
        self.assertEqual(
            resolve(
                reverse(
                    "bingos:bingo_carton_nuevo",
                    kwargs={"idbingo": self.bingo.pk},
                )
            ).url_name,
            "bingo_carton_nuevo",
        )
        self.assertEqual(
            list(GenerarCartonBingoForm().fields),
            ["idjugador", "preciopagado"],
        )


class FormatoBolasBingoTests(SimpleTestCase):
    def test_letra_bingo_1_es_b(self):
        self.assertEqual(letra_bingo(1), "B")

    def test_letra_bingo_15_es_b(self):
        self.assertEqual(letra_bingo(15), "B")

    def test_letra_bingo_16_es_i(self):
        self.assertEqual(letra_bingo(16), "I")

    def test_letra_bingo_30_es_i(self):
        self.assertEqual(letra_bingo(30), "I")

    def test_letra_bingo_31_es_n(self):
        self.assertEqual(letra_bingo(31), "N")

    def test_letra_bingo_45_es_n(self):
        self.assertEqual(letra_bingo(45), "N")

    def test_letra_bingo_46_es_g(self):
        self.assertEqual(letra_bingo(46), "G")

    def test_letra_bingo_60_es_g(self):
        self.assertEqual(letra_bingo(60), "G")

    def test_letra_bingo_61_es_o(self):
        self.assertEqual(letra_bingo(61), "O")

    def test_letra_bingo_75_es_o(self):
        self.assertEqual(letra_bingo(75), "O")

    def test_numeros_fuera_del_rango_producen_error_controlado(self):
        for numero in (0, 76, -1, "doce", None):
            with self.subTest(numero=numero):
                with self.assertRaises(BolaBingoError):
                    letra_bingo(numero)

    def test_formato_de_bola_incluye_letra_y_numero(self):
        self.assertEqual(formatear_bola_bingo(39), "N-39")

    def test_parser_admite_vacios_json_y_texto_separado(self):
        self.assertEqual(parsear_bolas_cantadas(None), [])
        self.assertEqual(parsear_bolas_cantadas(""), [])
        self.assertEqual(parsear_bolas_cantadas("[]"), [])
        self.assertEqual(
            parsear_bolas_cantadas("12, I-24; N-39"),
            [12, 24, 39],
        )

    def test_parser_admite_json_antiguo_con_objetos_y_elimina_duplicados(self):
        valor = '[{"numero":12,"codigo":"B-12"},{"codigo":"I-24"},12]'

        self.assertEqual(parsear_bolas_cantadas(valor), [12, 24])

    def test_serializacion_estable_es_json_compacto_de_enteros(self):
        self.assertEqual(
            serializar_bolas_cantadas([12, "I-24", {"numero": 39}]),
            "[12,24,39]",
        )


class ExtraccionBolasTests(SimpleTestCase):
    def _partida(self, estado=ESTADO_PARTIDA_EN_CURSO, historial="[]"):
        partida = Partidabingo(
            idpartidabingo=8,
            estadopartida=estado,
            bolascantadas=historial,
            ultimabola=0,
        )
        partida.save = Mock()
        return partida

    def _extraer_sin_base(self, historial="[]", bola=12):
        partida = self._partida(historial=historial)
        generador = Mock()
        generador.choice.return_value = bola
        with (
            patch(
                "apps.bingos.services.transaction.atomic",
                return_value=nullcontext(),
            ) as atomic_mock,
            patch(
                "apps.bingos.services.Partidabingo.objects.select_for_update"
            ) as select_for_update,
        ):
            select_for_update.return_value.get.return_value = partida
            nueva_bola = extraer_siguiente_bola(
                partida,
                generador_aleatorio=generador,
            )
        return partida, nueva_bola, generador, atomic_mock, select_for_update

    def test_una_bola_extraida_no_vuelve_a_salir(self):
        partida, nueva_bola, generador, _atomic, _select = self._extraer_sin_base(
            historial="[12,24]",
            bola=13,
        )

        disponibles_entregadas = generador.choice.call_args.args[0]
        self.assertNotIn(12, disponibles_entregadas)
        self.assertNotIn(24, disponibles_entregadas)
        self.assertEqual(nueva_bola, 13)
        self.assertEqual(parsear_bolas_cantadas(partida.bolascantadas), [12, 24, 13])

    def test_bolas_disponibles_disminuyen_correctamente(self):
        antes = obtener_bolas_disponibles([1, 2])
        despues = obtener_bolas_disponibles([1, 2, 3])

        self.assertEqual(len(antes), 73)
        self.assertEqual(len(despues), 72)
        self.assertNotIn(3, despues)

    def test_al_llegar_a_75_bolas_no_se_extrae_otra(self):
        historial = serializar_bolas_cantadas(range(1, 76))
        partida = self._partida(historial=historial)
        generador = Mock()

        with (
            patch(
                "apps.bingos.services.transaction.atomic",
                return_value=nullcontext(),
            ),
            patch(
                "apps.bingos.services.Partidabingo.objects.select_for_update"
            ) as select_for_update,
        ):
            select_for_update.return_value.get.return_value = partida
            with self.assertRaises(BolilleroAgotadoError):
                extraer_siguiente_bola(partida, generador_aleatorio=generador)

        self.assertEqual(partida.bolascantadas, historial)
        partida.save.assert_not_called()
        generador.choice.assert_not_called()

    def test_no_se_puede_sacar_bola_en_programada(self):
        partida = self._partida(estado=ESTADO_PARTIDA_PROGRAMADA)

        with patch("apps.bingos.services.transaction.atomic") as atomic_mock:
            with self.assertRaises(BolaBingoError):
                extraer_siguiente_bola(partida)

        atomic_mock.assert_not_called()
        partida.save.assert_not_called()

    def test_no_se_puede_sacar_bola_en_pausada(self):
        partida = self._partida(estado=ESTADO_PARTIDA_PAUSADA)

        with self.assertRaises(BolaBingoError):
            extraer_siguiente_bola(partida)

        partida.save.assert_not_called()

    def test_no_se_puede_sacar_bola_en_finalizada(self):
        partida = self._partida(estado=ESTADO_PARTIDA_FINALIZADA)

        with self.assertRaises(BolaBingoError):
            extraer_siguiente_bola(partida)

        partida.save.assert_not_called()

    def test_en_espera_desempate_y_cancelada_tambien_bloquean_extraccion(self):
        for estado in (
            ESTADO_PARTIDA_EN_ESPERA,
            ESTADO_PARTIDA_DESEMPATE,
            ESTADO_PARTIDA_CANCELADA,
        ):
            with self.subTest(estado=estado):
                partida = self._partida(estado=estado)
                with self.assertRaises(BolaBingoError):
                    extraer_siguiente_bola(partida)
                partida.save.assert_not_called()

    def test_si_se_puede_sacar_bola_en_curso(self):
        partida, nueva_bola, _generador, _atomic, _select = self._extraer_sin_base(
            bola=55
        )

        self.assertEqual(nueva_bola, 55)
        partida.save.assert_called_once_with(
            update_fields=["bolascantadas", "ultimabola"]
        )

    def test_extraccion_actualiza_ultima_bola(self):
        partida, _nueva_bola, _generador, _atomic, _select = self._extraer_sin_base(
            bola=73
        )

        self.assertEqual(partida.ultimabola, 73)

    def test_extraccion_conserva_historial_previo(self):
        partida, _nueva_bola, _generador, _atomic, _select = self._extraer_sin_base(
            historial="1, I-16, 31",
            bola=46,
        )

        self.assertEqual(partida.bolascantadas, "[1,16,31,46]")

    def test_extraccion_usa_transaccion_y_select_for_update(self):
        _partida, _bola, _generador, atomic_mock, select_for_update = (
            self._extraer_sin_base(bola=7)
        )

        atomic_mock.assert_called_once_with()
        select_for_update.assert_called_once_with()
        select_for_update.return_value.get.assert_called_once_with(pk=8)


class RutaExtraccionBolaTests(SimpleTestCase):
    def setUp(self):
        self.factory = RequestFactory()

    def test_ruta_post_rechaza_usuario_no_administrativo(self):
        request = self.factory.post("/partidas/1/sacar-bola/")
        request.user = User(username="jugador", is_staff=False, is_superuser=False)

        with patch("apps.bingos.views.extraer_siguiente_bola") as extraer_mock:
            with self.assertRaises(PermissionDenied):
                sacar_bola(request, 1)

        extraer_mock.assert_not_called()

    def test_ruta_get_no_ejecuta_extraccion(self):
        request = self.factory.get("/partidas/1/sacar-bola/")
        request.user = User(username="admin", is_staff=True)

        with patch("apps.bingos.views.extraer_siguiente_bola") as extraer_mock:
            response = sacar_bola(request, 1)

        self.assertEqual(response.status_code, 405)
        extraer_mock.assert_not_called()

    def test_ruta_post_extrae_y_redirige_a_consola(self):
        request = self.factory.post("/partidas/1/sacar-bola/")
        request.user = User(username="admin", is_staff=True)
        request.session = {}
        request._messages = FallbackStorage(request)
        partida = Partidabingo(
            idpartidabingo=1,
            estadopartida=ESTADO_PARTIDA_EN_CURSO,
        )

        with (
            patch("apps.bingos.views.get_object_or_404", return_value=partida),
            patch("apps.bingos.views.extraer_siguiente_bola", return_value=12),
            patch("apps.bingos.views.programar_publicacion_partida") as publicar_mock,
        ):
            response = sacar_bola(request, 1)

        self.assertEqual(response.status_code, 302)
        self.assertIn("/partidas/1/consola/", response["Location"])
        self.assertIn(
            "Bola B-12 extraída correctamente.",
            [message.message for message in get_messages(request)],
        )
        publicar_mock.assert_called_once_with(partida, "bola_extraida")


class ReglaGanadorCartonTests(SimpleTestCase):
    def setUp(self):
        self.matriz = matriz_carton_prueba()
        self.numeros = obtener_numeros_carton(self.matriz)

    def test_matriz_valida_5x5_con_libre_se_interpreta_correctamente(self):
        serializada = serializar_matriz_carton_bingo(self.matriz)

        self.assertEqual(deserializar_matriz_carton_bingo(serializada), self.matriz)
        self.assertEqual(obtener_numeros_carton(serializada), self.numeros)
        self.assertEqual(len(self.numeros), 24)

    def test_carton_con_todos_los_numeros_llamados_es_ganador(self):
        self.assertTrue(
            carton_tiene_bingo_completo(self.matriz, self.numeros)
        )

    def test_carton_con_un_numero_pendiente_no_es_ganador(self):
        bolas = self.numeros[:-1]

        self.assertFalse(carton_tiene_bingo_completo(self.matriz, bolas))
        self.assertEqual(
            obtener_numeros_faltantes_carton(self.matriz, bolas),
            [self.numeros[-1]],
        )

    def test_casilla_libre_no_cuenta_como_numero_pendiente(self):
        faltantes = obtener_numeros_faltantes_carton(
            self.matriz,
            self.numeros,
        )

        self.assertEqual(faltantes, [])
        self.assertNotIn(CASILLA_LIBRE, obtener_numeros_carton(self.matriz))

    def test_matriz_danada_no_puede_ganar(self):
        matriz_danada = "[[1,2,3],[4,5,6]]"

        self.assertFalse(carton_tiene_bingo_completo(matriz_danada, range(1, 76)))
        with self.assertRaises(MatrizCartonInvalidaError):
            obtener_numeros_carton(matriz_danada)


class PatronGanadorCartonTests(SimpleTestCase):
    def setUp(self):
        self.matriz = matriz_carton_prueba()

    def _matriz_marcada(self, numeros_marcados):
        marcados = set(numeros_marcados)
        return [
            [
                {
                    "valor": valor,
                    "libre": valor == CASILLA_LIBRE,
                    "marcada": valor == CASILLA_LIBRE or valor in marcados,
                }
                for valor in fila
            ]
            for fila in self.matriz
        ]

    def test_carton_lleno_complete_e_incompleto(self):
        completa = self._matriz_marcada(obtener_numeros_carton(self.matriz))
        incompleta = self._matriz_marcada([1, 2, 3])

        self.assertTrue(carton_cumple_patron_ganador(completa, "carton_lleno"))
        self.assertFalse(
            carton_cumple_patron_ganador(incompleta, "carton_lleno")
        )
        self.assertCountEqual(
            obtener_numeros_faltantes_patron_ganador(
                incompleta,
                "carton_lleno",
            ),
            [16, 31, 46, 61, 17, 32, 47, 62, 18, 48, 63, 4, 19, 33, 49, 64, 5, 20, 34, 50, 65],
        )
        self.assertEqual(total_casillas_patron_ganador("carton_lleno"), 24)

    def test_linea_horizontal_y_vertical_con_tie_elige_la_primera(self):
        horizontal = self._matriz_marcada([1, 16, 31, 46, 2, 17, 32, 47])
        vertical = self._matriz_marcada([1, 2, 3, 4, 16, 17, 18, 19])

        self.assertEqual(
            obtener_numeros_faltantes_patron_ganador(
                horizontal,
                "linea_horizontal",
            ),
            [61],
        )
        self.assertEqual(
            obtener_numeros_faltantes_patron_ganador(
                vertical,
                "linea_vertical",
            ),
            [5],
        )

    def test_diagonal_cuatro_esquinas_cruz_y_x(self):
        diagonal = self._matriz_marcada([17, 49, 65, 47, 19, 5])
        cuatro_esquinas = self._matriz_marcada([])
        cruz = self._matriz_marcada([3, 48, 63, 31, 32, 33, 34])
        x = self._matriz_marcada([17, 49, 65, 47, 19, 5])

        self.assertEqual(
            obtener_numeros_faltantes_patron_ganador(diagonal, "diagonal"),
            [1],
        )
        self.assertEqual(
            obtener_numeros_faltantes_patron_ganador(
                cuatro_esquinas,
                "cuatro_esquinas",
            ),
            [1, 61, 5, 65],
        )
        self.assertEqual(
            obtener_numeros_faltantes_patron_ganador(cruz, "cruz"),
            [18],
        )
        self.assertEqual(
            obtener_numeros_faltantes_patron_ganador(x, "x"),
            [1, 61],
        )
        self.assertEqual(total_casillas_patron_ganador("cuatro_esquinas"), 4)
        self.assertEqual(total_casillas_patron_ganador("cruz"), 8)
        self.assertEqual(total_casillas_patron_ganador("x"), 8)

    def test_casilla_libre_central_cuenta_como_marcada(self):
        matriz = self._matriz_marcada([3, 48, 63, 31, 32, 33, 34])

        self.assertTrue(matriz[2][2]["libre"])
        self.assertTrue(matriz[2][2]["marcada"])
        self.assertEqual(
            obtener_numeros_faltantes_patron_ganador(matriz, "cruz"),
            [18],
        )

    def test_patron_invalido_usando_fallback_carton_lleno(self):
        matriz = self._matriz_marcada([1, 2, 3])

        self.assertEqual(normalizar_patron_ganador("patron-raro"), "carton_lleno")
        self.assertEqual(obtener_etiqueta_patron_ganador("patron-raro"), "Cartón lleno")
        self.assertEqual(
            obtener_numeros_faltantes_patron_ganador(matriz, "patron-raro"),
            obtener_numeros_faltantes_patron_ganador(matriz, "carton_lleno"),
        )

    def test_resumen_patron_cuatro_esquinas_calcula_progreso_y_faltantes(self):
        resumen = preparar_resumen_patron_ganador(
            self.matriz,
            [1, 61, 5],
            "cuatro_esquinas",
        )

        self.assertEqual(resumen["patron_ganador_label"], "Cuatro esquinas")
        self.assertEqual(resumen["total_requeridos_patron"], 4)
        self.assertEqual(resumen["numeros_marcados_patron"], 3)
        self.assertEqual(resumen["progreso_patron"], 75)
        self.assertEqual(resumen["numeros_faltantes_patron"], [65])
        self.assertEqual(resumen["faltantes_patron_codigos"], ["O-65"])
        self.assertFalse(resumen["patron_completo"])
        self.assertNotIn("LIBRE", resumen["faltantes_patron_codigos"])

    def test_error_de_incompleto_menciona_el_patron_legible(self):
        error = CartonNoCompletoError([1, 61], "cuatro_esquinas")

        self.assertIn("Cuatro esquinas", str(error))
        self.assertIn("Faltan: B-1, O-61", str(error))


class ValidacionGanadorTests(SimpleTestCase):
    def _partida(self, estado=ESTADO_PARTIDA_EN_CURSO, bolas=None, pk=20):
        return Partidabingo(
            idpartidabingo=pk,
            estadopartida=estado,
            bolascantadas=serializar_bolas_cantadas(bolas or []),
            ultimabola=(bolas or [0])[-1],
            idjugadorganador=None,
            haydesempate=False,
            idbingadores=None,
        )

    def _carton(self, partida, pk=30, jugador_pk=40, matriz=None):
        jugador = Jugador(
            idjugador=jugador_pk,
            aliasjugador=f"Jugador {jugador_pk}",
        )
        carton = Carton(
            idcarton=pk,
            idpartida=partida,
            idjugador=jugador,
            codigocarton=f"P{partida.pk}-C-{pk}",
            matriznumeros=serializar_matriz_carton_bingo(
                matriz or matriz_carton_prueba()
            ),
            estadocarton="Vendido",
        )
        carton.save = Mock()
        return carton

    def _validar_sin_base(self, partida, carton, cartones, save_error=None):
        partida_bloqueada = Partidabingo(
            idpartidabingo=partida.pk,
            estadopartida=partida.estadopartida,
            bolascantadas=partida.bolascantadas,
            ultimabola=partida.ultimabola,
            patronganador=getattr(partida, "patronganador", None),
            idjugadorganador=partida.idjugadorganador,
            haydesempate=partida.haydesempate,
            idbingadores=partida.idbingadores,
        )
        partida_bloqueada.save = Mock(side_effect=save_error)

        with (
            patch(
                "apps.bingos.services.transaction.atomic",
                return_value=nullcontext(),
            ) as atomic_mock,
            patch(
                "apps.bingos.services.Partidabingo.objects.select_for_update"
            ) as lock_partida,
            patch(
                "apps.bingos.services.Carton.objects.select_for_update"
            ) as lock_cartones,
        ):
            lock_partida.return_value.get.return_value = partida_bloqueada
            consulta = lock_cartones.return_value.filter.return_value
            consulta.select_related.return_value.order_by.return_value = cartones
            resultado = validar_carton_ganador(partida, carton)

        return (
            resultado,
            partida_bloqueada,
            atomic_mock,
            lock_partida,
            lock_cartones,
        )

    def test_carton_no_puede_validarse_si_pertenece_a_otra_partida(self):
        partida = self._partida(pk=1)
        otra_partida = self._partida(pk=2)
        carton = self._carton(otra_partida)

        with patch("apps.bingos.services.transaction.atomic") as atomic_mock:
            with self.assertRaises(ValidacionCartonError):
                validar_carton_ganador(partida, carton)

        atomic_mock.assert_not_called()

    def test_no_se_puede_validar_en_programada(self):
        partida = self._partida(estado=ESTADO_PARTIDA_PROGRAMADA)
        carton = self._carton(partida)

        with self.assertRaises(ValidacionCartonError):
            validar_carton_ganador(partida, carton)

    def test_no_se_puede_validar_en_pausada(self):
        partida = self._partida(estado=ESTADO_PARTIDA_PAUSADA)
        carton = self._carton(partida)

        with self.assertRaises(ValidacionCartonError):
            validar_carton_ganador(partida, carton)

    def test_no_se_puede_validar_en_finalizada(self):
        partida = self._partida(estado=ESTADO_PARTIDA_FINALIZADA)
        carton = self._carton(partida)

        with self.assertRaises(ValidacionCartonError):
            validar_carton_ganador(partida, carton)

    def test_en_espera_desempate_y_cancelada_tambien_bloquean_validacion(self):
        for estado in (
            ESTADO_PARTIDA_EN_ESPERA,
            ESTADO_PARTIDA_DESEMPATE,
            ESTADO_PARTIDA_CANCELADA,
        ):
            with self.subTest(estado=estado):
                partida = self._partida(estado=estado)
                carton = self._carton(partida)
                with self.assertRaises(ValidacionCartonError):
                    validar_carton_ganador(partida, carton)

    def test_si_se_puede_validar_en_curso(self):
        numeros = obtener_numeros_carton(matriz_carton_prueba())
        partida = self._partida(bolas=numeros)
        carton = self._carton(partida)

        resultado, _bloqueada, _atomic, _lock_partida, _lock_cartones = (
            self._validar_sin_base(partida, carton, [carton])
        )

        self.assertEqual(resultado["resultado"], "ganador")

    def test_un_carton_ganador_registra_al_jugador(self):
        numeros = obtener_numeros_carton(matriz_carton_prueba())
        partida = self._partida(bolas=numeros)
        carton = self._carton(partida, jugador_pk=77)

        resultado, bloqueada, _atomic, _lock_partida, _lock_cartones = (
            self._validar_sin_base(partida, carton, [carton])
        )

        self.assertEqual(resultado["resultado"], "ganador")
        self.assertEqual(partida.idjugadorganador_id, 77)
        self.assertFalse(partida.haydesempate)
        registros = parsear_candidatos_desempate(partida.idbingadores)
        self.assertEqual(len(registros), 1)
        self.assertEqual(registros[0]["idcarton"], carton.pk)
        self.assertEqual(registros[0]["idjugador"], 77)
        self.assertEqual(partida.estadopartida, ESTADO_PARTIDA_EN_CURSO)
        bloqueada.save.assert_called_once_with(
            update_fields=[
                "idjugadorganador",
                "haydesempate",
                "idbingadores",
            ]
        )

    def test_varios_cartones_ganadores_cambian_a_desempate(self):
        numeros = obtener_numeros_carton(matriz_carton_prueba())
        partida = self._partida(bolas=numeros)
        carton_uno = self._carton(partida, pk=31, jugador_pk=41)
        carton_dos = self._carton(partida, pk=32, jugador_pk=42)

        resultado, _bloqueada, _atomic, _lock_partida, _lock_cartones = (
            self._validar_sin_base(
                partida,
                carton_uno,
                [carton_uno, carton_dos],
            )
        )

        self.assertEqual(resultado["resultado"], "desempate")
        self.assertEqual(partida.estadopartida, ESTADO_PARTIDA_DESEMPATE)
        self.assertTrue(partida.haydesempate)
        candidatos = parsear_candidatos_desempate(partida.idbingadores)
        self.assertEqual(
            [candidato["idcarton"] for candidato in candidatos],
            [31, 32],
        )

    def test_varios_ganadores_no_eligen_un_ganador_arbitrario(self):
        numeros = obtener_numeros_carton(matriz_carton_prueba())
        partida = self._partida(bolas=numeros)
        carton_uno = self._carton(partida, pk=31, jugador_pk=41)
        carton_dos = self._carton(partida, pk=32, jugador_pk=42)

        self._validar_sin_base(
            partida,
            carton_uno,
            [carton_uno, carton_dos],
        )

        self.assertIsNone(partida.idjugadorganador)

    def test_validacion_historica_con_cuatro_esquinas_reconoce_desempate(self):
        partida = self._partida(bolas=[1, 61, 5, 65])
        partida.patronganador = "cuatro_esquinas"
        carton_uno = self._carton(partida, pk=31, jugador_pk=41)
        carton_dos = self._carton(partida, pk=32, jugador_pk=42)

        resultado, _bloqueada, _atomic, _lock_partida, _lock_cartones = (
            self._validar_sin_base(
                partida,
                carton_uno,
                [carton_uno, carton_dos],
            )
        )

        self.assertEqual(resultado["resultado"], "desempate")
        self.assertEqual(partida.estadopartida, ESTADO_PARTIDA_DESEMPATE)
        self.assertTrue(partida.haydesempate)
        candidatos = parsear_candidatos_desempate(partida.idbingadores)
        self.assertEqual(
            [candidato["idcarton"] for candidato in candidatos],
            [31, 32],
        )

    def test_carton_incompleto_no_modifica_partida(self):
        numeros = obtener_numeros_carton(matriz_carton_prueba())
        partida = self._partida(bolas=numeros[:-1])
        carton = self._carton(partida)

        with self.assertRaises(CartonNoCompletoError) as error:
            self._validar_sin_base(partida, carton, [carton])

        self.assertIn(formatear_bola_bingo(numeros[-1]), str(error.exception))
        self.assertIsNone(partida.idjugadorganador)
        self.assertEqual(partida.estadopartida, ESTADO_PARTIDA_EN_CURSO)

    def test_validacion_conserva_bolas_y_no_modifica_cartones(self):
        numeros = obtener_numeros_carton(matriz_carton_prueba())
        partida = self._partida(bolas=numeros)
        carton = self._carton(partida)
        bolas_antes = partida.bolascantadas
        matriz_antes = carton.matriznumeros

        self._validar_sin_base(partida, carton, [carton])

        self.assertEqual(partida.bolascantadas, bolas_antes)
        self.assertEqual(carton.matriznumeros, matriz_antes)
        carton.save.assert_not_called()

    def test_validacion_usa_transaccion_y_bloquea_partida_y_cartones(self):
        numeros = obtener_numeros_carton(matriz_carton_prueba())
        partida = self._partida(bolas=numeros)
        carton = self._carton(partida)

        _resultado, _bloqueada, atomic, lock_partida, lock_cartones = (
            self._validar_sin_base(partida, carton, [carton])
        )

        atomic.assert_called_once_with()
        lock_partida.assert_called_once_with()
        lock_partida.return_value.get.assert_called_once_with(pk=partida.pk)
        lock_cartones.assert_called_once_with(of=("self",))

    def test_error_al_guardar_no_sincroniza_cambios_parciales(self):
        numeros = obtener_numeros_carton(matriz_carton_prueba())
        partida = self._partida(bolas=numeros)
        carton = self._carton(partida)

        with self.assertRaises(DatabaseError):
            self._validar_sin_base(
                partida,
                carton,
                [carton],
                save_error=DatabaseError("fallo simulado"),
            )

        self.assertIsNone(partida.idjugadorganador)
        self.assertFalse(partida.haydesempate)
        self.assertEqual(partida.estadopartida, ESTADO_PARTIDA_EN_CURSO)

    def test_validacion_en_partida_no_en_curso_sigue_bloqueada(self):
        for estado in (
            ESTADO_PARTIDA_PROGRAMADA,
            ESTADO_PARTIDA_PAUSADA,
            ESTADO_PARTIDA_EN_ESPERA,
            ESTADO_PARTIDA_DESEMPATE,
            ESTADO_PARTIDA_FINALIZADA,
            ESTADO_PARTIDA_CANCELADA,
        ):
            with self.subTest(estado=estado):
                partida = self._partida(estado=estado)
                carton = self._carton(partida)
                with self.assertRaises(ValidacionCartonError):
                    validar_carton_ganador(partida, carton)

    def test_evaluar_carton_danado_genera_error_controlado_y_evidencia(self):
        partida = self._partida(bolas=range(1, 76))
        carton = self._carton(partida)
        carton.matriznumeros = "[[1,2,3],[4,5,6]]"

        with self.assertRaises(MatrizCartonInvalidaError):
            evaluar_carton_en_partida(carton, partida)
        with self.assertLogs("apps.bingos.services", level="WARNING"):
            preparar_cartones_para_validacion(partida, [carton])

    def test_consola_solo_prepara_cartones_vendidos_y_asignados(self):
        partida = self._partida()
        vendido = self._carton(partida, pk=31)
        disponible = self._carton(partida, pk=32)
        disponible.estadocarton = "Disponible"
        sin_jugador = self._carton(partida, pk=33)
        sin_jugador.idjugador = None

        resultado = preparar_cartones_para_validacion(
            partida,
            [vendido, disponible, sin_jugador],
        )

        self.assertEqual(
            [item["carton"].pk for item in resultado],
            [vendido.pk],
        )

    def test_preparacion_con_cuatro_esquinas_calcula_total_y_faltantes(self):
        partida = self._partida(
            bolas=[1, 61, 5],
        )
        partida.patronganador = "cuatro_esquinas"
        carton = self._carton(partida, pk=31)

        resultado = preparar_cartones_para_validacion(partida, [carton])

        self.assertEqual(resultado[0]["patron_ganador"], "cuatro_esquinas")
        self.assertEqual(resultado[0]["patron_ganador_label"], "Cuatro esquinas")
        self.assertEqual(resultado[0]["total_requeridos"], 4)
        self.assertEqual(resultado[0]["cantidad_marcados"], 3)
        self.assertEqual(resultado[0]["progreso"], 75)
        self.assertEqual(resultado[0]["faltantes"], [65])
        self.assertEqual(resultado[0]["faltantes_codigos"], ["O-65"])
        self.assertFalse(resultado[0]["completo"])


class ParticipacionGanadoraHibridaTests(SimpleTestCase):
    def setUp(self):
        self.bingo = Bingo(idbingo=7, titulobingo="Bingo híbrido")
        self.jugador = Jugador(idjugador=4, aliasjugador="jugador4")
        self.matriz = matriz_carton_prueba()
        self.numeros = obtener_numeros_carton(self.matriz)

    def _partida(
        self,
        pk=21,
        bolas=None,
        estado=ESTADO_PARTIDA_EN_CURSO,
        bingo=None,
    ):
        bolas = self.numeros if bolas is None else list(bolas)
        return Partidabingo(
            idpartidabingo=pk,
            idbingo=bingo or self.bingo,
            nombreronda=f"Ronda {pk}",
            estadopartida=estado,
            bolascantadas=serializar_bolas_cantadas(bolas),
            ultimabola=bolas[-1] if bolas else 0,
            idjugadorganador=None,
            haydesempate=False,
            idbingadores=None,
        )

    def _carton(self, pk=40, jugador=None, bingo=None, matriz=None):
        carton = Carton(
            idcarton=pk,
            idbingo=bingo or self.bingo,
            idjugador=jugador or self.jugador,
            idpartida=None,
            indicevictoria=None,
            codigocarton=f"B7-C-{pk}",
            matriznumeros=serializar_matriz_carton_bingo(
                matriz or self.matriz
            ),
            estadocarton="Vendido",
        )
        carton.save = Mock()
        return carton

    def _participacion(
        self,
        carton,
        partida,
        pk=51,
        estado=CartonPartidaBingo.ESTADO_PENDIENTE,
        bingo=None,
    ):
        participacion = CartonPartidaBingo(
            idcartonpartidabingo=pk,
            idcarton=carton,
            idpartida=partida,
            idbingo=bingo or self.bingo,
            estado_participacion=estado,
            indicevictoria=None,
            es_asignacion_original=False,
            origen_asignacion=CartonPartidaBingo.ORIGEN_APLICACION,
            fechavalidacion=None,
        )
        participacion.save = Mock()
        return participacion

    def _partida_bloqueada(self, partida):
        bloqueada = Partidabingo(
            idpartidabingo=partida.pk,
            idbingo=partida.idbingo,
            nombreronda=partida.nombreronda,
            estadopartida=partida.estadopartida,
            bolascantadas=partida.bolascantadas,
            ultimabola=partida.ultimabola,
            patronganador=getattr(partida, "patronganador", None),
            idjugadorganador=partida.idjugadorganador,
            haydesempate=partida.haydesempate,
            idbingadores=partida.idbingadores,
        )
        bloqueada.save = Mock()
        return bloqueada

    def _validar_sin_base(
        self,
        partida,
        carton,
        participaciones,
        indice=24,
        now=None,
    ):
        bloqueada = self._partida_bloqueada(partida)
        cartones = {}
        for participacion in participaciones:
            participacion.idpartida = bloqueada
            cartones[participacion.idcarton_id] = participacion.idcarton
        cartones.setdefault(carton.pk, carton)
        with (
            patch(
                "apps.bingos.services.transaction.atomic",
                return_value=nullcontext(),
            ) as atomic_mock,
            patch(
                "apps.bingos.services._bloquear_contexto_participaciones",
                return_value=(bloqueada, participaciones, cartones),
            ) as bloquear_mock,
        ):
            resultado = validar_participacion_ganadora(
                partida,
                carton,
                indice,
                now=now,
            )
        return resultado, bloqueada, atomic_mock, bloquear_mock

    def test_resuelve_participacion_por_carton_y_partida(self):
        partida = self._partida()
        carton = self._carton()
        participacion = self._participacion(carton, partida)

        with (
            patch(
                "apps.bingos.services.Partidabingo.objects.get",
                return_value=partida,
            ) as partida_get,
            patch(
                "apps.bingos.services.Carton.objects.get",
                return_value=carton,
            ) as carton_get,
            patch(
                "apps.bingos.services.CartonPartidaBingo.objects.select_related"
            ) as seleccionar,
        ):
            seleccionar.return_value.get.return_value = participacion
            resultado = obtener_participacion_carton_en_partida(
                partida.pk,
                carton.pk,
            )

        self.assertEqual(resultado, participacion)
        partida_get.assert_called_once_with(pk=partida.pk)
        carton_get.assert_called_once_with(pk=carton.pk)
        seleccionar.return_value.get.assert_called_once_with(
            idcarton_id=carton.pk,
            idpartida_id=partida.pk,
        )

    def test_rechaza_pareja_sin_participacion(self):
        partida = self._partida()
        carton = self._carton()
        with (
            patch(
                "apps.bingos.services.Partidabingo.objects.get",
                return_value=partida,
            ),
            patch(
                "apps.bingos.services.Carton.objects.get",
                return_value=carton,
            ),
            patch(
                "apps.bingos.services.CartonPartidaBingo.objects.select_related"
            ) as seleccionar,
        ):
            seleccionar.return_value.get.side_effect = (
                CartonPartidaBingo.DoesNotExist
            )
            with self.assertRaisesMessage(
                ValidacionCartonError,
                "No existe una participación",
            ):
                obtener_participacion_carton_en_partida(partida, carton)

    def test_rechaza_participacion_de_otro_bingo(self):
        partida = self._partida()
        carton = self._carton()
        otro_bingo = Bingo(idbingo=8, titulobingo="Otro")
        participacion = self._participacion(
            carton,
            partida,
            bingo=otro_bingo,
        )

        with self.assertRaisesMessage(
            ValidacionCartonError,
            "participación pertenece a otro Bingo",
        ):
            evaluar_participacion_en_partida(participacion, partida)

    def test_evalua_solo_bolas_de_la_partida_exacta(self):
        carton = self._carton()
        primera = self._partida(pk=21, bolas=self.numeros)
        segunda = self._partida(pk=22, bolas=self.numeros[:-1])
        participacion_primera = self._participacion(carton, primera, pk=51)
        participacion_segunda = self._participacion(carton, segunda, pk=52)

        self.assertTrue(
            evaluar_participacion_en_partida(
                participacion_primera,
                primera,
            )["completo"]
        )
        resultado_segunda = evaluar_participacion_en_partida(
            participacion_segunda,
            segunda,
        )
        self.assertFalse(resultado_segunda["completo"])
        self.assertEqual(resultado_segunda["faltantes"], [self.numeros[-1]])

    def test_marca_solo_participacion_sin_cambiar_historicos_del_maestro(self):
        partida = self._partida()
        otra_partida = self._partida(pk=22)
        carton = self._carton()
        participacion = self._participacion(carton, partida, pk=51)
        otra_participacion = self._participacion(
            carton,
            otra_partida,
            pk=52,
        )
        fecha = timezone.now()

        resultado, bloqueada, atomic_mock, bloquear_mock = (
            self._validar_sin_base(
                partida,
                carton,
                [participacion],
                indice=24,
                now=fecha,
            )
        )

        self.assertEqual(resultado["resultado"], "ganador")
        self.assertEqual(
            participacion.estado_participacion,
            CartonPartidaBingo.ESTADO_GANADOR,
        )
        self.assertEqual(participacion.indicevictoria, 24)
        self.assertEqual(participacion.fechavalidacion, fecha)
        self.assertEqual(
            otra_participacion.estado_participacion,
            CartonPartidaBingo.ESTADO_PENDIENTE,
        )
        otra_participacion.save.assert_not_called()
        self.assertIsNone(carton.idpartida)
        self.assertIsNone(carton.indicevictoria)
        carton.save.assert_not_called()
        atomic_mock.assert_called_once_with()
        bloquear_mock.assert_called_once_with(
            partida.pk,
            carton_id_adicional=carton.pk,
        )
        participacion.save.assert_called_once_with(
            update_fields=[
                "estado_participacion",
                "indicevictoria",
                "fechavalidacion",
            ]
        )
        bloqueada.save.assert_called_once()

    def test_mismo_carton_gana_dos_rondas_sin_mezclar_resultados(self):
        primera = self._partida(pk=21)
        segunda = self._partida(pk=22)
        carton = self._carton()
        participacion_primera = self._participacion(carton, primera, pk=51)
        participacion_segunda = self._participacion(carton, segunda, pk=52)
        fecha_primera = timezone.now()
        fecha_segunda = fecha_primera + timedelta(seconds=1)

        self._validar_sin_base(
            primera,
            carton,
            [participacion_primera],
            indice=24,
            now=fecha_primera,
        )
        guardados_primera = participacion_primera.save.call_count
        self._validar_sin_base(
            segunda,
            carton,
            [participacion_segunda],
            indice=25,
            now=fecha_segunda,
        )

        self.assertEqual(
            participacion_primera.estado_participacion,
            CartonPartidaBingo.ESTADO_GANADOR,
        )
        self.assertEqual(
            participacion_segunda.estado_participacion,
            CartonPartidaBingo.ESTADO_GANADOR,
        )
        self.assertEqual(participacion_primera.indicevictoria, 24)
        self.assertEqual(participacion_segunda.indicevictoria, 25)
        self.assertEqual(participacion_primera.fechavalidacion, fecha_primera)
        self.assertEqual(participacion_segunda.fechavalidacion, fecha_segunda)
        self.assertEqual(participacion_primera.save.call_count, guardados_primera)
        self.assertIsNone(carton.idpartida)
        self.assertIsNone(carton.indicevictoria)
        carton.save.assert_not_called()

    def test_rechaza_indice_de_victoria_no_positivo(self):
        partida = self._partida()
        carton = self._carton()
        for indice in (0, -1, None, True, "no-numero"):
            with self.subTest(indice=indice):
                with patch(
                    "apps.bingos.services.transaction.atomic"
                ) as atomic_mock:
                    with self.assertRaisesMessage(
                        ValidacionCartonError,
                        "mayor que cero",
                    ):
                        validar_participacion_ganadora(
                            partida,
                            carton,
                            indice,
                        )
                atomic_mock.assert_not_called()

    def test_no_elige_arbitrariamente_si_hay_varias_ganadoras(self):
        partida = self._partida()
        otro_carton = self._carton(pk=41, jugador=self.jugador)
        carton = self._carton(pk=40, jugador=self.jugador)
        primera = self._participacion(carton, partida, pk=51)
        segunda = self._participacion(otro_carton, partida, pk=52)

        resultado, bloqueada, _atomic, _lock = self._validar_sin_base(
            partida,
            carton,
            [primera, segunda],
        )
        candidatos = normalizar_candidatos_desempate_participaciones(
            bloqueada.idbingadores,
            partida=bloqueada,
        )

        self.assertEqual(resultado["resultado"], "desempate")
        self.assertEqual(bloqueada.estadopartida, ESTADO_PARTIDA_DESEMPATE)
        self.assertEqual(
            [item["idcartonpartidabingo"] for item in candidatos],
            [51, 52],
        )
        primera.save.assert_not_called()
        segunda.save.assert_not_called()

    def test_validacion_hibrida_con_cuatro_esquinas_reconoce_desempate(self):
        partida = self._partida(bolas=[1, 61, 5, 65])
        partida.patronganador = "cuatro_esquinas"
        primer_carton = self._carton(pk=40, jugador=self.jugador)
        segundo_carton = self._carton(pk=41, jugador=self.jugador)
        primera = self._participacion(primer_carton, partida, pk=51)
        segunda = self._participacion(segundo_carton, partida, pk=52)

        resultado, bloqueada, _atomic, _lock = self._validar_sin_base(
            partida,
            primer_carton,
            [primera, segunda],
        )

        self.assertEqual(resultado["resultado"], "desempate")
        self.assertEqual(bloqueada.estadopartida, ESTADO_PARTIDA_DESEMPATE)
        self.assertTrue(bloqueada.haydesempate)
        candidatos = normalizar_candidatos_desempate_participaciones(
            bloqueada.idbingadores,
            partida=bloqueada,
        )
        self.assertEqual(
            [item["idcartonpartidabingo"] for item in candidatos],
            [51, 52],
        )

    def test_conserva_dos_candidatas_del_mismo_jugador(self):
        partida = self._partida()
        primer_carton = self._carton(pk=40, jugador=self.jugador)
        segundo_carton = self._carton(pk=41, jugador=self.jugador)
        candidatas = [
            construir_candidato_desempate_participacion(
                self._participacion(primer_carton, partida, pk=51)
            ),
            construir_candidato_desempate_participacion(
                self._participacion(segundo_carton, partida, pk=52)
            ),
        ]

        normalizadas = normalizar_candidatos_desempate_participaciones(
            candidatas,
            partida=partida,
        )

        self.assertEqual(len(normalizadas), 2)
        self.assertEqual(
            [item["idjugador"] for item in normalizadas],
            [self.jugador.pk, self.jugador.pk],
        )
        self.assertEqual(
            [item["idcartonpartidabingo"] for item in normalizadas],
            [51, 52],
        )

    def test_rechaza_candidata_y_participacion_de_otra_partida(self):
        primera = self._partida(pk=21)
        segunda = self._partida(pk=22)
        carton = self._carton()
        participacion = self._participacion(carton, primera, pk=51)
        candidato = construir_candidato_desempate_participacion(participacion)

        with self.assertRaisesMessage(
            DatosDesempateInvalidosError,
            "otra partida",
        ):
            normalizar_candidatos_desempate_participaciones(
                [candidato],
                partida=segunda,
            )
        with self.assertRaisesMessage(
            ValidacionCartonError,
            "partida indicada",
        ):
            evaluar_participacion_en_partida(participacion, segunda)

    def test_confirmar_desempate_actualiza_solo_candidatas_de_la_ronda(self):
        partida = self._partida(estado=ESTADO_PARTIDA_DESEMPATE)
        otra_partida = self._partida(pk=22)
        primer_carton = self._carton(pk=40, jugador=self.jugador)
        segundo_carton = self._carton(pk=41, jugador=self.jugador)
        primera = self._participacion(primer_carton, partida, pk=51)
        segunda = self._participacion(segundo_carton, partida, pk=52)
        externa = self._participacion(primer_carton, otra_partida, pk=53)
        candidatos = [
            construir_candidato_desempate_participacion(primera),
            construir_candidato_desempate_participacion(segunda),
        ]
        candidatos[0]["tiro_desempate"] = 70
        candidatos[1]["tiro_desempate"] = 60
        partida.idbingadores = json.dumps(candidatos)
        partida.save = Mock()
        fecha = timezone.now()

        with (
            patch(
                "apps.bingos.services.transaction.atomic",
                return_value=nullcontext(),
            ),
            patch(
                "apps.bingos.services.Partidabingo.objects.select_for_update"
            ) as lock_partida,
            patch(
                "apps.bingos.services._candidatos_y_participaciones_bloqueadas",
                return_value=(candidatos, {51: primera, 52: segunda}),
            ),
        ):
            lock_partida.return_value.get.return_value = partida
            resultado = confirmar_y_finalizar_desempate_participaciones(
                partida,
                indicevictoria=24,
                now=fecha,
            )

        self.assertEqual(resultado["participacion_ganadora"], primera)
        self.assertEqual(
            primera.estado_participacion,
            CartonPartidaBingo.ESTADO_GANADOR,
        )
        self.assertEqual(
            segunda.estado_participacion,
            CartonPartidaBingo.ESTADO_CERRADO,
        )
        self.assertEqual(primera.indicevictoria, 24)
        self.assertIsNone(segunda.indicevictoria)
        self.assertEqual(primera.fechavalidacion, fecha)
        self.assertEqual(segunda.fechavalidacion, fecha)
        self.assertEqual(
            externa.estado_participacion,
            CartonPartidaBingo.ESTADO_PENDIENTE,
        )
        externa.save.assert_not_called()
        primer_carton.save.assert_not_called()
        segundo_carton.save.assert_not_called()
        self.assertIsNone(primer_carton.idpartida)
        self.assertIsNone(primer_carton.indicevictoria)


@override_settings(
    STORAGES={
        "default": {
            "BACKEND": "django.core.files.storage.FileSystemStorage",
        },
        "staticfiles": {
            "BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage",
        },
    }
)
class ConsolaValidacionHibridaTests(SimpleTestCase):
    def setUp(self):
        self.factory = RequestFactory()
        self.bingo = Bingo(idbingo=7, titulobingo="Bingo híbrido")
        self.jugador = Jugador(idjugador=4, aliasjugador="jugador4")
        self.matriz = matriz_carton_prueba()
        self.numeros = obtener_numeros_carton(self.matriz)
        self.partida = Partidabingo(
            idpartidabingo=21,
            idbingo=self.bingo,
            nombreronda="Ronda híbrida",
            estadopartida=ESTADO_PARTIDA_EN_CURSO,
            bolascantadas=serializar_bolas_cantadas(self.numeros),
            ultimabola=self.numeros[-1],
            idbingadores=None,
            haydesempate=False,
        )
        self.carton_hibrido = Carton(
            idcarton=40,
            idbingo=self.bingo,
            idjugador=self.jugador,
            idpartida=None,
            indicevictoria=None,
            codigocarton="B7-C-HIBRIDO",
            matriznumeros=serializar_matriz_carton_bingo(self.matriz),
            estadocarton="Vendido",
        )
        self.participacion = CartonPartidaBingo(
            idcartonpartidabingo=51,
            idcarton=self.carton_hibrido,
            idpartida=self.partida,
            idbingo=self.bingo,
            estado_participacion=CartonPartidaBingo.ESTADO_EN_JUEGO,
            indicevictoria=9,
            es_asignacion_original=False,
            origen_asignacion=CartonPartidaBingo.ORIGEN_APLICACION,
            fechavalidacion=timezone.now(),
        )
        self.carton_heredado = Carton(
            idcarton=41,
            idbingo=self.bingo,
            idjugador=self.jugador,
            idpartida=self.partida,
            indicevictoria=3,
            codigocarton="P21-C-HEREDADO",
            matriznumeros=serializar_matriz_carton_bingo(self.matriz),
            estadocarton="Vendido",
        )

    def _request(self, carton=None):
        carton = carton or self.carton_hibrido
        request = self.factory.post(
            reverse(
                "bingos:validar_carton",
                kwargs={
                    "idpartidabingo": self.partida.pk,
                    "idcarton": carton.pk,
                },
            )
        )
        request.user = User(username="admin", is_staff=True)
        request.session = {}
        request._messages = FallbackStorage(request)
        return request

    def _consulta_participaciones(self, valores):
        consulta = Mock()
        consulta.select_related.return_value.order_by.return_value = valores
        return consulta

    def test_consulta_participaciones_de_la_partida_correcta(self):
        consulta = self._consulta_participaciones([self.participacion])
        with patch(
            "apps.bingos.services.CartonPartidaBingo.objects.filter",
            return_value=consulta,
        ) as filtrar:
            resultado = obtener_participaciones_hibridas_partida(self.partida)

        self.assertEqual(resultado, [self.participacion])
        filtrar.assert_called_once_with(
            idpartida=self.partida,
            idcarton__idpartida__isnull=True,
        )
        consulta.select_related.assert_called_once_with(
            "idcarton",
            "idcarton__idjugador",
            "idcarton__idbingo",
            "idpartida",
            "idpartida__idbingo",
            "idbingo",
        )

    def test_participacion_de_otra_partida_no_aparece(self):
        otra_partida = Partidabingo(
            idpartidabingo=22,
            idbingo=self.bingo,
            nombreronda="Otra ronda",
        )
        otra_participacion = CartonPartidaBingo(
            idcartonpartidabingo=52,
            idcarton=self.carton_hibrido,
            idpartida=otra_partida,
            idbingo=self.bingo,
        )
        consulta = self._consulta_participaciones([])
        with patch(
            "apps.bingos.services.CartonPartidaBingo.objects.filter",
            return_value=consulta,
        ):
            resultado = obtener_participaciones_hibridas_partida(self.partida)

        self.assertEqual(resultado, [])
        self.assertNotIn(otra_participacion, resultado)

    def test_participacion_de_otro_bingo_se_rechaza(self):
        otro_bingo = Bingo(idbingo=8, titulobingo="Otro Bingo")
        inconsistente = CartonPartidaBingo(
            idcartonpartidabingo=53,
            idcarton=self.carton_hibrido,
            idpartida=self.partida,
            idbingo=otro_bingo,
        )
        consulta = self._consulta_participaciones([inconsistente])
        with patch(
            "apps.bingos.services.CartonPartidaBingo.objects.filter",
            return_value=consulta,
        ):
            with self.assertRaisesMessage(
                ValidacionCartonError,
                "otro Bingo",
            ):
                obtener_participaciones_hibridas_partida(self.partida)

    def test_presentacion_usa_estado_e_indice_de_participacion(self):
        self.carton_hibrido.indicevictoria = 88
        datos = preparar_participaciones_hibridas_para_consola(
            self.partida,
            participaciones=[self.participacion],
        )

        self.assertEqual(len(datos), 1)
        self.assertEqual(
            datos[0]["estado_participacion"],
            CartonPartidaBingo.ESTADO_EN_JUEGO,
        )
        self.assertEqual(datos[0]["indicevictoria"], 9)
        self.assertNotEqual(
            datos[0]["indicevictoria"],
            self.carton_hibrido.indicevictoria,
        )
        self.assertEqual(datos[0]["cantidad_marcados"], 24)
        self.assertEqual(datos[0]["progreso"], 100)

    def test_presentacion_cuatro_esquinas_usa_total_cuatro_y_faltantes_de_esquina(self):
        self.partida.patronganador = "cuatro_esquinas"
        self.partida.bolascantadas = serializar_bolas_cantadas([1, 61, 5])
        self.partida.ultimabola = 5

        datos = preparar_participaciones_hibridas_para_consola(
            self.partida,
            participaciones=[self.participacion],
        )

        self.assertEqual(datos[0]["patron_ganador"], "cuatro_esquinas")
        self.assertEqual(datos[0]["patron_ganador_label"], "Cuatro esquinas")
        self.assertEqual(datos[0]["total_requeridos"], 4)
        self.assertEqual(datos[0]["cantidad_marcados"], 3)
        self.assertEqual(datos[0]["progreso"], 75)
        self.assertEqual(datos[0]["faltantes"], [65])
        self.assertEqual(datos[0]["faltantes_codigos"], ["O-65"])

    def test_consola_carga_grupos_heredado_e_hibrido_separados(self):
        request = self.factory.get(
            reverse(
                "bingos:consola_operador",
                kwargs={"idpartidabingo": self.partida.pk},
            )
        )
        request.user = User(username="admin", is_staff=True)
        consulta_cartones = Mock()
        consulta_cartones.select_related.return_value.order_by.return_value = [
            self.carton_heredado
        ]
        item_hibrido = {"participacion": self.participacion}
        with (
            patch(
                "apps.bingos.views.get_object_or_404",
                return_value=self.partida,
            ),
            patch(
                "apps.bingos.views.acciones_disponibles_consola",
                return_value=set(),
            ),
            patch(
                "apps.bingos.views.Carton.objects.filter",
                return_value=consulta_cartones,
            ),
            patch(
                "apps.bingos.views.obtener_participaciones_hibridas_partida",
                return_value=[self.participacion],
            ) as obtener_hibridas,
            patch(
                "apps.bingos.views.preparar_participaciones_hibridas_para_consola",
                return_value=[item_hibrido],
            ),
            patch(
                "apps.bingos.views.preparar_cartones_para_validacion",
                return_value=[{"carton": self.carton_heredado}],
            ),
            patch(
                "apps.bingos.views.preparar_datos_bolas_partida",
                return_value={},
            ),
            patch(
                "apps.bingos.views.render",
                return_value=HttpResponse(status=200),
            ) as render_mock,
        ):
            response = consola_operador(request, self.partida.pk)

        self.assertEqual(response.status_code, 200)
        obtener_hibridas.assert_called_once_with(self.partida)
        contexto = render_mock.call_args.args[2]
        self.assertEqual(contexto["cartones"], [self.carton_heredado])
        self.assertEqual(
            contexto["participaciones_hibridas"],
            [self.participacion],
        )
        self.assertEqual(
            contexto["participaciones_hibridas_validacion"],
            [item_hibrido],
        )

    def test_consola_muestra_total_requerido_del_patron_en_la_plantilla(self):
        item = {
            "participacion": self.participacion,
            "etiqueta_tipo": "Cartón de Bingo",
            "carton": self.carton_hibrido,
            "matriz": [],
            "patron_ganador": "cuatro_esquinas",
            "patron_ganador_label": "Cuatro esquinas",
            "faltantes": [65],
            "faltantes_codigos": ["O-65"],
            "cantidad_marcados": 3,
            "total_requeridos": 4,
            "progreso": 75,
            "completo": False,
            "estado_participacion": CartonPartidaBingo.ESTADO_EN_JUEGO,
            "indicevictoria": 9,
            "fechavalidacion": timezone.now(),
            "error": None,
            "puede_validar": True,
        }
        request = self.factory.get("/partidas/21/consola/")
        request.user = User(username="admin", is_staff=True)
        html = render_to_string(
            "bingos/consola_operador.html",
            {
                "partida": self.partida,
                "acciones_consola": [],
                "actualizacion_estados_pendiente": False,
                "cartones": [self.carton_heredado],
                "cartones_validacion": [],
                "participaciones_hibridas": [self.participacion],
                "participaciones_hibridas_validacion": [item],
                "error_participaciones_hibridas": None,
                "candidatos_desempate": [],
                "puede_asignar_cartones": False,
                "puede_validar_cartones": True,
            },
            request=request,
        )

        self.assertIn("3 de 4 números marcados · 75%", html)
        self.assertNotIn("24 de 24 números marcados", html)

    def test_validacion_heredada_conserva_servicio_existente(self):
        request = self._request(self.carton_heredado)
        resultado = {
            "resultado": "ganador",
            "partida": self.partida,
            "carton": self.carton_heredado,
        }
        with (
            patch(
                "apps.bingos.views.get_object_or_404",
                side_effect=[self.partida, self.carton_heredado],
            ),
            patch(
                "apps.bingos.views.validar_carton_ganador",
                return_value=resultado,
            ) as legado_mock,
            patch(
                "apps.bingos.views.validar_participacion_ganadora"
            ) as hibrido_mock,
            patch("apps.bingos.views.programar_publicacion_partida"),
        ):
            response = validar_carton(
                request,
                self.partida.pk,
                self.carton_heredado.pk,
            )

        self.assertEqual(response.status_code, 302)
        legado_mock.assert_called_once_with(self.partida, self.carton_heredado)
        hibrido_mock.assert_not_called()

    def test_validacion_hibrida_usa_servicio_nuevo_y_no_toca_maestro(self):
        request = self._request()
        resultado = {
            "resultado": "ganador",
            "partida": self.partida,
            "participacion": self.participacion,
        }
        with (
            patch(
                "apps.bingos.views.get_object_or_404",
                side_effect=[self.partida, self.carton_hibrido],
            ),
            patch(
                "apps.bingos.views.validar_participacion_ganadora",
                return_value=resultado,
            ) as hibrido_mock,
            patch(
                "apps.bingos.views.validar_carton_ganador"
            ) as legado_mock,
            patch("apps.bingos.views.programar_publicacion_partida"),
        ):
            response = validar_carton(
                request,
                self.partida.pk,
                self.carton_hibrido.pk,
            )

        self.assertEqual(response.status_code, 302)
        hibrido_mock.assert_called_once_with(
            partida=self.partida,
            carton=self.carton_hibrido,
            indicevictoria=len(self.numeros),
        )
        legado_mock.assert_not_called()
        self.assertIsNone(self.carton_hibrido.idpartida)
        self.assertIsNone(self.carton_hibrido.indicevictoria)

    def test_exito_hibrido_comunica_victoria_de_ronda(self):
        request = self._request()
        resultado = {
            "resultado": "ganador",
            "partida": self.partida,
            "participacion": self.participacion,
        }
        with (
            patch(
                "apps.bingos.views.get_object_or_404",
                side_effect=[self.partida, self.carton_hibrido],
            ),
            patch(
                "apps.bingos.views.validar_participacion_ganadora",
                return_value=resultado,
            ),
            patch("apps.bingos.views.programar_publicacion_partida"),
        ):
            validar_carton(
                request,
                self.partida.pk,
                self.carton_hibrido.pk,
            )

        mensajes = [mensaje.message for mensaje in get_messages(request)]
        self.assertTrue(
            any(
                self.carton_hibrido.codigocarton in mensaje
                and self.partida.nombreronda in mensaje
                and "ganó la ronda" in mensaje
                for mensaje in mensajes
            )
        )
        self.assertFalse(any("todo el Bingo" in mensaje for mensaje in mensajes))

    def test_error_hibrido_muestra_mensaje_util_sin_fallback(self):
        request = self._request()
        error = "No existe una participación para ese cartón y esa partida."
        with (
            patch(
                "apps.bingos.views.get_object_or_404",
                side_effect=[self.partida, self.carton_hibrido],
            ),
            patch(
                "apps.bingos.views.validar_participacion_ganadora",
                side_effect=ValidacionCartonError(error),
            ),
            patch(
                "apps.bingos.views.validar_carton_ganador"
            ) as legado_mock,
        ):
            response = validar_carton(
                request,
                self.partida.pk,
                self.carton_hibrido.pk,
            )

        self.assertEqual(response.status_code, 302)
        legado_mock.assert_not_called()
        self.assertIn(error, [mensaje.message for mensaje in get_messages(request)])

    def test_resultado_hibrido_de_desempate_muestra_senal_clara(self):
        request = self._request()
        resultado = {
            "resultado": "desempate",
            "partida": self.partida,
            "participacion": self.participacion,
        }
        with (
            patch(
                "apps.bingos.views.get_object_or_404",
                side_effect=[self.partida, self.carton_hibrido],
            ),
            patch(
                "apps.bingos.views.validar_participacion_ganadora",
                return_value=resultado,
            ),
            patch("apps.bingos.views.programar_publicacion_partida"),
        ):
            validar_carton(
                request,
                self.partida.pk,
                self.carton_hibrido.pk,
            )

        mensajes = [mensaje.message for mensaje in get_messages(request)]
        self.assertTrue(any("Desempate" in mensaje for mensaje in mensajes))

    def test_template_distingue_historico_y_carton_de_bingo(self):
        item = preparar_participaciones_hibridas_para_consola(
            self.partida,
            participaciones=[self.participacion],
        )[0]
        request = self.factory.get("/partidas/21/consola/")
        request.user = User(username="admin", is_staff=True)
        html = render_to_string(
            "bingos/consola_operador.html",
            {
                "partida": self.partida,
                "acciones_consola": [],
                "actualizacion_estados_pendiente": False,
                "cartones": [self.carton_heredado],
                "cartones_validacion": [],
                "participaciones_hibridas": [self.participacion],
                "participaciones_hibridas_validacion": [item],
                "error_participaciones_hibridas": None,
                "candidatos_desempate": [],
                "puede_asignar_cartones": False,
                "puede_validar_cartones": True,
            },
            request=request,
        )

        self.assertIn("Histórico por partida", html)
        self.assertIn("Cartón de Bingo", html)
        self.assertIn(self.carton_hibrido.codigocarton, html)
        self.assertIn(CartonPartidaBingo.ESTADO_EN_JUEGO, html)
        self.assertIn("Índice: 9", html)
        self.assertIn("24 de 24 números marcados", html)

    def test_rutas_principales_heredadas_siguen_resolviendo(self):
        consola_url = reverse(
            "bingos:consola_operador",
            kwargs={"idpartidabingo": self.partida.pk},
        )
        validacion_url = reverse(
            "bingos:validar_carton",
            kwargs={
                "idpartidabingo": self.partida.pk,
                "idcarton": self.carton_heredado.pk,
            },
        )

        self.assertEqual(resolve(consola_url).url_name, "consola_operador")
        self.assertEqual(resolve(validacion_url).url_name, "validar_carton")


class RutaValidacionCartonTests(SimpleTestCase):
    def setUp(self):
        self.factory = RequestFactory()

    def test_usuario_no_administrativo_recibe_acceso_denegado(self):
        request = self.factory.post("/partidas/1/cartones/2/validar/")
        request.user = User(username="jugador", is_staff=False, is_superuser=False)

        with patch("apps.bingos.views.validar_carton_ganador") as validar_mock:
            with self.assertRaises(PermissionDenied):
                validar_carton(request, 1, 2)

        validar_mock.assert_not_called()

    def test_get_no_modifica_ganador_ni_estado(self):
        request = self.factory.get("/partidas/1/cartones/2/validar/")
        request.user = User(username="admin", is_staff=True)

        with patch("apps.bingos.views.validar_carton_ganador") as validar_mock:
            response = validar_carton(request, 1, 2)

        self.assertEqual(response.status_code, 405)
        validar_mock.assert_not_called()

    def test_post_busca_el_carton_dentro_de_la_partida_de_la_url(self):
        request = self.factory.post("/partidas/1/cartones/2/validar/")
        request.user = User(username="admin", is_staff=True)
        request.session = {}
        request._messages = FallbackStorage(request)
        partida = Partidabingo(idpartidabingo=1)
        carton = Carton(idcarton=2, idpartida=partida)

        with (
            patch(
                "apps.bingos.views.get_object_or_404",
                side_effect=[partida, carton],
            ) as obtener_mock,
            patch(
                "apps.bingos.views.validar_carton_ganador",
                side_effect=ValidacionCartonError("validación simulada"),
            ),
        ):
            response = validar_carton(request, 1, 2)

        self.assertEqual(response.status_code, 302)
        self.assertEqual(
            obtener_mock.call_args_list[1],
            call(Carton, idcarton=2),
        )


class FormatoCandidatosDesempateTests(SimpleTestCase):
    def test_interpreta_json_plano_generado_por_etapa_tres(self):
        valor = (
            '[{"idcarton":31,"codigocarton":"P20-C-31",'
            '"idjugador":41,"jugador":"juan123"},'
            '{"idcarton":32,"codigocarton":"P20-C-32",'
            '"idjugador":42,"jugador":"maria456"}]'
        )

        candidatos = normalizar_candidatos_desempate(
            parsear_candidatos_desempate(valor)
        )

        self.assertEqual([item["idjugador"] for item in candidatos], [41, 42])
        self.assertEqual(candidatos[0]["cartones"][0]["idcarton"], 31)
        self.assertIsNone(candidatos[0]["tiro_desempate"])

    def test_interpreta_ids_antiguos_separados_por_comas(self):
        candidatos = normalizar_candidatos_desempate(
            parsear_candidatos_desempate("41,42")
        )

        self.assertEqual([item["idjugador"] for item in candidatos], [41, 42])
        self.assertEqual(candidatos[0]["cartones"], [])

    def test_unifica_jugador_repetido_y_conserva_todos_sus_cartones(self):
        candidatos = normalizar_candidatos_desempate(
            [
                {
                    "idjugador": 41,
                    "jugador": "juan123",
                    "idcarton": 31,
                    "codigocarton": "P20-C-31",
                },
                {
                    "idjugador": 41,
                    "jugador": "juan123",
                    "idcarton": 33,
                    "codigocarton": "P20-C-33",
                },
            ]
        )

        self.assertEqual(len(candidatos), 1)
        self.assertEqual(
            [carton["idcarton"] for carton in candidatos[0]["cartones"]],
            [31, 33],
        )

    def test_tiro_se_conserva_al_serializar_y_volver_a_leer(self):
        candidatos = candidatos_desempate_prueba(tiro_uno=67)

        persistidos = serializar_candidatos_desempate(candidatos)
        recargados = normalizar_candidatos_desempate(
            parsear_candidatos_desempate(persistidos)
        )

        self.assertEqual(recargados[0]["tiro_desempate"], 67)
        self.assertEqual(recargados[0]["cartones"], candidatos[0]["cartones"])

    def test_formato_persistido_es_json_compacto_y_canonico(self):
        persistidos = serializar_candidatos_desempate(
            [candidatos_desempate_prueba(tiro_uno=67)[0]]
        )

        self.assertEqual(
            persistidos,
            '[{"idjugador":41,"jugador":"juan123","cartones":'
            '[{"idcarton":31,"codigocarton":"P20-C-31"}],'
            '"tiro_desempate":67}]',
        )

    def test_recargar_datos_no_modifica_tiros_persistidos(self):
        valor = serializar_candidatos_desempate(
            candidatos_desempate_prueba(tiro_uno=67)
        )
        partida = Partidabingo(
            idpartidabingo=20,
            estadopartida=ESTADO_PARTIDA_DESEMPATE,
            idbingadores=valor,
        )

        primera_carga = preparar_datos_desempate(partida)
        segunda_carga = preparar_datos_desempate(partida)

        self.assertEqual(
            primera_carga["candidatos_desempate"],
            segunda_carga["candidatos_desempate"],
        )
        self.assertEqual(partida.idbingadores, valor)

    def test_letra_bingo_del_tiro_se_calcula_correctamente(self):
        self.assertEqual(
            [formatear_bola_bingo(numero) for numero in (1, 16, 31, 46, 61, 75)],
            ["B-1", "I-16", "N-31", "G-46", "O-61", "O-75"],
        )

    def test_rechaza_tiros_repetidos_en_datos_persistidos(self):
        with self.assertRaises(DatosDesempateInvalidosError):
            normalizar_candidatos_desempate(
                candidatos_desempate_prueba(tiro_uno=50, tiro_dos=50)
            )


class ServicioDesempateTests(SimpleTestCase):
    def _partida(
        self,
        estado=ESTADO_PARTIDA_DESEMPATE,
        candidatos=None,
        pk=20,
    ):
        candidatos = candidatos or candidatos_desempate_prueba()
        return Partidabingo(
            idpartidabingo=pk,
            estadopartida=estado,
            idbingadores=serializar_candidatos_desempate(candidatos),
            idjugadorganador=None,
            bolamayordesempate=None,
            haydesempate=True,
            bolascantadas="[1,22,73]",
            ultimabola=73,
            horafin=None,
        )

    def _partida_bloqueada(self, partida, save_error=None):
        bloqueada = Partidabingo(
            idpartidabingo=partida.pk,
            estadopartida=partida.estadopartida,
            idbingadores=partida.idbingadores,
            idjugadorganador_id=partida.idjugadorganador_id,
            bolamayordesempate=partida.bolamayordesempate,
            haydesempate=partida.haydesempate,
            bolascantadas=partida.bolascantadas,
            ultimabola=partida.ultimabola,
            horafin=partida.horafin,
        )
        bloqueada.save = Mock(side_effect=save_error)
        return bloqueada

    def _sortear(self, partida, idjugador, balota=67, save_error=None):
        self.bloqueada = self._partida_bloqueada(partida, save_error=save_error)
        generador = Mock()
        generador.choice.return_value = balota
        with (
            patch(
                "apps.bingos.services.transaction.atomic",
                return_value=nullcontext(),
            ) as atomic_mock,
            patch(
                "apps.bingos.services.Partidabingo.objects.select_for_update"
            ) as lock_mock,
        ):
            lock_mock.return_value.get.return_value = self.bloqueada
            self.atomic_mock = atomic_mock
            self.lock_mock = lock_mock
            self.generador = generador
            return sortear_balota_desempate(
                partida,
                idjugador,
                generador_aleatorio=generador,
            )

    def _confirmar(self, partida, now=None, save_error=None):
        self.bloqueada = self._partida_bloqueada(partida, save_error=save_error)
        with (
            patch(
                "apps.bingos.services.transaction.atomic",
                return_value=nullcontext(),
            ) as atomic_mock,
            patch(
                "apps.bingos.services.Partidabingo.objects.select_for_update"
            ) as lock_mock,
        ):
            lock_mock.return_value.get.return_value = self.bloqueada
            self.atomic_mock = atomic_mock
            self.lock_mock = lock_mock
            return confirmar_y_finalizar_desempate(partida, now=now)

    def test_cada_jugador_recibe_solo_un_tiro(self):
        partida = self._partida(
            candidatos=candidatos_desempate_prueba(tiro_uno=67)
        )

        with self.assertRaises(DesempateError):
            self._sortear(partida, 41, balota=68)

        self.bloqueada.save.assert_not_called()
        self.generador.choice.assert_not_called()

    def test_tiro_generado_esta_entre_uno_y_setenta_y_cinco(self):
        partida = self._partida()

        resultado = self._sortear(partida, 41, balota=75)

        self.assertGreaterEqual(resultado["balota"], 1)
        self.assertLessEqual(resultado["balota"], 75)

    def test_dos_candidatos_no_reciben_la_misma_balota(self):
        partida = self._partida(
            candidatos=candidatos_desempate_prueba(tiro_uno=67)
        )

        resultado = self._sortear(partida, 42, balota=68)

        disponibles = self.generador.choice.call_args.args[0]
        self.assertNotIn(67, disponibles)
        self.assertEqual(resultado["balota"], 68)
        self.assertEqual(obtener_tiros_desempate(resultado["candidatos"]), [67, 68])

    def test_jugador_no_candidato_no_puede_sortear(self):
        partida = self._partida()

        with self.assertRaises(DesempateError):
            self._sortear(partida, 99)

        self.bloqueada.save.assert_not_called()

    def test_no_se_puede_sortear_fuera_de_desempate(self):
        for estado in (
            ESTADO_PARTIDA_PROGRAMADA,
            ESTADO_PARTIDA_EN_ESPERA,
            ESTADO_PARTIDA_PAUSADA,
            ESTADO_PARTIDA_CANCELADA,
        ):
            with self.subTest(estado=estado):
                partida = self._partida(estado=estado)
                with self.assertRaises(DesempateError):
                    sortear_balota_desempate(partida, 41)

    def test_no_se_puede_sortear_en_curso(self):
        partida = self._partida(estado=ESTADO_PARTIDA_EN_CURSO)

        with self.assertRaises(DesempateError):
            sortear_balota_desempate(partida, 41)

    def test_no_se_puede_sortear_en_finalizada(self):
        partida = self._partida(estado=ESTADO_PARTIDA_FINALIZADA)

        with self.assertRaises(DesempateError):
            sortear_balota_desempate(partida, 41)

    def test_primer_tiro_convierte_formato_etapa_tres_sin_perder_cartones(self):
        partida = self._partida()
        partida.idbingadores = (
            '[{"idcarton":31,"codigocarton":"P20-C-31",'
            '"idjugador":41,"jugador":"juan123"},'
            '{"idcarton":33,"codigocarton":"P20-C-33",'
            '"idjugador":41,"jugador":"juan123"},'
            '{"idcarton":32,"codigocarton":"P20-C-32",'
            '"idjugador":42,"jugador":"maria456"}]'
        )

        resultado = self._sortear(partida, 41, balota=67)
        persistidos = normalizar_candidatos_desempate(
            parsear_candidatos_desempate(partida.idbingadores)
        )

        self.assertEqual(resultado["balota"], 67)
        self.assertEqual(persistidos[0]["tiro_desempate"], 67)
        self.assertEqual(
            [carton["idcarton"] for carton in persistidos[0]["cartones"]],
            [31, 33],
        )

    def test_balotas_disponibles_excluyen_tiros_realizados(self):
        disponibles = obtener_balotas_disponibles_desempate(
            candidatos_desempate_prueba(tiro_uno=1, tiro_dos=75)
        )

        self.assertNotIn(1, disponibles)
        self.assertNotIn(75, disponibles)
        self.assertEqual(len(disponibles), 73)

    def test_no_confirma_antes_de_completar_todos_los_tiros(self):
        partida = self._partida(
            candidatos=candidatos_desempate_prueba(tiro_uno=67)
        )

        with self.assertRaises(DesempateIncompletoError):
            self._confirmar(partida)

        self.bloqueada.save.assert_not_called()

    def test_no_se_puede_confirmar_fuera_de_desempate(self):
        partida = self._partida(
            estado=ESTADO_PARTIDA_FINALIZADA,
            candidatos=candidatos_desempate_prueba(tiro_uno=67, tiro_dos=58),
        )

        with self.assertRaises(DesempateError):
            confirmar_y_finalizar_desempate(partida)

    def test_identifica_correctamente_el_numero_maximo(self):
        resultado = obtener_resultado_desempate(
            candidatos_desempate_prueba(tiro_uno=67, tiro_dos=58)
        )

        self.assertEqual(resultado["idjugador"], 41)
        self.assertEqual(resultado["balota"], 67)
        self.assertEqual(resultado["codigo"], "O-67")

    def test_confirmar_registra_al_jugador_ganador(self):
        partida = self._partida(
            candidatos=candidatos_desempate_prueba(tiro_uno=67, tiro_dos=58)
        )

        self._confirmar(partida)

        self.assertEqual(partida.idjugadorganador_id, 41)

    def test_confirmar_guarda_la_balota_mayor(self):
        partida = self._partida(
            candidatos=candidatos_desempate_prueba(tiro_uno=67, tiro_dos=58)
        )

        self._confirmar(partida)

        self.assertEqual(partida.bolamayordesempate, 67)

    def test_confirmar_cambia_la_partida_a_finalizada(self):
        partida = self._partida(
            candidatos=candidatos_desempate_prueba(tiro_uno=67, tiro_dos=58)
        )

        self._confirmar(partida)

        self.assertEqual(partida.estadopartida, ESTADO_PARTIDA_FINALIZADA)

    def test_confirmar_guarda_hora_fin(self):
        partida = self._partida(
            candidatos=candidatos_desempate_prueba(tiro_uno=67, tiro_dos=58)
        )
        now = timezone.now()

        self._confirmar(partida, now=now)

        self.assertEqual(partida.horafin, now)

    def test_confirmar_conserva_hay_desempate_verdadero(self):
        partida = self._partida(
            candidatos=candidatos_desempate_prueba(tiro_uno=67, tiro_dos=58)
        )

        self._confirmar(partida)

        self.assertTrue(partida.haydesempate)

    def test_confirmar_no_altera_bolas_normales_ni_ultima_bola(self):
        partida = self._partida(
            candidatos=candidatos_desempate_prueba(tiro_uno=67, tiro_dos=58)
        )
        bolas_antes = partida.bolascantadas
        ultima_antes = partida.ultimabola

        self._confirmar(partida)

        self.assertEqual(partida.bolascantadas, bolas_antes)
        self.assertEqual(partida.ultimabola, ultima_antes)
        self.assertNotIn("bolascantadas", self.bloqueada.save.call_args.kwargs["update_fields"])
        self.assertNotIn("ultimabola", self.bloqueada.save.call_args.kwargs["update_fields"])

    def test_operaciones_usan_transaccion_y_bloquean_partida(self):
        partida = self._partida()

        self._sortear(partida, 41, balota=67)

        self.atomic_mock.assert_called_once_with()
        self.lock_mock.assert_called_once_with()
        self.lock_mock.return_value.get.assert_called_once_with(pk=partida.pk)

    def test_error_al_guardar_no_sincroniza_cambios_parciales(self):
        partida = self._partida(
            candidatos=candidatos_desempate_prueba(tiro_uno=67, tiro_dos=58)
        )
        valores_antes = (
            partida.idjugadorganador_id,
            partida.bolamayordesempate,
            partida.estadopartida,
            partida.horafin,
            partida.idbingadores,
        )

        with self.assertRaises(DatabaseError):
            self._confirmar(
                partida,
                save_error=DatabaseError("fallo simulado"),
            )

        self.assertEqual(
            (
                partida.idjugadorganador_id,
                partida.bolamayordesempate,
                partida.estadopartida,
                partida.horafin,
                partida.idbingadores,
            ),
            valores_antes,
        )


class RutasDesempateTests(SimpleTestCase):
    def setUp(self):
        self.factory = RequestFactory()

    def test_usuario_no_administrativo_recibe_acceso_denegado(self):
        usuario = User(username="jugador", is_staff=False, is_superuser=False)
        solicitudes = (
            (sortear_desempate, self.factory.post("/desempate/41/sortear/"), (20, 41)),
            (confirmar_desempate, self.factory.post("/desempate/confirmar/"), (20,)),
            (desempate_operador, self.factory.get("/desempate/"), (20,)),
        )

        for vista, request, argumentos in solicitudes:
            with self.subTest(vista=vista.__name__):
                request.user = usuario
                with self.assertRaises(PermissionDenied):
                    vista(request, *argumentos)

    def test_get_no_registra_tiros_ni_finaliza(self):
        usuario = User(username="admin", is_staff=True)
        request_sortear = self.factory.get("/desempate/41/sortear/")
        request_sortear.user = usuario
        request_confirmar = self.factory.get("/desempate/confirmar/")
        request_confirmar.user = usuario

        with (
            patch("apps.bingos.views.sortear_balota_desempate") as sortear_mock,
            patch("apps.bingos.views.confirmar_y_finalizar_desempate") as confirmar_mock,
        ):
            response_sortear = sortear_desempate(request_sortear, 20, 41)
            response_confirmar = confirmar_desempate(request_confirmar, 20)

        self.assertEqual(response_sortear.status_code, 405)
        self.assertEqual(response_confirmar.status_code, 405)
        sortear_mock.assert_not_called()
        confirmar_mock.assert_not_called()


class ServiciosPublicosBingoTests(SimpleTestCase):
    def test_tablero_prepara_ultima_bola_y_totales(self):
        partida = partida_publica_prueba(bolas=[1, 27, 73], ultima=73)

        datos = preparar_datos_tablero_publico(partida)

        self.assertEqual(datos["ultima_bola_codigo"], "O-73")
        self.assertEqual(datos["total_bolas_extraidas"], 3)
        self.assertEqual(datos["total_bolas_faltantes"], 72)
        self.assertEqual(datos["patron_ganador_label"], "Cartón lleno")
        self.assertEqual(datos["total_requeridos_patron"], 24)

    def test_payload_websocket_incluye_patron_ganador_y_total_requerido(self):
        partida = partida_publica_prueba(bolas=[1, 61, 5], ultima=5)
        partida.patronganador = "cuatro_esquinas"

        payload = construir_payload_publico_partida(partida, "bola_extraida")

        self.assertEqual(payload["partida"]["patron_ganador"], "cuatro_esquinas")
        self.assertEqual(payload["partida"]["patron_ganador_label"], "Cuatro esquinas")
        self.assertEqual(payload["partida"]["total_requeridos_patron"], 4)

    def test_tablero_marca_las_bolas_extraidas(self):
        partida = partida_publica_prueba(bolas=[1, 27, 73])

        datos = preparar_datos_tablero_publico(partida)
        bolas = {
            bola["numero"]: bola["extraida"]
            for columna in datos["tablero_bingo"]
            for bola in columna["bolas"]
        }

        self.assertTrue(bolas[1])
        self.assertTrue(bolas[27])
        self.assertTrue(bolas[73])
        self.assertFalse(bolas[2])

    def test_partida_sin_bolas_se_prepara_de_forma_controlada(self):
        partida = partida_publica_prueba(bolas=[], ultima=0)

        datos = preparar_datos_tablero_publico(partida)

        self.assertIsNone(datos["ultima_bola_codigo"])
        self.assertEqual(datos["total_bolas_extraidas"], 0)
        self.assertEqual(datos["historial_bolas"], [])

    def test_partida_pausada_muestra_mensaje_publico(self):
        partida = partida_publica_prueba(estado=ESTADO_PARTIDA_PAUSADA)

        datos = preparar_datos_tablero_publico(partida)

        self.assertEqual(
            datos["mensaje_estado_publico"],
            "La partida está pausada temporalmente.",
        )
        self.assertNotIn("puede_sacar_bola", datos)

    def test_ganador_solo_se_expone_al_finalizar(self):
        ganador = Jugador(idjugador=41, aliasjugador="campeon_publico")
        en_curso = partida_publica_prueba(ganador=ganador)
        finalizada = partida_publica_prueba(
            estado=ESTADO_PARTIDA_FINALIZADA,
            ganador=ganador,
        )

        self.assertIsNone(preparar_datos_tablero_publico(en_curso)["ganador_publico"])
        self.assertEqual(
            preparar_datos_tablero_publico(finalizada)["ganador_publico"],
            "campeon_publico",
        )

    def test_resumen_publico_no_incluye_datos_de_desempate(self):
        partida = partida_publica_prueba(bolas=[12])
        partida.idbingadores = '[{"idjugador":99,"tiro_desempate":75}]'

        resumen = preparar_resumen_partida_publica(partida)

        self.assertEqual(resumen["total_bolas_extraidas"], 1)
        self.assertNotIn("idbingadores", resumen)
        self.assertNotIn("candidatos", resumen)

    def test_carton_prepara_matriz_cinco_por_cinco_y_libre(self):
        carton = carton_publico_prueba()

        datos = preparar_datos_carton_jugador(carton)

        self.assertEqual(len(datos["matriz_carton"]), 5)
        self.assertTrue(all(len(fila) == 5 for fila in datos["matriz_carton"]))
        self.assertTrue(datos["matriz_carton"][2][2]["libre"])
        self.assertEqual(datos["matriz_carton"][2][2]["valor"], CASILLA_LIBRE)
        self.assertEqual(datos["patron_ganador_label"], "Cartón lleno")
        self.assertEqual(datos["total_requeridos_patron"], 24)
        self.assertEqual(datos["numeros_marcados_patron"], 0)

    def test_carton_diferencia_numeros_marcados_y_pendientes(self):
        partida = partida_publica_prueba(bolas=[1, 16])
        carton = carton_publico_prueba(partida=partida)

        datos = preparar_datos_carton_jugador(carton)

        self.assertTrue(datos["matriz_carton"][0][0]["marcada"])
        self.assertTrue(datos["matriz_carton"][0][1]["marcada"])
        self.assertFalse(datos["matriz_carton"][0][2]["marcada"])

    def test_carton_cuatro_esquinas_prepara_progreso_y_faltantes_del_patron(self):
        partida = partida_publica_prueba(bolas=[1, 61, 5], ultima=5)
        partida.patronganador = "cuatro_esquinas"
        carton = carton_publico_prueba(partida=partida)

        datos = preparar_datos_carton_jugador(carton)

        self.assertEqual(datos["patron_ganador_label"], "Cuatro esquinas")
        self.assertEqual(datos["total_requeridos_patron"], 4)
        self.assertEqual(datos["numeros_marcados_patron"], 3)
        self.assertEqual(datos["progreso_patron"], 75)
        self.assertEqual(datos["numeros_faltantes_patron"], [65])
        self.assertEqual(datos["faltantes_patron_codigos"], ["O-65"])
        self.assertFalse(datos["patron_completo"])

    def test_contador_ignora_libre(self):
        matriz = matriz_carton_prueba()

        self.assertEqual(contar_numeros_marcados_carton(matriz, []), 0)
        self.assertEqual(contar_numeros_marcados_carton(matriz, [1, 16]), 2)
        self.assertEqual(
            contar_numeros_marcados_carton(matriz, obtener_numeros_carton(matriz)),
            24,
        )

    def test_matriz_invalida_no_se_considera_dato_publico_valido(self):
        carton = carton_publico_prueba()
        carton.matriznumeros = "[[1,2],[3,4]]"

        with self.assertRaises(MatrizCartonInvalidaError):
            preparar_datos_carton_jugador(carton)

    def test_carton_sin_partida_genera_error_controlado(self):
        carton = carton_publico_prueba()
        carton.idpartida = None

        with self.assertRaises(CartonPublicoError):
            preparar_datos_carton_jugador(carton)


class VistasPublicasBingoTests(SimpleTestCase):
    def setUp(self):
        self.factory = RequestFactory()

    def _request(self, path, usuario=None, method="get", data=None):
        request = getattr(self.factory, method)(path, data=data or {})
        request.user = usuario or AnonymousUser()
        return request

    def test_sala_publica_abre_sin_iniciar_sesion(self):
        request = self._request("/juego/")
        partida = partida_publica_prueba()
        consulta = Mock()
        consulta.order_by.return_value = [partida]

        with (
            patch(
                "apps.bingos.views.Partidabingo.objects.select_related",
                return_value=consulta,
            ),
            patch(
                "apps.bingos.views.render",
                return_value=HttpResponse("Sala pública"),
            ) as render_mock,
        ):
            response = sala_juego_publica(request)

        self.assertEqual(response.status_code, 200)
        contexto = render_mock.call_args.args[2]
        self.assertEqual(contexto["partidas_publicas"][0]["partida"], partida)

    def test_tablero_publico_abre_sin_iniciar_sesion(self):
        request = self._request("/juego/partidas/20/tablero/")
        partida = partida_publica_prueba()

        with (
            patch("apps.bingos.views.get_object_or_404", return_value=partida),
            patch(
                "apps.bingos.views.render",
                return_value=HttpResponse("Tablero público"),
            ),
        ):
            response = tablero_publico(request, 20)

        self.assertEqual(response.status_code, 200)

    def test_usuario_no_administrativo_puede_abrir_rutas_publicas(self):
        usuario = User(username="publico", is_staff=False, is_superuser=False)
        request = self._request(
            "/juego/partidas/20/tablero/",
            usuario=usuario,
        )
        partida = partida_publica_prueba()

        with (
            patch("apps.bingos.views.get_object_or_404", return_value=partida),
            patch(
                "apps.bingos.views.render",
                return_value=HttpResponse("Tablero público"),
            ),
        ):
            response = tablero_publico(request, 20)

        self.assertEqual(response.status_code, 200)

    def test_post_al_tablero_no_puede_modificar_partida(self):
        request = self._request(
            "/juego/partidas/20/tablero/",
            method="post",
        )

        with patch("apps.bingos.views.get_object_or_404") as obtener_mock:
            response = tablero_publico(request, 20)

        self.assertEqual(response.status_code, 405)
        obtener_mock.assert_not_called()

    def test_get_publico_no_cambia_datos_de_partida(self):
        request = self._request("/juego/partidas/20/tablero/")
        partida = partida_publica_prueba(bolas=[1, 22], ultima=22)
        partida.idbingadores = "evidencia-privada"
        partida.save = Mock()
        antes = (
            partida.estadopartida,
            partida.bolascantadas,
            partida.ultimabola,
            partida.idjugadorganador_id,
            partida.idbingadores,
        )

        with (
            patch("apps.bingos.views.get_object_or_404", return_value=partida),
            patch(
                "apps.bingos.views.render",
                return_value=HttpResponse("Tablero público"),
            ),
        ):
            tablero_publico(request, 20)

        self.assertEqual(
            (
                partida.estadopartida,
                partida.bolascantadas,
                partida.ultimabola,
                partida.idjugadorganador_id,
                partida.idbingadores,
            ),
            antes,
        )
        partida.save.assert_not_called()

    def test_codigo_valido_muestra_el_carton_correcto(self):
        request = self._request("/juego/cartones/PUBLICO-001/")
        carton = carton_publico_prueba()
        consulta = Mock()
        consulta.filter.return_value.first.return_value = carton

        with (
            patch(
                "apps.bingos.views.Carton.objects.select_related",
                return_value=consulta,
            ),
            patch(
                "apps.bingos.views.render",
                return_value=HttpResponse("Mi cartón"),
            ) as render_mock,
        ):
            response = carton_publico(request, carton.codigocarton)

        self.assertEqual(response.status_code, 200)
        self.assertIs(render_mock.call_args.args[2]["carton"], carton)
        consulta.filter.assert_called_once_with(codigocarton="PUBLICO-001")

    def test_formulario_valido_redirige_al_carton_sin_modificarlo(self):
        request = self._request(
            "/juego/cartones/acceder/",
            method="post",
            data={"codigocarton": "PUBLICO-001"},
        )

        with patch("apps.bingos.views.Carton.objects.filter") as filtrar_mock:
            filtrar_mock.return_value.exists.return_value = True
            response = acceder_carton_publico(request)

        self.assertEqual(response.status_code, 302)
        self.assertIn("/juego/cartones/PUBLICO-001/", response["Location"])
        filtrar_mock.return_value.update.assert_not_called()

    def test_codigo_invalido_muestra_mensaje_sin_detalles_internos(self):
        request = self._request(
            "/juego/cartones/acceder/",
            method="post",
            data={"codigocarton": "NO-EXISTE"},
        )

        with (
            patch("apps.bingos.views.Carton.objects.filter") as filtrar_mock,
            patch(
                "apps.bingos.views.render",
                return_value=HttpResponse("Código no encontrado"),
            ) as render_mock,
        ):
            filtrar_mock.return_value.exists.return_value = False
            response = acceder_carton_publico(request)

        self.assertEqual(response.status_code, 200)
        errores = str(render_mock.call_args.args[2]["form"].errors)
        self.assertIn("No encontramos un cartón", errores)
        self.assertNotIn("SELECT", errores)
        self.assertNotIn("carton.idcarton", errores)

    def test_url_de_codigo_invalido_devuelve_404_controlado(self):
        request = self._request("/juego/cartones/NO-EXISTE/")
        consulta = Mock()
        consulta.filter.return_value.first.return_value = None

        def render_controlado(_request, _template, _context, status=200):
            return HttpResponse("No encontramos un cartón", status=status)

        with (
            patch(
                "apps.bingos.views.Carton.objects.select_related",
                return_value=consulta,
            ),
            patch("apps.bingos.views.render", side_effect=render_controlado),
        ):
            response = carton_publico(request, "NO-EXISTE")

        self.assertEqual(response.status_code, 404)
        self.assertContains(response, "No encontramos un cartón", status_code=404)

    def test_matriz_invalida_no_rompe_vista_y_deja_evidencia(self):
        request = self._request("/juego/cartones/PUBLICO-001/")
        carton = carton_publico_prueba()
        carton.matriznumeros = "[[1,2],[3,4]]"
        consulta = Mock()
        consulta.filter.return_value.first.return_value = carton

        with (
            patch(
                "apps.bingos.views.Carton.objects.select_related",
                return_value=consulta,
            ),
            patch(
                "apps.bingos.views.render",
                return_value=HttpResponse("Error controlado"),
            ) as render_mock,
            self.assertLogs("apps.bingos.views", level="WARNING"),
        ):
            response = carton_publico(request, carton.codigocarton)

        self.assertEqual(response.status_code, 200)
        self.assertIn("no está disponible", render_mock.call_args.args[2]["error_carton"])

    def test_rutas_administrativas_conservan_restriccion(self):
        request = self._request(
            "/partidas/",
            usuario=User(username="publico", is_staff=False),
        )

        with self.assertRaises(PermissionDenied):
            partidas_lista(request)


@override_settings(
    STORAGES={
        "default": {
            "BACKEND": "django.core.files.storage.FileSystemStorage",
        },
        "staticfiles": {
            "BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage",
        },
    }
)
class PlantillasPublicasBingoTests(SimpleTestCase):
    def setUp(self):
        self.factory = RequestFactory()

    def _request(self, path):
        request = self.factory.get(path)
        request.user = AnonymousUser()
        return request

    def test_sala_no_expone_cartones_jugadores_ni_candidatos(self):
        partida = partida_publica_prueba()
        partida.idbingadores = (
            '[{"jugador":"jugador_privado","codigocarton":'
            '"CARTON-PRIVADO","tiro_desempate":75}]'
        )

        html = render_to_string(
            "bingos/sala_juego_publica.html",
            {
                "partidas_publicas": [
                    preparar_resumen_partida_publica(partida)
                ]
            },
            request=self._request("/juego/"),
        )

        self.assertNotIn("jugador_privado", html)
        self.assertNotIn("CARTON-PRIVADO", html)
        self.assertNotIn("tiro_desempate", html)

    def test_tablero_renderiza_ultima_bola_conteo_y_resaltados(self):
        partida = partida_publica_prueba(bolas=[1, 27, 73], ultima=73)
        contexto = {
            "partida": partida,
            **preparar_datos_tablero_publico(partida),
        }

        html = render_to_string(
            "bingos/tablero_publico.html",
            contexto,
            request=self._request("/juego/partidas/20/tablero/"),
        )

        self.assertIn("O-73", html)
        self.assertIn("3", html)
        self.assertIn("<dt>Extraídas</dt>", html)
        self.assertIn("<dt>Restantes</dt>", html)
        self.assertIn("Patrón de esta ronda", html)
        self.assertIn("Cartón lleno", html)
        self.assertIn("Casillas requeridas", html)
        self.assertIn("data-realtime-patron-label", html)
        self.assertIn("data-realtime-patron-total", html)
        self.assertEqual(html.count('class="public-bingo-column col"'), 5)
        self.assertIn("public-bingo-ball is-drawn", html)
        self.assertIn("public-ball-check", html)
        self.assertIn("Extraída", html)
        self.assertEqual(html.count("public-history-ball"), 3)
        self.assertIn("B-1", html)
        self.assertIn("I-27", html)
        self.assertIn("O-73", html)
        self.assertNotIn("Sacar siguiente bola", html)
        self.assertNotIn("Validar cartón", html)

    def test_tablero_cuatro_esquinas_muestra_patron_legible(self):
        partida = partida_publica_prueba(bolas=[1, 61, 5, 65], ultima=65)
        partida.patronganador = "cuatro_esquinas"
        html = render_to_string(
            "bingos/tablero_publico.html",
            {"partida": partida, **preparar_datos_tablero_publico(partida)},
            request=self._request("/juego/partidas/20/tablero/"),
        )

        self.assertIn("Patrón de esta ronda", html)
        self.assertIn("Cuatro esquinas", html)

    def test_tablero_sin_bolas_muestra_mensaje_controlado(self):
        partida = partida_publica_prueba(bolas=[], ultima=0)

        html = render_to_string(
            "bingos/tablero_publico.html",
            {"partida": partida, **preparar_datos_tablero_publico(partida)},
            request=self._request("/juego/partidas/20/tablero/"),
        )

        self.assertIn("Aún no se han extraído bolas.", html)

    def test_tablero_pausado_no_muestra_acciones_administrativas(self):
        partida = partida_publica_prueba(estado=ESTADO_PARTIDA_PAUSADA)

        html = render_to_string(
            "bingos/tablero_publico.html",
            {"partida": partida, **preparar_datos_tablero_publico(partida)},
            request=self._request("/juego/partidas/20/tablero/"),
        )

        self.assertIn("La partida está pausada temporalmente.", html)
        self.assertNotIn("Reanudar partida", html)
        self.assertNotIn("Pausar partida", html)

    def test_finalizada_muestra_ganador_solo_si_existe(self):
        ganador = Jugador(idjugador=41, aliasjugador="campeon_publico")
        con_ganador = partida_publica_prueba(
            estado=ESTADO_PARTIDA_FINALIZADA,
            ganador=ganador,
        )
        sin_ganador = partida_publica_prueba(
            estado=ESTADO_PARTIDA_FINALIZADA,
            ganador=None,
        )

        html_con = render_to_string(
            "bingos/tablero_publico.html",
            {"partida": con_ganador, **preparar_datos_tablero_publico(con_ganador)},
            request=self._request("/juego/partidas/20/tablero/"),
        )
        html_sin = render_to_string(
            "bingos/tablero_publico.html",
            {"partida": sin_ganador, **preparar_datos_tablero_publico(sin_ganador)},
            request=self._request("/juego/partidas/20/tablero/"),
        )

        self.assertIn("Ganador: campeon_publico", html_con)
        self.assertNotIn("campeon_publico", html_sin)

    def test_carton_renderiza_matriz_libre_marcados_pendientes_y_contador(self):
        partida = partida_publica_prueba(bolas=[1, 16])
        carton = carton_publico_prueba(partida=partida)
        datos = preparar_datos_carton_jugador(carton)

        html = render_to_string(
            "bingos/carton_publico.html",
            {
                "carton": carton,
                "partida": partida,
                "error_carton": None,
                **datos,
            },
            request=self._request("/juego/cartones/PUBLICO-001/"),
        )

        self.assertEqual(html.count("<tr>"), 6)
        self.assertIn("<dt>Bingo</dt>", html)
        self.assertIn("<dt>Ronda</dt>", html)
        self.assertIn("<dt>Estado</dt>", html)
        self.assertIn("Patrón de esta ronda", html)
        self.assertIn("Cartón lleno", html)
        self.assertIn("<dt>Progreso del cartón</dt>", html)
        self.assertIn("data-realtime-patron-label", html)
        self.assertIn("data-realtime-patron-progreso", html)
        self.assertIn("data-realtime-patron-completo", html)
        self.assertIn("data-realtime-patron-faltantes", html)
        self.assertIn("LIBRE", html)
        self.assertIn("public-card-cell--marked", html)
        self.assertIn("public-card-cell--pending", html)
        self.assertIn("public-card-cell--free", html)
        self.assertIn("✓ Marcado", html)
        self.assertIn("○ Pendiente", html)
        self.assertIn("★ Automática", html)
        self.assertIn("2 de 24 números marcados.", html)
        self.assertIn("La casilla LIBRE se marca automáticamente.", html)

    def test_carton_cuatro_esquinas_muestra_progreso_y_mensaje_de_patron(self):
        partida = partida_publica_prueba(bolas=[1, 61, 5, 65], ultima=65)
        partida.patronganador = "cuatro_esquinas"
        carton = carton_publico_prueba(partida=partida)
        datos = preparar_datos_carton_jugador(carton)

        html = render_to_string(
            "bingos/carton_publico.html",
            {
                "carton": carton,
                "partida": partida,
                "error_carton": None,
                **datos,
            },
            request=self._request("/juego/cartones/PUBLICO-001/"),
        )

        self.assertIn("Patrón de esta ronda", html)
        self.assertIn("Cuatro esquinas", html)
        self.assertIn("4 de 4 números marcados", html)
        self.assertIn("¡Patrón completado! Espera la validación del operador.", html)
        self.assertNotIn("LIBRE", datos["numeros_faltantes_patron"])
        self.assertEqual(datos["total_requeridos_patron"], 4)
        self.assertIn("data-realtime-patron-completo", html)
        self.assertIn("data-realtime-patron-faltantes", html)

    def test_mi_carton_detalle_muestra_patron_y_progreso_real(self):
        partida = partida_publica_prueba(bolas=[1, 61, 5], ultima=5)
        partida.patronganador = "cuatro_esquinas"
        carton = carton_publico_prueba(partida=partida)
        html = render_to_string(
            "bingos/mi_carton_detalle.html",
            {
                "partida": partida,
                "carton": carton,
                "jugador": carton.idjugador,
                "error_carton": None,
                **preparar_datos_carton_jugador(carton),
            },
            request=self._request("/mis-cartones/PUBLICO-001/"),
        )

        self.assertIn("Patrón de esta ronda", html)
        self.assertIn("Cuatro esquinas", html)
        self.assertIn("3 de 4 números marcados", html)
        self.assertIn("Faltan:", html)
        self.assertIn("data-realtime-patron-label", html)
        self.assertIn("data-realtime-patron-progreso", html)
        self.assertIn("data-realtime-patron-faltantes", html)

    def test_carton_no_expone_precio_otros_cartones_ni_otros_jugadores(self):
        partida = partida_publica_prueba(bolas=[1])
        carton = carton_publico_prueba(partida=partida)

        html = render_to_string(
            "bingos/carton_publico.html",
            {
                "carton": carton,
                "partida": partida,
                "error_carton": None,
                **preparar_datos_carton_jugador(carton),
            },
            request=self._request("/juego/cartones/PUBLICO-001/"),
        )

        self.assertNotIn("9876.54", html)
        self.assertNotIn("OTRO-CARTON-PRIVADO", html)
        self.assertNotIn("otro_jugador_privado", html)
        self.assertNotIn("Editar", html)
        self.assertNotIn("Validar cartón", html)


@override_settings(
    STORAGES={
        "default": {
            "BACKEND": "django.core.files.storage.FileSystemStorage",
        },
        "staticfiles": {
            "BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage",
        },
    }
)
class PlantillasAdministrativasPartidasTests(SimpleTestCase):
    def setUp(self):
        self.factory = RequestFactory()

    def _request(self, path):
        request = self.factory.get(path)
        request.user = User(username="admin", is_staff=True)
        return request

    def test_lista_partidas_muestra_patron_ganador_legible(self):
        bingo = Bingo(idbingo=7, titulobingo="Bingo administrativo")
        partida = Partidabingo(
            idpartidabingo=21,
            idbingo=bingo,
            nombreronda="Ronda 21",
            valorefectivo=Decimal("25.00"),
            premiomaterial="",
            estadopartida=ESTADO_PARTIDA_PROGRAMADA,
            patronganador="cuatro_esquinas",
            bolascantadas="[]",
            ultimabola=0,
            haydesempate=False,
            horainicio=timezone.now(),
        )

        html = render_to_string(
            "bingos/partidas_lista.html",
            {"page_obj": [partida], "busqueda": "", "total": 1},
            request=self._request("/partidas/"),
        )

        self.assertIn("Patrón ganador", html)
        self.assertIn("Cuatro esquinas", html)


class EstadoPartidaTests(SimpleTestCase):
    def test_en_juego_se_normaliza_a_en_curso(self):
        self.assertEqual(normalizar_estado_partida("En Juego"), ESTADO_PARTIDA_EN_CURSO)
        self.assertEqual(estado_partida_mostrar("En Juego"), ESTADO_PARTIDA_EN_CURSO)

    def test_verificando_se_normaliza_a_en_espera(self):
        self.assertEqual(normalizar_estado_partida("Verificando"), ESTADO_PARTIDA_EN_ESPERA)
        self.assertEqual(estado_partida_mostrar("Verificando"), ESTADO_PARTIDA_EN_ESPERA)

    def test_iniciar_partida_programada_cambia_a_en_curso(self):
        partida = Partidabingo(estadopartida=ESTADO_PARTIDA_PROGRAMADA)

        update_fields = aplicar_accion_consola(partida, "iniciar")

        self.assertEqual(partida.estadopartida, ESTADO_PARTIDA_EN_CURSO)
        self.assertEqual(update_fields, ["estadopartida"])

    def test_pausar_partida_programada_no_es_valido(self):
        partida = Partidabingo(estadopartida=ESTADO_PARTIDA_PROGRAMADA)

        with self.assertRaises(EstadoPartidaError):
            aplicar_accion_consola(partida, "pausar")

    def test_finalizar_partida_en_curso_asigna_hora_fin(self):
        now = timezone.now()
        partida = Partidabingo(estadopartida=ESTADO_PARTIDA_EN_CURSO)

        update_fields = aplicar_accion_consola(partida, "finalizar", now=now)

        self.assertEqual(partida.estadopartida, ESTADO_PARTIDA_FINALIZADA)
        self.assertEqual(partida.horafin, now)
        self.assertEqual(update_fields, ["estadopartida", "horafin"])

    def test_reanudar_partida_pausada_cambia_a_en_curso(self):
        partida = Partidabingo(estadopartida=ESTADO_PARTIDA_PAUSADA)

        aplicar_accion_consola(partida, "reanudar")

        self.assertEqual(partida.estadopartida, ESTADO_PARTIDA_EN_CURSO)

    def test_en_juego_puede_pasar_a_pausada(self):
        partida = Partidabingo(estadopartida="En Juego")

        aplicar_accion_consola(partida, "pausar")

        self.assertEqual(partida.estadopartida, ESTADO_PARTIDA_PAUSADA)

    def test_transicion_invalida_no_modifica_estado(self):
        partida = Partidabingo(estadopartida=ESTADO_PARTIDA_PROGRAMADA)

        with self.assertRaises(EstadoPartidaError):
            aplicar_accion_consola(partida, "reanudar")

        self.assertEqual(partida.estadopartida, ESTADO_PARTIDA_PROGRAMADA)

    def test_acciones_disponibles_por_estado(self):
        self.assertEqual(
            acciones_disponibles_consola(Partidabingo(estadopartida=ESTADO_PARTIDA_PROGRAMADA)),
            {"iniciar"},
        )
        self.assertEqual(
            acciones_disponibles_consola(Partidabingo(estadopartida=ESTADO_PARTIDA_EN_ESPERA)),
            {"iniciar"},
        )
        self.assertEqual(
            acciones_disponibles_consola(Partidabingo(estadopartida="En Juego")),
            {"pausar", "finalizar"},
        )
        self.assertEqual(
            acciones_disponibles_consola(Partidabingo(estadopartida=ESTADO_PARTIDA_PAUSADA)),
            {"reanudar", "finalizar"},
        )
        self.assertEqual(
            acciones_disponibles_consola(Partidabingo(estadopartida=ESTADO_PARTIDA_DESEMPATE)),
            set(),
        )
        self.assertEqual(
            acciones_disponibles_consola(Partidabingo(estadopartida=ESTADO_PARTIDA_FINALIZADA)),
            set(),
        )
        self.assertEqual(
            acciones_disponibles_consola(Partidabingo(estadopartida=ESTADO_PARTIDA_CANCELADA)),
            set(),
        )


class PermisosBingoTests(SimpleTestCase):
    def setUp(self):
        self.factory = RequestFactory()

    def test_usuario_no_staff_no_accede_a_lista_de_partidas(self):
        request = self.factory.get("/partidas/")
        request.user = User(username="jugador", is_staff=False, is_superuser=False)

        with self.assertRaises(PermissionDenied):
            partidas_lista(request)

    def test_usuario_anonimo_es_redirigido_a_login(self):
        request = self.factory.get("/partidas/")
        request.user = AnonymousUser()

        response = partidas_lista(request)

        self.assertEqual(response.status_code, 302)
        self.assertIn("/login/", response["Location"])


class ConsolaOperadorTests(SimpleTestCase):
    def setUp(self):
        self.factory = RequestFactory()

    def _request_with_messages(self, accion):
        request = self.factory.post("/partidas/1/consola/", {"accion": accion})
        request.user = User(username="admin", is_staff=True)
        request.session = {}
        request._messages = FallbackStorage(request)
        return request

    def test_consola_muestra_exito_despues_de_pausar(self):
        request = self._request_with_messages("pausar")
        partida = Partidabingo(idpartidabingo=1, estadopartida="En Juego")
        partida.save = Mock()

        with (
            patch("apps.bingos.views._base_datos_permite_estado_partida", return_value=True),
            patch("apps.bingos.views.transaction.atomic", return_value=nullcontext()),
            patch("apps.bingos.views.programar_publicacion_partida"),
        ):
            resultado = _procesar_accion_consola(request, partida, "pausar")

        self.assertTrue(resultado)
        self.assertEqual(partida.estadopartida, ESTADO_PARTIDA_PAUSADA)
        partida.save.assert_called_once_with(update_fields=["estadopartida"])
        self.assertIn(
            "Pausar partida realizada correctamente.",
            [message.message for message in get_messages(request)],
        )

    def test_consola_muestra_exito_despues_de_reanudar(self):
        request = self._request_with_messages("reanudar")
        partida = Partidabingo(idpartidabingo=1, estadopartida=ESTADO_PARTIDA_PAUSADA)
        partida.save = Mock()

        with (
            patch("apps.bingos.views._base_datos_permite_estado_partida", return_value=True),
            patch("apps.bingos.views.transaction.atomic", return_value=nullcontext()),
            patch("apps.bingos.views.programar_publicacion_partida"),
        ):
            resultado = _procesar_accion_consola(request, partida, "reanudar")

        self.assertTrue(resultado)
        self.assertEqual(partida.estadopartida, ESTADO_PARTIDA_EN_CURSO)
        partida.save.assert_called_once_with(update_fields=["estadopartida"])
        self.assertIn(
            "Reanudar partida realizada correctamente.",
            [message.message for message in get_messages(request)],
        )

    def test_consola_bloquea_estado_si_check_no_esta_actualizada(self):
        request = self._request_with_messages("pausar")
        partida = Partidabingo(idpartidabingo=1, estadopartida="En Juego")
        partida.save = Mock()

        with patch("apps.bingos.views._base_datos_permite_estado_partida", return_value=False):
            resultado = _procesar_accion_consola(request, partida, "pausar")

        self.assertFalse(resultado)
        self.assertEqual(partida.estadopartida, "En Juego")
        partida.save.assert_not_called()
        self.assertTrue(
            any(
                "Esta acción requiere actualizar la restricción de estados en PostgreSQL."
                in message.message
                for message in get_messages(request)
            )
        )


@override_settings(
    ALLOWED_HOSTS=["localhost", "testserver"],
    CHANNEL_LAYERS={
        "default": {
            "BACKEND": "channels.layers.InMemoryChannelLayer",
        }
    },
)
class WebSocketPublicoBingoTests(SimpleTestCase):
    class CapaCanalesPrueba:
        def __init__(self):
            self.grupos_agregados = []
            self.grupos_descartados = []

        async def group_add(self, group_name, channel_name):
            self.grupos_agregados.append((group_name, channel_name))

        async def group_discard(self, group_name, channel_name):
            self.grupos_descartados.append((group_name, channel_name))

    def test_aplicacion_asgi_carga_correctamente(self):
        from config.asgi import application

        self.assertIsNotNone(application)

    def test_ruta_websocket_publica_existe(self):
        from apps.bingos.routing import websocket_urlpatterns

        rutas_partida = [
            ruta
            for ruta in websocket_urlpatterns
            if getattr(ruta, "name", None) == "ws_partida_publica"
        ]

        self.assertEqual(len(rutas_partida), 1)
        self.assertEqual(
            rutas_partida[0].callback.consumer_class,
            PartidaPublicaConsumer,
        )

    def test_consumidor_acepta_partida_existente(self):
        async def partida_existe(_consumer, _idpartidabingo):
            return True

        async def escenario():
            consumidor = PartidaPublicaConsumer()
            consumidor.scope = {
                "url_route": {"kwargs": {"idpartidabingo": "20"}}
            }
            consumidor.channel_layer = self.CapaCanalesPrueba()
            consumidor.channel_name = "canal-prueba"
            consumidor.base_send = AsyncMock()

            with patch.object(
                PartidaPublicaConsumer,
                "_partida_existe",
                new=partida_existe,
            ):
                await consumidor.connect()

            return consumidor

        consumidor = asyncio.run(escenario())
        consumidor.base_send.assert_awaited_once_with(
            {"type": "websocket.accept", "subprotocol": None}
        )
        self.assertEqual(
            consumidor.channel_layer.grupos_agregados,
            [(nombre_grupo_partida(20), "canal-prueba")],
        )

    def test_partida_inexistente_se_cierra_de_forma_controlada(self):
        async def partida_existe(_consumer, _idpartidabingo):
            return False

        async def escenario():
            consumidor = PartidaPublicaConsumer()
            consumidor.scope = {
                "url_route": {"kwargs": {"idpartidabingo": "9999"}}
            }
            consumidor.channel_layer = self.CapaCanalesPrueba()
            consumidor.channel_name = "canal-prueba"
            consumidor.base_send = AsyncMock()

            with patch.object(
                PartidaPublicaConsumer,
                "_partida_existe",
                new=partida_existe,
            ):
                await consumidor.connect()

            return consumidor

        consumidor = asyncio.run(escenario())
        consumidor.base_send.assert_awaited_once_with(
            {"type": "websocket.close", "code": 4404}
        )
        self.assertEqual(consumidor.channel_layer.grupos_agregados, [])

    def test_mensajes_del_cliente_se_ignoran_sin_acciones_admin(self):
        consumidor = PartidaPublicaConsumer()
        with patch("apps.bingos.models.Partidabingo.save") as guardar_mock:
            resultado = asyncio.run(
                consumidor.receive(
                    text_data='{"accion": "sacar_bola", "estadopartida": "Finalizada"}'
                )
            )

        self.assertIsNone(resultado)
        guardar_mock.assert_not_called()

    def test_evento_de_bola_llega_al_cliente_con_payload_publico(self):
        from config.asgi import application

        async def partida_existe(_consumer, _idpartidabingo):
            return True

        partida = partida_publica_prueba(bolas=[2, 23], ultima=23)
        payload = construir_payload_publico_partida(
            partida,
            "bola_extraida",
        )

        async def escenario():
            with patch.object(
                PartidaPublicaConsumer,
                "_partida_existe",
                new=partida_existe,
            ), patch("apps.bingos.models.Partidabingo.save") as guardar_mock:
                comunicador = WebsocketCommunicator(
                    application,
                    "/ws/juego/partidas/20/",
                    headers=[(b"origin", b"http://localhost")],
                )
                await get_channel_layer().flush()
                try:
                    conectado, _subprotocolo = await comunicador.connect(
                        timeout=0.5
                    )
                    await comunicador.send_json_to(
                        {
                            "accion": "sacar_bola",
                            "estadopartida": "Finalizada",
                        }
                    )
                    sin_respuesta = await comunicador.receive_nothing(
                        timeout=0.05
                    )
                    guardar_mock.assert_not_called()
                    await get_channel_layer().group_send(
                        nombre_grupo_partida(20),
                        {
                            "type": "partida.actualizada",
                            "payload": payload,
                        },
                    )
                    await asyncio.sleep(0.05)
                    recibido = await comunicador.receive_json_from(
                        timeout=0.5
                    )
                finally:
                    comunicador.stop(exceptions=False)
                return conectado, sin_respuesta, recibido

        conectado, sin_respuesta, recibido = asyncio.run(escenario())
        self.assertTrue(conectado)
        self.assertTrue(sin_respuesta)
        self.assertEqual(recibido, payload)
        self.assertEqual(recibido["partida"]["ultima_bola"]["codigo"], "I-23")
        self.assertEqual(recibido["partida"]["estado"], ESTADO_PARTIDA_EN_CURSO)
        self.assertEqual(recibido["partida"]["total_extraidas"], 2)
        self.assertEqual(recibido["partida"]["cantidad_extraida"], 2)
        self.assertEqual(recibido["partida"]["restantes"], 73)

    def test_consumidor_sin_capa_de_canales_cierra_controladamente(self):
        async def partida_existe(_consumer, _idpartidabingo):
            return True

        async def escenario():
            consumidor = PartidaPublicaConsumer()
            consumidor.scope = {
                "url_route": {"kwargs": {"idpartidabingo": "20"}}
            }
            consumidor.channel_layer = None
            consumidor.channel_name = "canal-prueba"
            consumidor.base_send = AsyncMock()

            with patch.object(
                PartidaPublicaConsumer,
                "_partida_existe",
                new=partida_existe,
            ):
                await consumidor.connect()

            return consumidor

        consumidor = asyncio.run(escenario())
        consumidor.base_send.assert_awaited_once_with(
            {"type": "websocket.close", "code": 1011}
        )

    def test_desconexion_descarta_el_grupo_publico(self):
        async def escenario():
            consumidor = PartidaPublicaConsumer()
            consumidor.group_name = nombre_grupo_partida(20)
            consumidor.channel_layer = self.CapaCanalesPrueba()
            consumidor.channel_name = "canal-prueba"

            await consumidor.disconnect(1000)
            return consumidor

        consumidor = asyncio.run(escenario())
        self.assertEqual(
            consumidor.channel_layer.grupos_descartados,
            [(nombre_grupo_partida(20), "canal-prueba")],
        )

    def test_payload_websocket_no_incluye_privados_y_contiene_resumen_publico(self):
        partida = partida_publica_prueba(bolas=[2, 23], ultima=23)
        partida.idbingadores = (
            '[{"idjugador":41,"tiro_desempate":75,'
            '"codigocarton":"PRIVADO-001"}]'
        )

        payload = construir_payload_publico_partida(
            partida,
            "bola_extraida",
        )
        partida_payload = payload["partida"]
        serializado = json.dumps(payload)

        self.assertEqual(partida_payload["estado"], ESTADO_PARTIDA_EN_CURSO)
        self.assertEqual(partida_payload["ultima_bola"]["codigo"], "I-23")
        self.assertEqual(partida_payload["bolas_extraidas"], [2, 23])
        self.assertEqual(partida_payload["total_extraidas"], 2)
        self.assertEqual(partida_payload["cantidad_extraida"], 2)
        self.assertEqual(partida_payload["restantes"], 73)
        self.assertIsNone(partida_payload["ganador"])
        self.assertNotIn("idbingadores", serializado)
        self.assertNotIn("tiro_desempate", serializado)
        self.assertNotIn("codigocarton", serializado)
        self.assertNotIn("preciopagado", serializado)
        self.assertNotIn("PRIVADO-001", serializado)


class PayloadPublicoTiempoRealTests(SimpleTestCase):
    def test_payload_excluye_datos_privados(self):
        partida = partida_publica_prueba(bolas=[12])
        partida.idbingadores = (
            '[{"idjugador":41,"tiro_desempate":75,'
            '"codigocarton":"PRIVADO-001"}]'
        )
        payload = construir_payload_publico_partida(partida, "actualizacion")
        serializado = json.dumps(payload)

        self.assertNotIn("idbingadores", serializado)
        self.assertNotIn("tiro_desempate", serializado)
        self.assertNotIn("codigocarton", serializado)
        self.assertNotIn("preciopagado", serializado)
        self.assertNotIn("PRIVADO-001", serializado)

    def test_payload_refleja_pausa_reanudacion_y_desempate(self):
        partida = partida_publica_prueba(estado=ESTADO_PARTIDA_PAUSADA)
        pausada = construir_payload_publico_partida(partida, "partida_pausada")
        partida.estadopartida = ESTADO_PARTIDA_EN_CURSO
        reanudada = construir_payload_publico_partida(
            partida,
            "partida_reanudada",
        )
        partida.estadopartida = ESTADO_PARTIDA_DESEMPATE
        partida.idbingadores = '[{"idjugador":41,"tiro_desempate":70}]'
        desempate = construir_payload_publico_partida(
            partida,
            "desempate_detectado",
        )

        self.assertEqual(pausada["partida"]["estado"], ESTADO_PARTIDA_PAUSADA)
        self.assertEqual(reanudada["partida"]["estado"], ESTADO_PARTIDA_EN_CURSO)
        self.assertEqual(desempate["partida"]["estado"], ESTADO_PARTIDA_DESEMPATE)
        self.assertNotIn("idbingadores", desempate["partida"])

    def test_ganador_solo_se_confirma_sin_publicar_alias(self):
        ganador = Jugador(idjugador=41, aliasjugador="campeon_publico")
        partida = partida_publica_prueba(ganador=ganador)
        en_curso = construir_payload_publico_partida(partida, "ganador_detectado")
        partida.estadopartida = ESTADO_PARTIDA_FINALIZADA
        partida.haydesempate = True
        finalizada = construir_payload_publico_partida(
            partida,
            "desempate_finalizado",
        )

        self.assertIsNone(en_curso["partida"]["ganador"])
        self.assertEqual(finalizada["partida"]["ganador"], "Confirmado")
        self.assertNotIn("campeon_publico", json.dumps(finalizada))
        self.assertTrue(finalizada["partida"]["finalizada"])
        self.assertTrue(finalizada["partida"]["resuelta_por_desempate"])

    def test_publicacion_se_difiere_hasta_on_commit(self):
        partida = partida_publica_prueba(bolas=[12])
        callbacks = []
        capa = Mock()
        capa.group_send = AsyncMock()

        with (
            patch(
                "apps.bingos.realtime.transaction.on_commit",
                side_effect=callbacks.append,
            ),
            patch("apps.bingos.realtime.get_channel_layer", return_value=capa),
        ):
            programar_publicacion_partida(partida, "bola_extraida")
            capa.group_send.assert_not_awaited()
            self.assertEqual(len(callbacks), 1)
            callbacks[0]()

        capa.group_send.assert_awaited_once()


class IntegracionPublicacionTiempoRealTests(SimpleTestCase):
    def setUp(self):
        self.factory = RequestFactory()

    def _request_con_mensajes(self, path, data=None):
        request = self.factory.post(path, data or {})
        request.user = User(username="admin", is_staff=True)
        request.session = {}
        request._messages = FallbackStorage(request)
        return request

    def test_pausar_y_reanudar_publican_el_estado_confirmado(self):
        casos = (
            ("pausar", ESTADO_PARTIDA_EN_CURSO, "partida_pausada"),
            ("reanudar", ESTADO_PARTIDA_PAUSADA, "partida_reanudada"),
        )
        for accion, estado_inicial, evento in casos:
            with self.subTest(accion=accion):
                request = self._request_con_mensajes(
                    "/partidas/20/consola/",
                    {"accion": accion},
                )
                partida = Partidabingo(
                    idpartidabingo=20,
                    estadopartida=estado_inicial,
                    bolascantadas="[]",
                )
                partida.save = Mock()
                with (
                    patch(
                        "apps.bingos.views._base_datos_permite_estado_partida",
                        return_value=True,
                    ),
                    patch(
                        "apps.bingos.views.transaction.atomic",
                        return_value=nullcontext(),
                    ),
                    patch(
                        "apps.bingos.views.programar_publicacion_partida"
                    ) as publicar_mock,
                ):
                    resultado = _procesar_accion_consola(
                        request,
                        partida,
                        accion,
                    )

                self.assertTrue(resultado)
                publicar_mock.assert_called_once_with(partida, evento)

    def test_detectar_desempate_publica_solo_actualizacion_publica(self):
        request = self._request_con_mensajes(
            "/partidas/20/cartones/31/validar/"
        )
        partida = partida_publica_prueba(estado=ESTADO_PARTIDA_DESEMPATE)
        carton = carton_publico_prueba(partida=partida)
        resultado = {
            "resultado": "desempate",
            "partida": partida,
            "carton": carton,
        }
        with (
            patch(
                "apps.bingos.views.get_object_or_404",
                side_effect=[partida, carton],
            ),
            patch(
                "apps.bingos.views.validar_carton_ganador",
                return_value=resultado,
            ),
            patch(
                "apps.bingos.views.programar_publicacion_partida"
            ) as publicar_mock,
        ):
            response = validar_carton(request, 20, 31)

        self.assertEqual(response.status_code, 302)
        publicar_mock.assert_called_once_with(partida, "desempate_detectado")

    def test_confirmar_desempate_publica_finalizacion_con_ganador(self):
        request = self._request_con_mensajes(
            "/partidas/20/desempate/confirmar/"
        )
        partida = partida_publica_prueba(estado=ESTADO_PARTIDA_FINALIZADA)
        confirmacion = {
            "partida": partida,
            "resultado": {
                "idjugador": 41,
                "jugador": "juan123",
                "codigo": "O-67",
            },
        }
        with (
            patch("apps.bingos.views.get_object_or_404", return_value=partida),
            patch(
                "apps.bingos.views.confirmar_y_finalizar_desempate",
                return_value=confirmacion,
            ),
            patch(
                "apps.bingos.views.programar_publicacion_partida"
            ) as publicar_mock,
        ):
            response = confirmar_desempate(request, 20)

        self.assertEqual(response.status_code, 302)
        publicar_mock.assert_called_once_with(
            partida,
            "desempate_finalizado",
            ganador_publico="juan123",
        )

    def test_error_de_transaccion_no_publica_cambio_parcial(self):
        request = self._request_con_mensajes(
            "/partidas/20/consola/",
            {"accion": "pausar"},
        )
        partida = Partidabingo(
            idpartidabingo=20,
            estadopartida=ESTADO_PARTIDA_EN_CURSO,
        )
        partida.save = Mock(side_effect=DatabaseError("fallo simulado"))
        with (
            patch(
                "apps.bingos.views._base_datos_permite_estado_partida",
                return_value=True,
            ),
            patch(
                "apps.bingos.views.transaction.atomic",
                return_value=nullcontext(),
            ),
            patch(
                "apps.bingos.views.programar_publicacion_partida"
            ) as publicar_mock,
        ):
            resultado = _procesar_accion_consola(
                request,
                partida,
                "pausar",
            )

        self.assertFalse(resultado)
        self.assertEqual(partida.estadopartida, ESTADO_PARTIDA_EN_CURSO)
        publicar_mock.assert_not_called()


class TiempoRealParticipacionesHibridasTests(SimpleTestCase):
    def setUp(self):
        self.factory = RequestFactory()
        self.bingo = Bingo(idbingo=70, titulobingo="Bingo tiempo real")
        self.partida_a = partida_publica_prueba(
            bolas=[1, 16, 31],
            pk=701,
        )
        self.partida_a.idbingo = self.bingo
        self.partida_b = partida_publica_prueba(
            bolas=[2, 17],
            pk=702,
        )
        self.partida_b.idbingo = self.bingo
        self.jugador = Jugador(
            idjugador=81,
            aliasjugador="alias_privado_ws",
            correojugador="privado-ws@example.com",
        )
        self.carton_hibrido = Carton(
            idcarton=801,
            idjugador=self.jugador,
            idbingo=self.bingo,
            idpartida=None,
            codigocarton="B70-C-WS",
            matriznumeros=serializar_matriz_carton_bingo(matriz_carton_prueba()),
            indicevictoria=999,
            preciopagado=Decimal("25.00"),
            estadocarton="Vendido",
        )
        self.carton_historico = Carton(
            idcarton=802,
            idjugador=self.jugador,
            idbingo=self.bingo,
            idpartida=self.partida_a,
            codigocarton="P701-C-WS",
            matriznumeros=self.carton_hibrido.matriznumeros,
            indicevictoria=3,
            estadocarton="Vendido",
        )

    def _request_con_mensajes(self, path):
        request = self.factory.post(path)
        request.user = User(username="admin_ws", is_staff=True)
        request.session = {}
        request._messages = FallbackStorage(request)
        return request

    def test_grupo_se_construye_con_id_de_partida(self):
        self.assertEqual(nombre_grupo_partida(self.partida_a.pk), "partida_701")

    def test_partidas_distintas_producen_grupos_distintos(self):
        self.assertNotEqual(
            nombre_grupo_partida(self.partida_a.pk),
            nombre_grupo_partida(self.partida_b.pk),
        )

    def test_publicacion_de_a_solo_se_dirige_al_grupo_de_a(self):
        callbacks = []
        capa = Mock()
        capa.group_send = AsyncMock()
        with (
            patch(
                "apps.bingos.realtime.transaction.on_commit",
                side_effect=callbacks.append,
            ),
            patch("apps.bingos.realtime.get_channel_layer", return_value=capa),
        ):
            payload = programar_publicacion_partida(
                self.partida_a,
                "bola_extraida",
            )
            callbacks[0]()

        capa.group_send.assert_awaited_once_with(
            nombre_grupo_partida(self.partida_a.pk),
            {
                "type": "partida.actualizada",
                "payload": payload,
            },
        )
        self.assertNotEqual(
            capa.group_send.await_args.args[0],
            nombre_grupo_partida(self.partida_b.pk),
        )

    def test_extraccion_publica_la_partida_operada(self):
        request = self._request_con_mensajes(
            "/partidas/701/sacar-bola/"
        )
        with (
            patch(
                "apps.bingos.views.get_object_or_404",
                return_value=self.partida_a,
            ),
            patch("apps.bingos.views.extraer_siguiente_bola", return_value=46),
            patch(
                "apps.bingos.views.programar_publicacion_partida"
            ) as publicar,
        ):
            response = sacar_bola(request, self.partida_a.pk)

        self.assertEqual(response.status_code, 302)
        publicar.assert_called_once_with(self.partida_a, "bola_extraida")

    def test_validacion_hibrida_publica_partida_del_resultado(self):
        request = self._request_con_mensajes(
            "/partidas/701/cartones/801/validar/"
        )
        resultado = {
            "resultado": "ganador",
            "partida": self.partida_a,
            "carton": self.carton_hibrido,
        }
        with (
            patch(
                "apps.bingos.views.get_object_or_404",
                side_effect=[self.partida_a, self.carton_hibrido],
            ),
            patch(
                "apps.bingos.views.validar_participacion_ganadora",
                return_value=resultado,
            ) as validar_hibrida,
            patch("apps.bingos.views.validar_carton_ganador") as validar_historica,
            patch(
                "apps.bingos.views.programar_publicacion_partida"
            ) as publicar,
        ):
            response = validar_carton(
                request,
                self.partida_a.pk,
                self.carton_hibrido.pk,
            )

        self.assertEqual(response.status_code, 302)
        validar_hibrida.assert_called_once_with(
            partida=self.partida_a,
            carton=self.carton_hibrido,
            indicevictoria=3,
        )
        validar_historica.assert_not_called()
        publicar.assert_called_once_with(self.partida_a, "ganador_detectado")

    def test_validacion_historica_conserva_publicacion_por_partida(self):
        request = self._request_con_mensajes(
            "/partidas/701/cartones/802/validar/"
        )
        resultado = {
            "resultado": "ganador",
            "partida": self.partida_a,
            "carton": self.carton_historico,
        }
        with (
            patch(
                "apps.bingos.views.get_object_or_404",
                side_effect=[self.partida_a, self.carton_historico],
            ),
            patch(
                "apps.bingos.views.validar_carton_ganador",
                return_value=resultado,
            ) as validar_historica,
            patch("apps.bingos.views.validar_participacion_ganadora") as validar_hibrida,
            patch(
                "apps.bingos.views.programar_publicacion_partida"
            ) as publicar,
        ):
            response = validar_carton(
                request,
                self.partida_a.pk,
                self.carton_historico.pk,
            )

        self.assertEqual(response.status_code, 302)
        validar_historica.assert_called_once_with(
            self.partida_a,
            self.carton_historico,
        )
        validar_hibrida.assert_not_called()
        publicar.assert_called_once_with(self.partida_a, "ganador_detectado")

    def test_payload_publico_no_expone_datos_privados(self):
        self.partida_a.estadopartida = ESTADO_PARTIDA_FINALIZADA
        self.partida_a.idjugadorganador = self.jugador
        self.partida_a.idbingadores = json.dumps(
            [
                {
                    "idjugador": self.jugador.pk,
                    "aliasjugador": self.jugador.aliasjugador,
                    "correo": self.jugador.correojugador,
                    "matriznumeros": self.carton_hibrido.matriznumeros,
                    "indicevictoria": 7,
                    "estado_participacion": "Ganador",
                }
            ]
        )
        payload = construir_payload_publico_partida(
            self.partida_a,
            "desempate_finalizado",
            ganador_publico="Nombre privado",
        )
        serializado = json.dumps(payload)

        self.assertEqual(payload["partida"]["ganador"], "Confirmado")
        for privado in (
            "alias_privado_ws",
            "privado-ws@example.com",
            "Nombre privado",
            "idjugador",
            "aliasjugador",
            "correo",
            "matriznumeros",
            "indicevictoria",
            "estado_participacion",
            "preciopagado",
        ):
            self.assertNotIn(privado, serializado)

    def test_eventos_de_resultado_solicitan_recarga(self):
        for evento in EVENTOS_REQUIEREN_RECARGA:
            with self.subTest(evento=evento):
                payload = construir_payload_publico_partida(
                    self.partida_a,
                    evento,
                )
                self.assertTrue(payload["requiere_recarga"])
        payload_bola = construir_payload_publico_partida(
            self.partida_a,
            "bola_extraida",
        )
        self.assertFalse(payload_bola["requiere_recarga"])

    def test_javascript_filtra_partida_antes_de_recargar_url_actual(self):
        javascript = Path("static/js/realtime_bingo.js").read_text(
            encoding="utf-8"
        )
        filtro = 'String(payload.partida.id) !== partidaId'
        recarga = "recargarVistaActual();"

        self.assertIn(filtro, javascript)
        self.assertIn(recarga, javascript)
        self.assertIn("window.location.reload();", javascript)
        self.assertLess(javascript.index(filtro), javascript.index(recarga))
        self.assertNotIn("window.location.href =", javascript)
        self.assertNotIn("window.location.assign(", javascript)
        self.assertNotIn("window.location.replace(", javascript)

    def test_ruta_y_tipo_de_evento_principales_se_conservan(self):
        from apps.bingos.routing import websocket_urlpatterns

        ruta = next(
            item
            for item in websocket_urlpatterns
            if getattr(item, "name", None) == "ws_partida_publica"
        )
        payload = construir_payload_publico_partida(
            self.partida_a,
            "bola_extraida",
        )

        self.assertIs(ruta.callback.consumer_class, PartidaPublicaConsumer)
        self.assertEqual(payload["tipo"], "partida_actualizada")


@override_settings(
    STORAGES={
        "default": {"BACKEND": "django.core.files.storage.FileSystemStorage"},
        "staticfiles": {
            "BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage"
        },
    }
)
class PlantillasTiempoRealBingoTests(SimpleTestCase):
    def setUp(self):
        self.request = RequestFactory().get("/juego/")
        self.request.user = AnonymousUser()

    def _contexto_consola(self, partida):
        bolas = parsear_bolas_cantadas(partida.bolascantadas)
        bolas_disponibles = obtener_bolas_disponibles(bolas)
        return {
            "partida": partida,
            "acciones_consola": [],
            "cartones": [],
            "cartones_validacion": [],
            "participaciones_hibridas": [],
            "participaciones_hibridas_validacion": [],
            "candidatos_desempate": [],
            "tablero_bingo": preparar_datos_tablero_publico(partida)["tablero_bingo"],
            "historial_bolas": [
                {"numero": numero, "codigo": formatear_bola_bingo(numero)}
                for numero in bolas
            ],
            "ultima_bola_codigo": formatear_bola_bingo(bolas[-1]) if bolas else None,
            "total_bolas_extraidas": len(bolas),
            "total_bolas_faltantes": len(bolas_disponibles),
            "hay_bolas_disponibles": bool(bolas_disponibles),
            "puede_sacar_bola": (
                partida.estadopartida == ESTADO_PARTIDA_EN_CURSO
                and bool(bolas_disponibles)
            ),
        }

    def test_tablero_conserva_actualizacion_manual_y_no_agrega_acciones_admin(self):
        partida = partida_publica_prueba(bolas=[1, 23])
        html = render_to_string(
            "bingos/tablero_publico.html",
            {"partida": partida, **preparar_datos_tablero_publico(partida)},
            request=self.request,
        )

        self.assertIn("Actualizar tablero", html)
        self.assertIn("data-realtime-bingo", html)
        self.assertIn("realtime_bingo.js", html)
        self.assertIn("data-realtime-audio-toggle", html)
        self.assertIn('aria-pressed="false"', html)
        self.assertIn("Activar sonido", html)
        self.assertNotIn("Sacar siguiente bola", html)
        self.assertNotIn("Validar cartón", html)

    def test_carton_conserva_actualizacion_manual_sin_reclamo(self):
        partida = partida_publica_prueba(bolas=[1, 23])
        carton = carton_publico_prueba(partida=partida)
        html = render_to_string(
            "bingos/carton_publico.html",
            {
                "partida": partida,
                "carton": carton,
                "error_carton": None,
                **preparar_datos_carton_jugador(carton),
            },
            request=self.request,
        )

        self.assertIn("Actualizar cartón", html)
        self.assertIn("data-carton-cell", html)
        self.assertIn("realtime_bingo.js", html)
        self.assertIn("data-realtime-audio-toggle", html)
        self.assertNotIn("Reclamar", html)
        self.assertNotIn("Validar cartón", html)

    def test_mi_carton_privado_incluye_control_de_sonido(self):
        partida = partida_publica_prueba(bolas=[1, 23])
        carton = carton_publico_prueba(partida=partida)
        html = render_to_string(
            "bingos/mi_carton_detalle.html",
            {
                "partida": partida,
                "carton": carton,
                "jugador": carton.idjugador,
                "error_carton": None,
                **preparar_datos_carton_jugador(carton),
            },
            request=self.request,
        )

        self.assertIn("data-realtime-audio-toggle", html)
        self.assertIn("realtime_bingo.js", html)

    def test_consola_operador_incluye_control_y_receptor_tiempo_real(self):
        partida = partida_publica_prueba(bolas=[1, 23])
        html = render_to_string(
            "bingos/consola_operador.html",
            self._contexto_consola(partida),
            request=self.request,
        )

        self.assertIn("data-realtime-audio-toggle", html)
        self.assertIn('data-realtime-mode="operador"', html)
        self.assertIn("realtime_bingo.js", html)
        self.assertIn("bolillero_automatico.js", html)
        self.assertIn("Patrón ganador", html)
        self.assertIn("Cartón lleno", html)

    def test_consola_operador_incluye_bolillero_automatico(self):
        partida = partida_publica_prueba(bolas=[1, 23])
        html = render_to_string(
            "bingos/consola_operador.html",
            self._contexto_consola(partida),
            request=self.request,
        )

        self.assertIn("Bolillero automático", html)
        self.assertIn("data-bolillero-intervalo", html)
        self.assertIn("3 segundos", html)
        self.assertIn("5 segundos", html)
        self.assertIn("8 segundos", html)
        self.assertIn("10 segundos", html)
        self.assertIn("Iniciar extracción automática", html)
        self.assertIn("Detener extracción automática", html)
        self.assertIn("data-bolillero-detener disabled", html)
        self.assertIn('aria-live="polite"', html)
        self.assertIn("Automático detenido", html)
        self.assertIn("Sacar siguiente bola", html)
        self.assertIn("data-bolillero-manual-form", html)

    def test_consola_operador_deshabilita_inicio_automatico_si_no_puede_extraer(self):
        partida = partida_publica_prueba(
            estado=ESTADO_PARTIDA_PAUSADA,
            bolas=[1, 23],
        )
        html = render_to_string(
            "bingos/consola_operador.html",
            self._contexto_consola(partida),
            request=self.request,
        )

        self.assertIn("data-bolillero-iniciar disabled", html)
        self.assertIn("Sacar siguiente bola", html)

    def test_consola_operador_tiene_datos_para_actualizacion_visual(self):
        partida = partida_publica_prueba(bolas=[1, 23])
        html = render_to_string(
            "bingos/consola_operador.html",
            self._contexto_consola(partida),
            request=self.request,
        )

        self.assertIn("data-realtime-estado", html)
        self.assertIn("data-realtime-ultima-bola", html)
        self.assertIn("data-realtime-sin-bolas", html)
        self.assertIn("data-realtime-total-extraidas", html)
        self.assertIn("data-realtime-restantes", html)
        self.assertIn("data-realtime-history", html)
        self.assertIn("data-realtime-history-empty", html)
        self.assertIn('data-bola-numero="1"', html)
        self.assertIn('data-bola-numero="75"', html)

    def test_javascript_anuncia_solo_bolas_nuevas_sin_duplicados(self):
        javascript = Path("static/js/realtime_bingo.js").read_text(
            encoding="utf-8"
        )

        self.assertIn("window.localStorage", javascript)
        self.assertIn('payload.evento !== "bola_extraida"', javascript)
        self.assertIn("balotasProcesadas.has(claveBalota)", javascript)
        self.assertIn("balotasProcesadas.add(claveBalota)", javascript)
        self.assertIn("window.speechSynthesis.speak(utterance)", javascript)
        self.assertIn('utterance.lang = "es-EC"', javascript)
        self.assertEqual(javascript.count("speechSynthesis.speak"), 1)
        self.assertLess(
            javascript.index('payload.evento !== "bola_extraida"'),
            javascript.index("window.speechSynthesis.speak(utterance)"),
        )

    def test_javascript_notifica_actualizaciones_para_modulos_de_consola(self):
        javascript = Path("static/js/realtime_bingo.js").read_text(
            encoding="utf-8"
        )

        self.assertIn('"siab:partidaActualizada"', javascript)
        self.assertIn("detail: payload", javascript)
        self.assertIn("bubbles: true", javascript)

    def test_javascript_actualiza_el_patron_en_vistas_publicas(self):
        javascript = Path("static/js/realtime_bingo.js").read_text(
            encoding="utf-8"
        )

        self.assertIn("actualizarPatronTablero", javascript)
        self.assertIn("actualizarPatronCarton", javascript)
        self.assertIn("[data-realtime-patron-label]", javascript)
        self.assertIn("[data-realtime-patron-progreso]", javascript)
        self.assertIn("[data-realtime-patron-faltantes]", javascript)
        self.assertIn("[data-realtime-patron-completo]", javascript)

    def test_javascript_actualiza_tablero_en_modo_operador(self):
        javascript = Path("static/js/realtime_bingo.js").read_text(
            encoding="utf-8"
        )

        self.assertIn('raiz.dataset.realtimeMode === "operador"', javascript)
        self.assertLess(
            javascript.index('raiz.dataset.realtimeMode === "operador"'),
            javascript.index("actualizarTablero(raiz, payload.partida)"),
        )

    def test_javascript_bolillero_automatico_reutiliza_post_manual_y_bloquea_concurrencia(self):
        javascript = Path("static/js/bolillero_automatico.js").read_text(
            encoding="utf-8"
        )

        self.assertIn("[data-bolillero-manual-form]", javascript)
        self.assertIn("window.fetch(contexto.formularioManual.action", javascript)
        self.assertIn('"POST"', javascript)
        self.assertIn('"X-CSRFToken": obtenerTokenCsrf(contexto)', javascript)
        self.assertIn("new window.FormData(contexto.formularioManual)", javascript)
        self.assertIn("solicitudPendiente", javascript)
        self.assertIn("contexto.solicitudPendiente", javascript)
        self.assertIn("Espera a que termine la extracción automática pendiente.", javascript)

    def test_javascript_bolillero_automatico_posterga_detencion_si_hay_peticion_pendiente(self):
        javascript = Path("static/js/bolillero_automatico.js").read_text(
            encoding="utf-8"
        )

        self.assertIn("detenerSolicitado", javascript)
        self.assertIn("contexto.solicitudPendiente", javascript)
        self.assertIn("Se detendrá después de la balota actual.", javascript)
        self.assertIn('escribirEstado(contexto, "Automático detenido")', javascript)
        self.assertLess(
            javascript.index("Se detendrá después de la balota actual."),
            javascript.index('escribirEstado(contexto, "Automático detenido")'),
        )

    def test_javascript_bolillero_automatico_detiene_estados_no_permitidos(self):
        javascript = Path("static/js/bolillero_automatico.js").read_text(
            encoding="utf-8"
        )

        self.assertIn('var ESTADO_EN_CURSO = "En curso";', javascript)
        self.assertIn("contexto.estadoPartida === ESTADO_EN_CURSO", javascript)
        self.assertIn("!extraccionDisponible || contexto.activo || contexto.solicitudPendiente", javascript)
        self.assertIn("payload.evento === \"desempate_detectado\"", javascript)
        self.assertIn("Extracción detenida: la partida ya no está En curso", javascript)
        self.assertIn("Extracción detenida: ya no quedan bolas disponibles", javascript)
        self.assertIn("beforeunload", javascript)


class FakeReportQuerySet(list):
    def select_related(self, *fields):
        return self

    def order_by(self, *fields):
        return self

    def filter(self, *args, **kwargs):
        return self


class ReportesAdministrativosTests(SimpleTestCase):
    def setUp(self):
        self.factory = RequestFactory()
        self.bingo = self._bingo()
        self.ganador = Jugador(idjugador=41, aliasjugador="juan123")
        self.partida = self._partida(
            estado=ESTADO_PARTIDA_FINALIZADA,
            ganador=self.ganador,
            hay_desempate=True,
            balota=67,
        )
        self.cartones = [
            self._carton(31, "P20-C-31", self.ganador, Decimal("5.00"), "Vendido"),
            self._carton(32, "P20-C-32", Jugador(idjugador=42, aliasjugador="maria456"), Decimal("5.00"), "Vendido"),
            self._carton(33, "P20-C-33", None, None, "Disponible"),
        ]

    def _request(self, path, user):
        request = self.factory.get(path)
        request.user = user
        return request

    def _staff(self):
        return User(username="admin", is_staff=True)

    def _usuario_normal(self):
        return User(username="usuario", is_staff=False, is_superuser=False)

    def _bingo(self):
        return Bingo(
            idbingo=7,
            titulobingo="Bingo de reportes",
            fechaprogramadabingo=timezone.now(),
            tipobingo="Virtual",
            lugarbingo="CoopBingo",
            preciocarton=Decimal("5.00"),
            premiomayor=Decimal("100.00"),
            descripcionpremiomayor="Premio mayor",
            estadobingo="Programado",
        )

    def _partida(
        self,
        estado=ESTADO_PARTIDA_EN_CURSO,
        ganador=None,
        hay_desempate=False,
        balota=None,
    ):
        return Partidabingo(
            idpartidabingo=20,
            idbingo=self.bingo,
            idjugadorganador=ganador,
            nombreronda="Ronda reportes",
            valorefectivo=Decimal("100.00"),
            premiomaterial="Canasta",
            estadopartida=estado,
            bolascantadas=serializar_bolas_cantadas([1, 16, 31]),
            ultimabola=31,
            haydesempate=hay_desempate,
            idbingadores='[{"privado": "NO_EXPORTAR"}]',
            bolamayordesempate=balota,
            horainicio=timezone.now(),
            horafin=timezone.now() if estado == ESTADO_PARTIDA_FINALIZADA else None,
        )

    def _carton(self, pk, codigo, jugador, precio, estado):
        return Carton(
            idcarton=pk,
            idjugador=jugador,
            idpartida=self.partida,
            codigocarton=codigo,
            matriznumeros=serializar_matriz_carton_bingo(matriz_carton_prueba()),
            indicevictoria=1 if jugador == self.ganador else None,
            preciopagado=precio,
            fechacompra=timezone.now(),
            estadocarton=estado,
        )

    def _patch_partida_reportes(self):
        return (
            patch("apps.bingos.views.get_object_or_404", return_value=self.partida),
            patch(
                "apps.bingos.views.Carton.objects.filter",
                return_value=FakeReportQuerySet(self.cartones),
            ),
            patch(
                "apps.bingos.views.CartonPartidaBingo.objects.filter",
                return_value=FakeReportQuerySet(),
            ),
        )

    def _patch_bingo_resumen(self):
        return (
            patch("apps.bingos.views.get_object_or_404", return_value=self.bingo),
            patch(
                "apps.bingos.views.Partidabingo.objects.filter",
                return_value=FakeReportQuerySet([self.partida]),
            ),
            patch(
                "apps.bingos.views.Carton.objects.filter",
                return_value=FakeReportQuerySet(self.cartones),
            ),
            patch(
                "apps.bingos.views.CartonPartidaBingo.objects.filter",
                return_value=FakeReportQuerySet(),
            ),
        )

    def test_anonimo_es_redirigido_al_descargar_reportes(self):
        casos = (
            (partida_reporte_pdf, "/partidas/20/reporte/pdf/", (20,)),
            (partida_cartones_excel, "/partidas/20/cartones/excel/", (20,)),
            (bingo_resumen_excel, "/bingos/7/resumen/excel/", (7,)),
        )
        for view, path, args in casos:
            with self.subTest(path=path):
                response = view(self._request(path, AnonymousUser()), *args)

                self.assertEqual(response.status_code, 302)
                self.assertIn("/login/", response["Location"])

    def test_usuario_normal_y_jugador_reciben_403(self):
        casos = (
            (partida_reporte_pdf, "/partidas/20/reporte/pdf/", (20,)),
            (partida_cartones_excel, "/partidas/20/cartones/excel/", (20,)),
            (bingo_resumen_excel, "/bingos/7/resumen/excel/", (7,)),
        )
        for user in (self._usuario_normal(), User(username="jugador", is_staff=False)):
            for view, path, args in casos:
                with self.subTest(user=user.username, path=path):
                    with self.assertRaises(PermissionDenied):
                        view(self._request(path, user), *args)

    def test_staff_descarga_pdf_de_partida(self):
        reportes_patches = self._patch_partida_reportes()
        with reportes_patches[0], reportes_patches[1], reportes_patches[2]:
            response = partida_reporte_pdf(
                self._request("/partidas/20/reporte/pdf/", self._staff()),
                20,
            )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], PDF_CONTENT_TYPE)
        self.assertIn("attachment", response["Content-Disposition"])
        self.assertIn("reporte_partida_20.pdf", response["Content-Disposition"])
        self.assertTrue(response.content.startswith(b"%PDF"))

    def test_superusuario_descarga_reporte(self):
        superuser = User(username="root", is_superuser=True)
        reportes_patches = self._patch_partida_reportes()
        with reportes_patches[0], reportes_patches[1], reportes_patches[2]:
            response = partida_reporte_pdf(
                self._request("/partidas/20/reporte/pdf/", superuser),
                20,
            )

        self.assertEqual(response.status_code, 200)

    def test_staff_descarga_excel_de_cartones(self):
        reportes_patches = self._patch_partida_reportes()
        with reportes_patches[0], reportes_patches[1], reportes_patches[2]:
            response = partida_cartones_excel(
                self._request("/partidas/20/cartones/excel/", self._staff()),
                20,
            )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], XLSX_CONTENT_TYPE)
        self.assertIn("attachment", response["Content-Disposition"])
        self.assertIn("cartones_partida_20.xlsx", response["Content-Disposition"])
        workbook = load_workbook(BytesIO(response.content))
        self.assertIn("Cartones", workbook.sheetnames)

    def test_staff_descarga_excel_resumen_de_bingo(self):
        resumen_patches = self._patch_bingo_resumen()
        with (
            resumen_patches[0],
            resumen_patches[1],
            resumen_patches[2],
            resumen_patches[3],
        ):
            response = bingo_resumen_excel(
                self._request("/bingos/7/resumen/excel/", self._staff()),
                7,
            )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], XLSX_CONTENT_TYPE)
        self.assertIn("attachment", response["Content-Disposition"])
        self.assertIn("resumen_bingo_7.xlsx", response["Content-Disposition"])
        workbook = load_workbook(BytesIO(response.content))
        self.assertIn("Resumen de partidas", workbook.sheetnames)

    def test_excel_cartones_contiene_headers_resumen_y_no_privados(self):
        contenido = generar_excel_cartones_partida(self.partida, self.cartones)
        workbook = load_workbook(BytesIO(contenido))
        worksheet = workbook["Cartones"]
        headers = [cell.value for cell in worksheet[1]]

        self.assertEqual(
            headers,
            [
                "Tipo de registro",
                "ID de cartón",
                "Código de cartón",
                "Jugador",
                "Estado del cartón",
                "Estado de participación",
                "Fecha de compra",
                "Importe registrado del cartón",
                "Índice de victoria",
                "Fecha de validación",
                "Ganador de la ronda",
                "ID de partida",
                "Nombre de ronda",
                "Estado de partida",
            ],
        )
        headers_lower = " ".join(str(header).lower() for header in headers)
        for prohibido in ("correo", "contraseña", "password", "hash", "idbingadores"):
            self.assertNotIn(prohibido, headers_lower)

        resumen = {
            worksheet.cell(row=row, column=1).value: worksheet.cell(row=row, column=2).value
            for row in range(worksheet.max_row - 5, worksheet.max_row + 1)
        }
        self.assertEqual(resumen["Total de cartones"], 3)
        self.assertNotIn("Total recaudado", resumen)
        self.assertEqual(resumen["Cartones vendidos"], 2)
        self.assertEqual(resumen["Cartones disponibles"], 1)

    def test_excel_resumen_bingo_contiene_headers_resumen_y_no_privados(self):
        contenido = generar_excel_resumen_bingo(
            self.bingo,
            [self.partida],
            self.cartones,
        )
        workbook = load_workbook(BytesIO(contenido))
        worksheet = workbook["Resumen de partidas"]
        headers = [cell.value for cell in worksheet[18]]

        self.assertEqual(worksheet["A1"].value, "Resumen del Bingo")
        self.assertEqual(
            worksheet["A3"].value,
            "A. Resumen financiero del Bingo",
        )
        self.assertEqual(worksheet["A17"].value, "B. Resumen de rondas")
        self.assertIn("Bingo", headers)
        self.assertNotIn("Recaudación total", headers)
        self.assertIn("Cartones participantes en la ronda", headers)
        self.assertNotIn(
            "Cartones registrados directamente en la ronda",
            headers,
        )
        self.assertIn("Participaciones por ronda", headers)
        self.assertIn("Balota mayor de desempate", headers)
        headers_lower = " ".join(str(header).lower() for header in headers)
        for prohibido in ("correo", "contraseña", "password", "hash", "idbingadores"):
            self.assertNotIn(prohibido, headers_lower)

        resumen = {
            worksheet.cell(row=row, column=1).value: worksheet.cell(row=row, column=2).value
            for row in range(4, 13)
        }
        self.assertEqual(resumen["Nombre del Bingo"], self.bingo.titulobingo)
        self.assertEqual(resumen["Total de rondas"], 1)
        self.assertEqual(resumen["Rondas finalizadas"], 1)
        self.assertEqual(resumen["Total de cartones maestros vendidos"], 3)
        self.assertEqual(resumen["Total recaudado"], 10)
        self.assertEqual(resumen["Rondas con desempate"], 1)

    def test_generar_reportes_no_modifica_objetos_ni_llama_save(self):
        estado = self.partida.estadopartida
        bolas = self.partida.bolascantadas
        ganador = self.partida.idjugadorganador
        self.partida.save = Mock()
        for carton in self.cartones:
            carton.save = Mock()

        generar_pdf_reporte_partida(self.partida, self.cartones)
        generar_excel_cartones_partida(self.partida, self.cartones)
        generar_excel_resumen_bingo(self.bingo, [self.partida], self.cartones)

        self.partida.save.assert_not_called()
        for carton in self.cartones:
            carton.save.assert_not_called()
        self.assertEqual(self.partida.estadopartida, estado)
        self.assertEqual(self.partida.bolascantadas, bolas)
        self.assertEqual(self.partida.idjugadorganador, ganador)

    def test_datos_partida_finalizada_incluyen_ganador_y_desempate_seguro(self):
        datos = construir_datos_reporte_partida(self.partida, self.cartones)
        pdf = generar_pdf_reporte_partida(self.partida, self.cartones)

        self.assertTrue(datos["finalizada"])
        self.assertEqual(datos["ganador"], "juan123")
        self.assertEqual(datos["carton_ganador"], "P20-C-31")
        self.assertTrue(datos["hubo_desempate"])
        self.assertEqual(datos["balota_mayor_desempate"], "O-67")
        self.assertNotIn("idbingadores", datos)
        self.assertIn(b"juan123", pdf)
        self.assertNotIn(b"NO_EXPORTAR", pdf)
        self.assertNotIn(b"idbingadores", pdf)

    def test_datos_partida_no_finalizada_no_inventan_ganador(self):
        partida = self._partida(
            estado=ESTADO_PARTIDA_EN_CURSO,
            ganador=self.ganador,
            hay_desempate=False,
        )
        datos = construir_datos_reporte_partida(partida, self.cartones)

        self.assertFalse(datos["finalizada"])
        self.assertIsNone(datos["ganador"])
        self.assertIsNone(datos["carton_ganador"])
        self.assertEqual(datos["mensaje_resultado"], "La partida aún no está finalizada.")


class ReportesHibridosTests(SimpleTestCase):
    def setUp(self):
        self.factory = RequestFactory()
        self.bingo = Bingo(
            idbingo=70,
            titulobingo="Bingo híbrido de reportes",
            fechaprogramadabingo=timezone.now(),
            tipobingo="Virtual",
            preciocarton=Decimal("12.00"),
            premiomayor=Decimal("100.00"),
            descripcionpremiomayor="Premio",
        )
        self.otro_bingo = Bingo(idbingo=71, titulobingo="Otro Bingo")
        self.jugador_historico = Jugador(
            idjugador=81,
            aliasjugador="historico_reporte",
        )
        self.jugador_hibrido = Jugador(
            idjugador=82,
            aliasjugador="hibrido_reporte",
        )
        self.partida_uno = self._partida(
            701,
            self.bingo,
            ESTADO_PARTIDA_FINALIZADA,
            self.jugador_hibrido,
        )
        self.partida_dos = self._partida(
            702,
            self.bingo,
            ESTADO_PARTIDA_EN_CURSO,
        )
        self.carton_historico = Carton(
            idcarton=801,
            idjugador=self.jugador_historico,
            idbingo=self.bingo,
            idpartida=self.partida_uno,
            codigocarton="P701-C-HIST",
            matriznumeros=serializar_matriz_carton_bingo(matriz_carton_prueba()),
            indicevictoria=4,
            preciopagado=Decimal("5.00"),
            fechacompra=timezone.now(),
            estadocarton="Vendido",
        )
        self.carton_hibrido = Carton(
            idcarton=802,
            idjugador=self.jugador_hibrido,
            idbingo=self.bingo,
            idpartida=None,
            codigocarton="B70-C-MAESTRO",
            matriznumeros=serializar_matriz_carton_bingo(matriz_carton_prueba()),
            indicevictoria=999,
            preciopagado=Decimal("12.00"),
            fechacompra=timezone.now(),
            estadocarton="Vendido",
        )
        self.participacion_uno = self._participacion(
            901,
            self.partida_uno,
            CartonPartidaBingo.ESTADO_GANADOR,
            indice=7,
        )
        self.participacion_dos = self._participacion(
            902,
            self.partida_dos,
            CartonPartidaBingo.ESTADO_PENDIENTE,
        )

    def _partida(self, pk, bingo, estado, ganador=None):
        ahora = timezone.now()
        return Partidabingo(
            idpartidabingo=pk,
            idbingo=bingo,
            idjugadorganador=ganador,
            nombreronda=f"Ronda {pk}",
            valorefectivo=Decimal("100.00"),
            premiomaterial="Premio",
            estadopartida=estado,
            bolascantadas=serializar_bolas_cantadas([1, 16, 31]),
            ultimabola=31,
            haydesempate=False,
            horainicio=ahora,
            horafin=ahora if estado == ESTADO_PARTIDA_FINALIZADA else None,
        )

    def _participacion(self, pk, partida, estado, indice=None, bingo=None):
        ahora = timezone.now()
        return CartonPartidaBingo(
            idcartonpartidabingo=pk,
            idcarton=self.carton_hibrido,
            idpartida=partida,
            idbingo=bingo or partida.idbingo,
            estado_participacion=estado,
            indicevictoria=indice,
            es_asignacion_original=False,
            origen_asignacion=CartonPartidaBingo.ORIGEN_APLICACION,
            fechacreacion=ahora,
            fechavalidacion=ahora if indice is not None else None,
        )

    def _staff_request(self, path):
        request = self.factory.get(path)
        request.user = User(username="admin_reportes", is_staff=True)
        return request

    def test_reporte_partida_combina_historico_e_hibrido(self):
        filas = construir_filas_reporte_partida(
            self.partida_uno,
            [self.carton_historico],
            [self.participacion_uno],
        )

        self.assertEqual([fila["tipo"] for fila in filas], [TIPO_HISTORICO, TIPO_HIBRIDO])
        self.assertEqual(
            [fila["codigo_carton"] for fila in filas],
            ["P701-C-HIST", "B70-C-MAESTRO"],
        )

    def test_participacion_de_otra_ronda_no_aparece(self):
        filas = construir_filas_reporte_partida(
            self.partida_uno,
            [],
            [self.participacion_dos],
        )

        self.assertEqual(filas, [])

    def test_participacion_de_otro_bingo_produce_error_controlado(self):
        inconsistente = self._participacion(
            903,
            self.partida_uno,
            CartonPartidaBingo.ESTADO_PENDIENTE,
            bingo=self.otro_bingo,
        )

        with self.assertRaisesMessage(ReporteHibridoError, "no coinciden"):
            construir_filas_reporte_partida(
                self.partida_uno,
                [],
                [inconsistente],
            )

    def test_hibrido_usa_maestro_y_resultado_de_participacion(self):
        fila = construir_filas_reporte_partida(
            self.partida_uno,
            [],
            [self.participacion_uno],
        )[0]

        self.assertEqual(fila["matriznumeros"], self.carton_hibrido.matriznumeros)
        self.assertIs(fila["jugador"], self.jugador_hibrido)
        self.assertIs(fila["partida"], self.partida_uno)
        self.assertEqual(fila["indicevictoria"], 7)
        self.assertNotEqual(fila["indicevictoria"], self.carton_hibrido.indicevictoria)
        self.assertTrue(fila["es_ganador"])

    def test_maestro_aparece_en_reportes_de_dos_rondas(self):
        fila_uno = construir_filas_reporte_partida(
            self.partida_uno,
            [],
            [self.participacion_uno],
        )[0]
        fila_dos = construir_filas_reporte_partida(
            self.partida_dos,
            [],
            [self.participacion_dos],
        )[0]

        self.assertEqual(fila_uno["idcarton"], fila_dos["idcarton"])
        self.assertNotEqual(fila_uno["partida"].pk, fila_dos["partida"].pk)
        self.assertEqual(fila_uno["indicevictoria"], 7)
        self.assertIsNone(fila_dos["indicevictoria"])

    def test_pdf_partida_identifica_ambos_tipos_y_ganador_de_ronda(self):
        contenido = generar_pdf_reporte_partida(
            self.partida_uno,
            [self.carton_historico],
            [self.participacion_uno],
        )

        self.assertTrue(contenido.startswith(b"%PDF"))
        self.assertIn(b"P701-C-HIST", contenido)
        self.assertIn(b"B70-C-MAESTRO", contenido)
        self.assertIn(b"Ganador", contenido)

    def test_excel_partida_exporta_tipo_e_indice_de_participacion(self):
        contenido = generar_excel_cartones_partida(
            self.partida_uno,
            [self.carton_historico],
            [self.participacion_uno],
        )
        worksheet = load_workbook(BytesIO(contenido))["Cartones"]
        headers = [cell.value for cell in worksheet[1]]
        tipo = headers.index("Tipo de registro") + 1
        codigo = headers.index("Código de cartón") + 1
        indice = headers.index("Índice de victoria") + 1
        estado = headers.index("Estado de participación") + 1
        filas = {
            worksheet.cell(row=row, column=codigo).value: row
            for row in (2, 3)
        }

        fila_hibrida = filas[self.carton_hibrido.codigocarton]
        self.assertEqual(worksheet.cell(fila_hibrida, tipo).value, "Cartón de Bingo")
        self.assertEqual(worksheet.cell(fila_hibrida, indice).value, 7)
        self.assertEqual(worksheet.cell(fila_hibrida, estado).value, "Ganador")

    def test_resumen_bingo_no_duplica_maestro_y_cuenta_rondas(self):
        resumenes = construir_resumenes_cartones_bingo(
            self.bingo,
            [self.carton_historico, self.carton_hibrido, self.carton_hibrido],
            [self.participacion_uno, self.participacion_dos],
        )

        self.assertEqual(len(resumenes), 2)
        hibrido = next(item for item in resumenes if item["tipo"] == TIPO_HIBRIDO)
        self.assertEqual(hibrido["total_participaciones"], 2)
        self.assertEqual(hibrido["rondas_ganadas"], 1)
        self.assertEqual(hibrido["rondas_pendientes_activas"], 1)

    def test_excel_resumen_muestra_inventario_unico_y_rondas(self):
        contenido = generar_excel_resumen_bingo(
            self.bingo,
            [self.partida_uno, self.partida_dos],
            [self.carton_historico, self.carton_hibrido],
            [self.participacion_uno, self.participacion_dos],
        )
        workbook = load_workbook(BytesIO(contenido))
        inventario = workbook["Cartones del Bingo"]
        codigos = [
            inventario.cell(row, 2).value
            for row in range(4, inventario.max_row + 1)
        ]

        self.assertEqual(codigos.count(self.carton_hibrido.codigocarton), 1)
        fila_hibrida = codigos.index(self.carton_hibrido.codigocarton) + 4
        self.assertEqual(inventario.cell(fila_hibrida, 7).value, 2)
        self.assertEqual(inventario.cell(fila_hibrida, 8).value, 1)
        resumen_partidas = workbook["Resumen de partidas"]
        self.assertEqual(resumen_partidas.cell(19, 8).value, 1)
        self.assertEqual(resumen_partidas.cell(20, 8).value, 1)
        participaciones = workbook["Participaciones por ronda"]
        codigos_participacion = [
            participaciones.cell(row, 5).value
            for row in range(4, participaciones.max_row + 1)
        ]
        self.assertEqual(
            codigos_participacion.count(self.carton_hibrido.codigocarton),
            2,
        )
        # Verify no per-round monetary column exists
        headers_partidas = [
            cell.value for cell in resumen_partidas[18]
        ]
        self.assertNotIn("Recaudación total", headers_partidas)

    def test_vista_devuelve_400_para_participacion_inconsistente(self):
        inconsistente = self._participacion(
            904,
            self.partida_uno,
            CartonPartidaBingo.ESTADO_PENDIENTE,
            bingo=self.otro_bingo,
        )
        with (
            patch("apps.bingos.views.get_object_or_404", return_value=self.partida_uno),
            patch(
                "apps.bingos.views.Carton.objects.filter",
                return_value=FakeReportQuerySet([self.carton_historico]),
            ),
            patch(
                "apps.bingos.views.CartonPartidaBingo.objects.filter",
                return_value=FakeReportQuerySet([inconsistente]),
            ),
        ):
            response = partida_cartones_excel(
                self._staff_request("/partidas/701/cartones/excel/"),
                701,
            )

        self.assertEqual(response.status_code, 400)
        self.assertIn("no coinciden", response.content.decode())

    def test_vistas_pdf_y_excel_siguen_respondiendo(self):
        with (
            patch("apps.bingos.views.get_object_or_404", return_value=self.partida_uno),
            patch(
                "apps.bingos.views.Carton.objects.filter",
                return_value=FakeReportQuerySet([self.carton_historico]),
            ),
            patch(
                "apps.bingos.views.CartonPartidaBingo.objects.filter",
                return_value=FakeReportQuerySet([self.participacion_uno]),
            ),
        ):
            pdf = partida_reporte_pdf(
                self._staff_request("/partidas/701/reporte/pdf/"),
                701,
            )
            excel = partida_cartones_excel(
                self._staff_request("/partidas/701/cartones/excel/"),
                701,
            )

        self.assertEqual(pdf.status_code, 200)
        self.assertEqual(excel.status_code, 200)

    def test_rutas_de_reportes_conservan_resolucion(self):
        self.assertIs(
            resolve("/partidas/701/reporte/pdf/").func,
            partida_reporte_pdf,
        )
        self.assertIs(
            resolve("/partidas/701/cartones/excel/").func,
            partida_cartones_excel,
        )
        self.assertIs(
            resolve("/bingos/70/resumen/excel/").func,
            bingo_resumen_excel,
        )


class _CartonesConsultaQuerySet(list):
    _prefetch_related_lookups = ()

    def all(self):
        return self

    def select_related(self, *fields):
        return self

    def order_by(self, *fields):
        return self

    def filter(self, *args, **kwargs):
        return self

    def annotate(self, *args, **kwargs):
        return self

    def iterator(self):
        return iter(self)

    def count(self):
        return len(self)

    def first(self):
        return self[0] if self else None


@override_settings(
    STORAGES={
        "default": {
            "BACKEND": "django.core.files.storage.FileSystemStorage",
        },
        "staticfiles": {
            "BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage",
        },
    }
)
class VisibilidadCartonesHibridosAdminTests(SimpleTestCase):
    def setUp(self):
        self.factory = RequestFactory()
        self.jugador = Jugador(
            idjugador=41,
            aliasjugador="jugador_admin",
        )
        self.bingo = Bingo(
            idbingo=70,
            titulobingo="Bingo administrativo",
            fechaprogramadabingo=timezone.now(),
            tipobingo="Presencial",
            preciocarton=Decimal("5.00"),
            premiomayor=Decimal("100.00"),
            descripcionpremiomayor="Premio mayor",
            estadobingo="Programado",
        )
        self.partida = self._partida(701, "Ronda principal")
        self.otra_partida = self._partida(702, "Ronda final")
        self.carton_maestro = Carton(
            idcarton=801,
            idjugador=self.jugador,
            idbingo=self.bingo,
            idpartida=None,
            codigocarton="B70-C-MAESTRO",
            matriznumeros=serializar_matriz_carton_bingo(
                matriz_carton_prueba()
            ),
            preciopagado=Decimal("5.00"),
            fechacompra=timezone.now(),
            estadocarton="Vendido",
        )
        self.carton_maestro.total_participaciones = 2
        self.carton_heredado = Carton(
            idcarton=802,
            idjugador=self.jugador,
            idbingo=self.bingo,
            idpartida=self.partida,
            codigocarton="P701-C-HEREDADO",
            matriznumeros=self.carton_maestro.matriznumeros,
            preciopagado=Decimal("5.00"),
            fechacompra=timezone.now(),
            estadocarton="Vendido",
        )
        self.carton_heredado.total_participaciones = 0
        self.participacion = CartonPartidaBingo(
            idcartonpartidabingo=901,
            idcarton=self.carton_maestro,
            idpartida=self.partida,
            idbingo=self.bingo,
            estado_participacion=CartonPartidaBingo.ESTADO_GANADOR,
            indicevictoria=24,
            es_asignacion_original=False,
            origen_asignacion=CartonPartidaBingo.ORIGEN_APLICACION,
            fechacreacion=timezone.now(),
            fechavalidacion=timezone.now(),
        )

    def _partida(self, pk, nombre):
        return Partidabingo(
            idpartidabingo=pk,
            idbingo=self.bingo,
            nombreronda=nombre,
            valorefectivo=Decimal("25.00"),
            premiomaterial="",
            estadopartida=ESTADO_PARTIDA_PROGRAMADA,
            bolascantadas="[]",
            ultimabola=0,
            haydesempate=False,
            horainicio=timezone.now(),
        )

    def _request(self, path):
        request = self.factory.get(path)
        request.user = User(username="admin", is_staff=True)
        request.session = {}
        return request

    def test_detalle_ronda_muestra_participacion_hibrida_sin_vacio(self):
        self.partida.patronganador = "cuatro_esquinas"
        cartones = _CartonesConsultaQuerySet()
        with (
            patch(
                "apps.bingos.views.get_object_or_404",
                return_value=self.partida,
            ),
            patch(
                "apps.bingos.views.Carton.objects.filter",
                return_value=cartones,
            ) as filtrar_cartones,
            patch(
                "apps.bingos.views.obtener_participaciones_hibridas_partida",
                return_value=[self.participacion],
            ) as consultar_participaciones,
            patch.object(Carton, "save") as guardar_carton,
            patch.object(CartonPartidaBingo, "save") as guardar_participacion,
        ):
            response = partida_detalle(
                self._request("/partidas/701/"),
                self.partida.pk,
            )

        self.assertContains(response, self.carton_maestro.codigocarton, count=1)
        self.assertContains(response, "Cartón del Bingo")
        self.assertContains(response, "Patrón ganador")
        self.assertContains(response, "Cuatro esquinas")
        self.assertContains(response, CartonPartidaBingo.ESTADO_GANADOR)
        self.assertContains(response, "Índice de victoria: 24")
        self.assertNotContains(
            response,
            "No hay cartones registrados para esta partida.",
        )
        filtrar_cartones.assert_called_once_with(idpartida=self.partida)
        consultar_participaciones.assert_called_once_with(self.partida)
        guardar_carton.assert_not_called()
        guardar_participacion.assert_not_called()

    def test_detalle_ronda_conserva_carton_heredado(self):
        with (
            patch(
                "apps.bingos.views.get_object_or_404",
                return_value=self.partida,
            ),
            patch(
                "apps.bingos.views.Carton.objects.filter",
                return_value=_CartonesConsultaQuerySet(
                    [self.carton_heredado]
                ),
            ),
            patch(
                "apps.bingos.views.obtener_participaciones_hibridas_partida",
                return_value=[],
            ),
            patch.object(Carton, "save") as guardar_carton,
            patch.object(CartonPartidaBingo, "save") as guardar_participacion,
        ):
            response = partida_detalle(
                self._request("/partidas/701/"),
                self.partida.pk,
            )

        self.assertContains(response, self.carton_heredado.codigocarton)
        self.assertContains(response, "Cartón heredado por ronda")
        self.assertNotContains(
            response,
            "No hay cartones registrados para esta partida.",
        )
        guardar_carton.assert_not_called()
        guardar_participacion.assert_not_called()

    def test_detalle_bingo_muestra_maestro_una_sola_vez(self):
        cartones = _CartonesConsultaQuerySet([self.carton_maestro])
        with (
            patch(
                "apps.bingos.views.get_object_or_404",
                return_value=self.bingo,
            ),
            patch(
                "apps.bingos.views.Partidabingo.objects.filter",
                return_value=_CartonesConsultaQuerySet(
                    [self.partida, self.otra_partida]
                ),
            ),
            patch(
                "apps.bingos.views.Carton.objects.filter",
                return_value=cartones,
            ) as filtrar_cartones,
            patch.object(Carton, "save") as guardar_carton,
            patch.object(CartonPartidaBingo, "save") as guardar_participacion,
        ):
            response = bingo_detalle(
                self._request("/bingos/70/"),
                self.bingo.pk,
            )

        self.assertContains(response, self.carton_maestro.codigocarton, count=1)
        self.assertContains(response, "Cartón del Bingo")
        self.assertContains(response, "2 rondas")
        filtrar_cartones.assert_called_once_with(idbingo=self.bingo)
        guardar_carton.assert_not_called()
        guardar_participacion.assert_not_called()

    def test_listado_global_describe_maestro_y_carton_heredado(self):
        cartones = _CartonesConsultaQuerySet(
            [self.carton_maestro, self.carton_heredado]
        )
        with (
            patch(
                "apps.bingos.views.Carton.objects.select_related",
                return_value=cartones,
            ),
            patch.object(Carton, "save") as guardar_carton,
            patch.object(CartonPartidaBingo, "save") as guardar_participacion,
        ):
            response = cartones_lista(self._request("/cartones/"))

        self.assertContains(response, self.carton_maestro.codigocarton)
        self.assertContains(response, self.bingo.titulobingo, count=2)
        self.assertContains(response, "Cartón del Bingo")
        self.assertContains(response, "2 rondas")
        self.assertContains(response, self.carton_heredado.codigocarton)
        self.assertContains(response, "Cartón heredado por ronda")
        self.assertContains(response, self.partida.nombreronda)
        guardar_carton.assert_not_called()
        guardar_participacion.assert_not_called()


@override_settings(
    STORAGES={
        "default": {
            "BACKEND": "django.core.files.storage.FileSystemStorage",
        },
        "staticfiles": {
            "BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage",
        },
    }
)
class CartonesPrivadosPublicosHibridosTests(SimpleTestCase):
    def setUp(self):
        self.factory = RequestFactory()
        self.bingo = Bingo(idbingo=70, titulobingo="Bingo por maestro")
        self.otro_bingo = Bingo(idbingo=71, titulobingo="Otro Bingo")
        self.jugador = Jugador(
            idjugador=41,
            aliasjugador="dueno_privado",
            correojugador="privado@example.com",
        )
        self.otro_jugador = Jugador(
            idjugador=42,
            aliasjugador="otro_jugador",
            correojugador="otro@example.com",
        )
        ahora = timezone.now()
        self.partida_activa = self._partida(
            701,
            self.bingo,
            ESTADO_PARTIDA_EN_CURSO,
            [1, 16, 31],
            ahora,
        )
        self.partida_pendiente = self._partida(
            702,
            self.bingo,
            ESTADO_PARTIDA_EN_ESPERA,
            [1],
            ahora + timedelta(minutes=1),
        )
        self.partida_finalizada = self._partida(
            703,
            self.bingo,
            ESTADO_PARTIDA_FINALIZADA,
            [1, 16],
            ahora - timedelta(days=1),
        )
        self.partida_otro_bingo = self._partida(
            704,
            self.otro_bingo,
            ESTADO_PARTIDA_EN_CURSO,
            [1, 16, 31, 46],
            ahora,
        )
        self.carton = Carton(
            idcarton=801,
            idjugador=self.jugador,
            idbingo=self.bingo,
            idpartida=None,
            codigocarton="B70-C-HIBRIDO",
            matriznumeros=serializar_matriz_carton_bingo(
                matriz_carton_prueba()
            ),
            indicevictoria=999,
            preciopagado=Decimal("9876.54"),
            estadocarton="Vendido",
        )
        self.participacion_activa = self._participacion(
            901,
            self.carton,
            self.partida_activa,
            CartonPartidaBingo.ESTADO_EN_JUEGO,
            indice=7,
        )
        self.participacion_pendiente = self._participacion(
            902,
            self.carton,
            self.partida_pendiente,
            CartonPartidaBingo.ESTADO_PENDIENTE,
        )
        self.participacion_finalizada = self._participacion(
            903,
            self.carton,
            self.partida_finalizada,
            CartonPartidaBingo.ESTADO_CERRADO,
        )
        self.participaciones = [
            self.participacion_activa,
            self.participacion_pendiente,
            self.participacion_finalizada,
        ]
        self.carton.participaciones_hibridas_cargadas = self.participaciones
        self.carton_historico = Carton(
            idcarton=802,
            idjugador=self.jugador,
            idbingo=self.bingo,
            idpartida=self.partida_activa,
            codigocarton="P701-C-HISTORICO",
            matriznumeros=self.carton.matriznumeros,
            indicevictoria=4,
            estadocarton="Vendido",
        )

    def _partida(self, pk, bingo, estado, bolas, fecha):
        return Partidabingo(
            idpartidabingo=pk,
            idbingo=bingo,
            nombreronda=f"Ronda {pk}",
            estadopartida=estado,
            bolascantadas=serializar_bolas_cantadas(bolas),
            ultimabola=bolas[-1] if bolas else 0,
            horainicio=fecha,
            horafin=fecha if estado == ESTADO_PARTIDA_FINALIZADA else None,
            haydesempate=False,
        )

    def _participacion(self, pk, carton, partida, estado, indice=None):
        return CartonPartidaBingo(
            idcartonpartidabingo=pk,
            idcarton=carton,
            idpartida=partida,
            idbingo=partida.idbingo,
            estado_participacion=estado,
            indicevictoria=indice,
            es_asignacion_original=False,
            origen_asignacion=CartonPartidaBingo.ORIGEN_APLICACION,
            fechavalidacion=timezone.now() if indice else None,
        )

    def _request(self, path, query=None, jugador=False):
        request = self.factory.get(path, query or {})
        request.user = (
            User(username=self.jugador.aliasjugador)
            if jugador
            else AnonymousUser()
        )
        if getattr(request.user, "is_authenticated", False):
            request.user.pk = 50
        request.session = {}
        return request

    def test_mis_cartones_lista_un_maestro_una_sola_vez(self):
        capturado = {}

        def fake_render(_request, _template, contexto):
            capturado.update(contexto)
            return HttpResponse(status=200)

        with (
            patch(
                "apps.jugadores.services.obtener_jugador_autenticado",
                return_value=self.jugador,
            ),
            patch(
                "apps.bingos.views.Carton.objects.filter",
                return_value=_CartonesConsultaQuerySet([self.carton]),
            ),
            patch("apps.bingos.views.render", side_effect=fake_render),
        ):
            response = mis_cartones(
                self._request("/mis-cartones/", jugador=True)
            )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(capturado["cartones_resumen"]), 1)
        self.assertEqual(capturado["cartones_resumen"][0]["total_rondas"], 3)

    def test_carton_historico_conserva_resumen_heredado(self):
        resumen = _resumen_carton_privado(self.carton_historico)

        self.assertEqual(resumen["tipo_carton"], "historico")
        self.assertEqual(resumen["partida"], self.partida_activa)
        self.assertEqual(resumen["total_rondas"], 1)

    def test_detalle_hibrido_selecciona_pareja_carton_ronda(self):
        seleccionada = _seleccionar_participacion_hibrida(
            self.participaciones,
            str(self.partida_pendiente.pk),
        )

        self.assertIs(seleccionada, self.participacion_pendiente)
        self.assertEqual(seleccionada.idcarton_id, self.carton.pk)

    def test_rechaza_ronda_perteneciente_a_otro_carton(self):
        otro_carton = Carton(
            idcarton=803,
            idbingo=self.bingo,
            idjugador=self.otro_jugador,
            idpartida=None,
        )
        otra_partida = self._partida(
            705,
            self.bingo,
            ESTADO_PARTIDA_EN_CURSO,
            [],
            timezone.now(),
        )
        self._participacion(
            904,
            otro_carton,
            otra_partida,
            CartonPartidaBingo.ESTADO_PENDIENTE,
        )

        with self.assertRaises(Http404):
            _seleccionar_participacion_hibrida(
                self.participaciones,
                str(otra_partida.pk),
            )

    def test_rechaza_ronda_de_otro_bingo(self):
        with self.assertRaises(Http404):
            _seleccionar_participacion_hibrida(
                self.participaciones,
                str(self.partida_otro_bingo.pk),
            )

    def test_consulta_limita_selector_a_participaciones_del_maestro(self):
        del self.carton.participaciones_hibridas_cargadas
        consulta = Mock()
        consulta.select_related.return_value.order_by.return_value = (
            self.participaciones
        )
        with patch(
            "apps.bingos.views.CartonPartidaBingo.objects.filter",
            return_value=consulta,
        ) as filtrar:
            resultado = _participaciones_hibridas_carton(self.carton)

        self.assertEqual(resultado, self.participaciones)
        filtrar.assert_called_once_with(idcarton=self.carton)

    def test_participacion_de_otro_bingo_se_rechaza(self):
        inconsistente = self._participacion(
            905,
            self.carton,
            self.partida_otro_bingo,
            CartonPartidaBingo.ESTADO_PENDIENTE,
        )
        self.carton.participaciones_hibridas_cargadas = [inconsistente]

        with self.assertRaisesMessage(CartonPublicoError, "mismo Bingo"):
            _participaciones_hibridas_carton(self.carton)

    def test_ronda_por_defecto_prioriza_la_activa(self):
        seleccionada = _seleccionar_participacion_hibrida(
            [self.participacion_finalizada, self.participacion_pendiente,
             self.participacion_activa],
            None,
        )

        self.assertIs(seleccionada, self.participacion_activa)

    def test_progreso_usa_bolas_de_la_ronda_seleccionada(self):
        datos_activa = _preparar_datos_carton_hibrido(
            self.carton,
            self.partida_activa,
        )
        datos_pendiente = _preparar_datos_carton_hibrido(
            self.carton,
            self.partida_pendiente,
        )

        self.assertEqual(datos_activa["numeros_marcados"], 3)
        self.assertEqual(datos_pendiente["numeros_marcados"], 1)

    def test_detalle_privado_no_permite_carton_de_otro_jugador(self):
        consulta = _CartonesConsultaQuerySet()
        with (
            patch(
                "apps.jugadores.services.obtener_jugador_autenticado",
                return_value=self.jugador,
            ),
            patch(
                "apps.bingos.views.Carton.objects.select_related",
                return_value=consulta,
            ),
        ):
            with self.assertRaises(Http404):
                mi_carton_detalle(
                    self._request(
                        "/mis-cartones/AJENO/",
                        jugador=True,
                    ),
                    "AJENO",
                )

    def test_acceso_publico_hibrido_no_expone_dueno_ni_precio(self):
        consulta = Mock()
        consulta.filter.return_value.first.return_value = self.carton
        with patch(
            "apps.bingos.views.Carton.objects.select_related",
            return_value=consulta,
        ):
            response = carton_publico(
                self._request(f"/juego/cartones/{self.carton.codigocarton}/"),
                self.carton.codigocarton,
            )

        contenido = response.content.decode()
        self.assertNotIn(self.jugador.aliasjugador, contenido)
        self.assertNotIn(self.jugador.correojugador, contenido)
        self.assertNotIn("9876.54", contenido)
        self.assertNotIn("Precio pagado", contenido)

    def test_acceso_publico_historico_usa_flujo_existente(self):
        consulta = Mock()
        consulta.filter.return_value.first.return_value = self.carton_historico
        with (
            patch(
                "apps.bingos.views.Carton.objects.select_related",
                return_value=consulta,
            ),
            patch(
                "apps.bingos.views.preparar_datos_carton_jugador",
                return_value={
                    **_preparar_datos_carton_hibrido(
                        self.carton_historico,
                        self.partida_activa,
                    )
                },
            ) as legado,
            patch(
                "apps.bingos.views.render",
                return_value=HttpResponse(status=200),
            ),
        ):
            response = carton_publico(
                self._request(
                    f"/juego/cartones/{self.carton_historico.codigocarton}/"
                ),
                self.carton_historico.codigocarton,
            )

        self.assertEqual(response.status_code, 200)
        legado.assert_called_once_with(self.carton_historico)

    def test_acceso_publico_hibrido_rechaza_partida_sin_participacion(self):
        consulta = Mock()
        consulta.filter.return_value.first.return_value = self.carton
        with patch(
            "apps.bingos.views.Carton.objects.select_related",
            return_value=consulta,
        ):
            with self.assertRaises(Http404):
                carton_publico(
                    self._request(
                        f"/juego/cartones/{self.carton.codigocarton}/",
                        {"partida": self.partida_otro_bingo.pk},
                    ),
                    self.carton.codigocarton,
                )

    def test_hibrido_usa_resultado_de_participacion_no_del_maestro(self):
        capturado = {}
        consulta = Mock()
        consulta.filter.return_value.first.return_value = self.carton

        def fake_render(_request, _template, contexto):
            capturado.update(contexto)
            return HttpResponse(status=200)

        with (
            patch(
                "apps.bingos.views.Carton.objects.select_related",
                return_value=consulta,
            ),
            patch("apps.bingos.views.render", side_effect=fake_render),
        ):
            carton_publico(
                self._request(f"/juego/cartones/{self.carton.codigocarton}/"),
                self.carton.codigocarton,
            )

        self.assertIsNone(self.carton.idpartida)
        self.assertEqual(self.carton.indicevictoria, 999)
        self.assertEqual(
            capturado["participacion_seleccionada"].indicevictoria,
            7,
        )

    def test_rutas_privadas_y_publicas_conservan_nombres(self):
        self.assertIs(resolve("/mis-cartones/").func, mis_cartones)
        self.assertIs(
            resolve(f"/mis-cartones/{self.carton.codigocarton}/").func,
            mi_carton_detalle,
        )
        self.assertIs(
            resolve(f"/juego/cartones/{self.carton.codigocarton}/").func,
            carton_publico,
        )


class RecaudacionRegistradaDuplicacionTests(SimpleTestCase):
    """Pruebas obligatorias de la Fase 2B: recaudación sin duplicación."""

    def setUp(self):
        self.bingo = Bingo(
            idbingo=90,
            titulobingo="Bingo recaudación",
            fechaprogramadabingo=timezone.now(),
            tipobingo="Virtual",
            preciocarton=Decimal("5.00"),
            premiomayor=Decimal("50.00"),
            descripcionpremiomayor="Premio",
            estadobingo="Programado",
        )
        self.otro_bingo = Bingo(
            idbingo=91,
            titulobingo="Otro Bingo",
            fechaprogramadabingo=timezone.now(),
            tipobingo="Virtual",
            preciocarton=Decimal("8.00"),
            premiomayor=Decimal("80.00"),
            descripcionpremiomayor="Premio",
            estadobingo="Programado",
        )
        self.jugador = Jugador(idjugador=50, aliasjugador="jugador_recaud")
        ahora = timezone.now()
        self.partidas = [
            self._partida(901, self.bingo, ESTADO_PARTIDA_FINALIZADA, ahora),
            self._partida(902, self.bingo, ESTADO_PARTIDA_EN_CURSO, ahora),
            self._partida(903, self.bingo, ESTADO_PARTIDA_PROGRAMADA, ahora),
        ]
        # Tres cartones maestros de $5 cada uno
        self.cartones = [
            self._carton_maestro(501, "B90-C-A01", Decimal("5.00")),
            self._carton_maestro(502, "B90-C-A02", Decimal("5.00")),
            self._carton_maestro(503, "B90-C-A03", Decimal("5.00")),
        ]
        # Cada cartón participa en las 3 rondas: 9 participaciones
        self.participaciones = []
        pk_counter = 9000
        for carton in self.cartones:
            for partida in self.partidas:
                pk_counter += 1
                self.participaciones.append(
                    self._participacion(pk_counter, carton, partida)
                )

    def _partida(self, pk, bingo, estado, ahora):
        return Partidabingo(
            idpartidabingo=pk,
            idbingo=bingo,
            nombreronda=f"Ronda {pk}",
            valorefectivo=Decimal("25.00"),
            premiomaterial="Premio",
            estadopartida=estado,
            bolascantadas=serializar_bolas_cantadas([1, 16, 31]),
            ultimabola=31,
            haydesempate=False,
            horainicio=ahora,
            horafin=ahora if estado == ESTADO_PARTIDA_FINALIZADA else None,
        )

    def _carton_maestro(self, pk, codigo, precio):
        return Carton(
            idcarton=pk,
            idjugador=self.jugador,
            idbingo=self.bingo,
            idpartida=None,
            codigocarton=codigo,
            matriznumeros=serializar_matriz_carton_bingo(matriz_carton_prueba()),
            indicevictoria=None,
            preciopagado=precio,
            fechacompra=timezone.now(),
            estadocarton="Vendido",
        )

    def _participacion(self, pk, carton, partida, estado=None):
        return CartonPartidaBingo(
            idcartonpartidabingo=pk,
            idcarton=carton,
            idpartida=partida,
            idbingo=partida.idbingo,
            estado_participacion=(
                estado or CartonPartidaBingo.ESTADO_PENDIENTE
            ),
            indicevictoria=None,
            es_asignacion_original=False,
            origen_asignacion=CartonPartidaBingo.ORIGEN_APLICACION,
            fechacreacion=timezone.now(),
            fechavalidacion=None,
        )

    def test_01_un_carton_de_5_con_tres_participaciones_aporta_5(self):
        """Un cartón maestro de $5 con tres participaciones aporta $5."""
        carton = self.cartones[0]
        total = calcular_recaudacion_registrada([carton])
        self.assertEqual(total, Decimal("5.00"))

    def test_02_tres_cartones_de_5_con_tres_rondas_producen_15_no_45(self):
        """Tres cartones maestros de $5 con tres rondas producen $15."""
        total = calcular_recaudacion_registrada(self.cartones)
        self.assertEqual(total, Decimal("15.00"))
        # Verificar en el Excel resumen
        contenido = generar_excel_resumen_bingo(
            self.bingo,
            self.partidas,
            self.cartones,
            self.participaciones,
        )
        workbook = load_workbook(BytesIO(contenido))
        worksheet = workbook["Resumen de partidas"]
        resumen = {
            worksheet.cell(row=row, column=1).value: worksheet.cell(row=row, column=2).value
            for row in range(4, 13)
            if worksheet.cell(row=row, column=1).value is not None
        }
        self.assertEqual(resumen["Total recaudado"], 15)
        self.assertNotIn("Recaudación total", resumen)

    def test_03_excel_ronda_no_contiene_total_recaudado(self):
        """El Excel de ronda no contiene 'Total recaudado'."""
        contenido = generar_excel_cartones_partida(
            self.partidas[0],
            [],
            self.participaciones[:3],
        )
        workbook = load_workbook(BytesIO(contenido))
        worksheet = workbook["Cartones"]
        todas_celdas = [
            worksheet.cell(row=row, column=col).value
            for row in range(1, worksheet.max_row + 1)
            for col in range(1, worksheet.max_column + 1)
        ]
        self.assertNotIn("Total recaudado", todas_celdas)

    def test_04_excel_ronda_contiene_nota_operativa(self):
        """El Excel de ronda contiene la nota operativa."""
        contenido = generar_excel_cartones_partida(
            self.partidas[0],
            [],
            self.participaciones[:3],
        )
        workbook = load_workbook(BytesIO(contenido))
        worksheet = workbook["Cartones"]
        textos = [
            str(worksheet.cell(row=row, column=1).value or "")
            for row in range(1, worksheet.max_row + 1)
        ]
        nota_encontrada = any(
            "no representa la liquidación general del Bingo" in texto
            for texto in textos
        )
        self.assertTrue(
            nota_encontrada,
            "No se encontró la nota operativa en el Excel de ronda.",
        )

    def test_05_excel_resumen_sin_columna_monetaria_repetida_por_ronda(self):
        """El Excel resumen no muestra columna monetaria repetida por ronda."""
        contenido = generar_excel_resumen_bingo(
            self.bingo,
            self.partidas,
            self.cartones,
            self.participaciones,
        )
        workbook = load_workbook(BytesIO(contenido))
        worksheet = workbook["Resumen de partidas"]
        headers = [cell.value for cell in worksheet[18]]
        self.assertNotIn("Recaudación total", headers)

    def test_06_excel_resumen_muestra_total_recaudado_una_vez(self):
        """El Excel resumen muestra el total recaudado una sola vez."""
        contenido = generar_excel_resumen_bingo(
            self.bingo,
            self.partidas,
            self.cartones,
            self.participaciones,
        )
        workbook = load_workbook(BytesIO(contenido))
        worksheet = workbook["Resumen de partidas"]
        etiquetas = [
            worksheet.cell(row=row, column=1).value
            for row in range(1, worksheet.max_row + 1)
        ]
        self.assertEqual(
            etiquetas.count("Total recaudado"),
            1,
            "Debe aparecer exactamente una vez.",
        )

    def test_07_carton_de_otro_bingo_no_afecta_total(self):
        """Un cartón de otro Bingo no afecta el total."""
        carton_otro = Carton(
            idcarton=599,
            idjugador=self.jugador,
            idbingo=self.otro_bingo,
            idpartida=None,
            codigocarton="B91-C-OTRO",
            matriznumeros=serializar_matriz_carton_bingo(
                matriz_carton_prueba()
            ),
            preciopagado=Decimal("100.00"),
            estadocarton="Vendido",
        )
        total = calcular_recaudacion_registrada(
            self.cartones + [carton_otro]
        )
        self.assertEqual(total, Decimal("115.00"))
        # Pero el resumen del Bingo 90 solo debería tener los del Bingo 90
        resumenes = construir_resumenes_cartones_bingo(
            self.bingo,
            self.cartones,
            self.participaciones,
        )
        total_bingo = calcular_recaudacion_registrada(resumenes)
        self.assertEqual(total_bingo, Decimal("15.00"))

    def test_08_precio_null_no_rompe_calculo_ni_incrementa_total(self):
        """Un precio NULL no rompe el cálculo ni incrementa el total."""
        carton_sin_precio = self._carton_maestro(504, "B90-C-NULL", None)
        todos = self.cartones + [carton_sin_precio]
        total = calcular_recaudacion_registrada(todos)
        self.assertEqual(total, Decimal("15.00"))

    def test_09_varias_participaciones_no_multiplican_importe(self):
        """Varias participaciones del mismo cartón no multiplican su importe."""
        # Pasar el mismo cartón como resumen dict múltiples veces
        resumenes_con_duplicados = [
            {"idcarton": 501, "precio_pagado": Decimal("5.00")},
            {"idcarton": 501, "precio_pagado": Decimal("5.00")},
            {"idcarton": 501, "precio_pagado": Decimal("5.00")},
        ]
        total = calcular_recaudacion_registrada(resumenes_con_duplicados)
        self.assertEqual(total, Decimal("5.00"))

    def test_10_reportes_existentes_sin_riesgo_continuan_funcionando(self):
        """Los reportes existentes sin riesgo continúan funcionando."""
        # PDF no calcula recaudación, should still work
        partida = self.partidas[0]
        pdf = generar_pdf_reporte_partida(
            partida,
            [],
            self.participaciones[:3],
        )
        self.assertTrue(pdf.startswith(b"%PDF"))

        # Excel de ronda still works
        excel = generar_excel_cartones_partida(
            partida,
            [],
            self.participaciones[:3],
        )
        workbook = load_workbook(BytesIO(excel))
        self.assertIn("Cartones", workbook.sheetnames)

        # Excel resumen still works
        contenido = generar_excel_resumen_bingo(
            self.bingo,
            self.partidas,
            self.cartones,
            self.participaciones,
        )
        workbook = load_workbook(BytesIO(contenido))
        self.assertIn("Resumen de partidas", workbook.sheetnames)
        self.assertIn("Cartones del Bingo", workbook.sheetnames)
        self.assertIn("Participaciones por ronda", workbook.sheetnames)

    def test_nota_gastos_adicionales_en_resumen(self):
        """El Excel resumen incluye la nota sobre gastos."""
        contenido = generar_excel_resumen_bingo(
            self.bingo,
            self.partidas,
            self.cartones,
            self.participaciones,
        )
        workbook = load_workbook(BytesIO(contenido))
        worksheet = workbook["Resumen de partidas"]
        textos = [
            str(worksheet.cell(row=row, column=1).value or "")
            for row in range(1, worksheet.max_row + 1)
        ]
        self.assertTrue(
            any("Gastos adicionales: no registrados" in t for t in textos),
            "No se encontró la nota de gastos.",
        )

    def test_nota_recaudacion_por_carton_maestro_en_resumen(self):
        """El Excel resumen incluye la nota de recaudación por cartón maestro."""
        contenido = generar_excel_resumen_bingo(
            self.bingo,
            self.partidas,
            self.cartones,
            self.participaciones,
        )
        workbook = load_workbook(BytesIO(contenido))
        worksheet = workbook["Resumen de partidas"]
        textos = [
            str(worksheet.cell(row=row, column=1).value or "")
            for row in range(1, worksheet.max_row + 1)
        ]
        self.assertTrue(
            any(
                "se calcula por cartón maestro vendido" in t
                and "no generan un cobro adicional" in t
                for t in textos
            ),
            "No se encontró la nota sobre cálculo por cartón maestro.",
        )

    def test_excel_resumen_tiene_secciones_y_formato_profesional(self):
        contenido = generar_excel_resumen_bingo(
            self.bingo,
            self.partidas,
            self.cartones,
            self.participaciones,
        )
        workbook = load_workbook(BytesIO(contenido))
        resumen = workbook["Resumen de partidas"]
        inventario = workbook["Cartones del Bingo"]
        participaciones = workbook["Participaciones por ronda"]

        self.assertEqual(resumen["A1"].value, "Resumen del Bingo")
        self.assertTrue(resumen["A1"].font.bold)
        self.assertEqual(resumen["A1"].font.sz, 18)
        self.assertEqual(resumen.freeze_panes, "A19")
        self.assertTrue(resumen.auto_filter.ref.startswith("A18:L"))
        self.assertEqual(resumen["A18"].alignment.horizontal, "center")
        self.assertEqual(resumen["B7"].number_format, '"$"#,##0.00')
        self.assertGreater(resumen.column_dimensions["A"].width, 12)

        self.assertEqual(
            inventario["A1"].value,
            "C. Cartones maestros vendidos",
        )
        self.assertEqual(inventario.freeze_panes, "A4")
        self.assertTrue(inventario.auto_filter.ref.startswith("A3:J"))
        self.assertEqual(
            participaciones["A1"].value,
            "D. Participaciones por ronda",
        )
        self.assertEqual(participaciones.freeze_panes, "A4")
        self.assertTrue(participaciones.auto_filter.ref.startswith("A3:I"))

        texto_visible = " ".join(
            str(cell.value or "").lower()
            for worksheet in workbook.worksheets
            for row in worksheet.iter_rows()
            for cell in row
        )
        for termino_interno in (
            "híbrido",
            "heredado",
            "histórico",
            "flujo antiguo",
            "cartonpartidabingo",
        ):
            self.assertNotIn(termino_interno, texto_visible)

    def test_seis_maestros_de_cinco_recaudan_treinta_sin_duplicarse(self):
        cartones = [
            self._carton_maestro(
                600 + indice,
                f"B90-C-{indice:02d}",
                Decimal("5.00"),
            )
            for indice in range(1, 7)
        ]
        participaciones = []
        pk = 10000
        for carton in cartones:
            for partida in self.partidas[:2]:
                pk += 1
                participaciones.append(
                    self._participacion(pk, carton, partida)
                )

        contenido = generar_excel_resumen_bingo(
            self.bingo,
            self.partidas[:2],
            cartones,
            participaciones,
        )
        workbook = load_workbook(BytesIO(contenido))
        resumen = workbook["Resumen de partidas"]
        detalle = workbook["Participaciones por ronda"]

        self.assertEqual(resumen["B6"].value, 6)
        self.assertEqual(resumen["B7"].value, 30)
        self.assertEqual(resumen["B9"].value, 12)
        codigos = [
            detalle.cell(row=row, column=5).value
            for row in range(4, detalle.max_row + 1)
        ]
        for carton in cartones:
            self.assertEqual(codigos.count(carton.codigocarton), 2)


class PreliquidacionFinancieraServicioTests(SimpleTestCase):
    def setUp(self):
        self.bingo = Bingo(
            idbingo=120,
            titulobingo="Bingo preliquidacion",
            fechaprogramadabingo=timezone.now(),
            tipobingo="Virtual",
            preciocarton=Decimal("5.00"),
            premiomayor=Decimal("50.00"),
            descripcionpremiomayor="Premio",
            estadobingo="Programado",
        )
        self.otro_bingo = Bingo(
            idbingo=121,
            titulobingo="Otro Bingo",
            fechaprogramadabingo=timezone.now(),
            tipobingo="Virtual",
            preciocarton=Decimal("5.00"),
            premiomayor=Decimal("50.00"),
            descripcionpremiomayor="Premio",
            estadobingo="Programado",
        )
        self.jugador = Jugador(idjugador=50, aliasjugador="jugador_preliq")
        ahora = timezone.now()
        self.partidas = [
            self._partida(
                1201,
                self.bingo,
                ESTADO_PARTIDA_FINALIZADA,
                Decimal("25.00"),
                "Canasta",
                ahora,
            ),
            self._partida(
                1202,
                self.bingo,
                ESTADO_PARTIDA_PROGRAMADA,
                Decimal("50.00"),
                "",
                ahora,
            ),
            self._partida(
                1203,
                self.bingo,
                ESTADO_PARTIDA_EN_CURSO,
                Decimal("99.00"),
                "Bono",
                ahora,
            ),
            self._partida(
                1204,
                self.bingo,
                ESTADO_PARTIDA_CANCELADA,
                Decimal("70.00"),
                "",
                ahora,
            ),
        ]
        self.cartones = [
            self._carton_maestro(501, "B120-C-01", Decimal("5.00")),
            self._carton_maestro(502, "B120-C-02", Decimal("5.00")),
            self._carton_maestro(503, "B120-C-03", Decimal("5.00")),
            self._carton_maestro(503, "B120-C-03-DUP", Decimal("5.00")),
            self._carton_maestro(
                504,
                "B120-C-DISP",
                Decimal("5.00"),
                estado="Disponible",
            ),
            self._carton_historico(505, "P1201-C-HIST", Decimal("5.00")),
            self._carton_maestro(
                506,
                "B121-C-OTRO",
                Decimal("100.00"),
                bingo=self.otro_bingo,
            ),
        ]
        self.participaciones = [
            self._participacion(9000 + indice, carton, partida)
            for indice, (carton, partida) in enumerate(
                (
                    (carton, partida)
                    for carton in self.cartones[:3]
                    for partida in self.partidas[:3]
                ),
                start=1,
            )
        ]

    def _partida(self, pk, bingo, estado, valor, premio_material, ahora):
        return Partidabingo(
            idpartidabingo=pk,
            idbingo=bingo,
            nombreronda=f"Ronda {pk}",
            valorefectivo=valor,
            premiomaterial=premio_material,
            estadopartida=estado,
            bolascantadas="[]",
            ultimabola=0,
            haydesempate=False,
            horainicio=ahora,
            horafin=ahora if estado == ESTADO_PARTIDA_FINALIZADA else None,
        )

    def _carton_maestro(self, pk, codigo, precio, estado="Vendido", bingo=None):
        return Carton(
            idcarton=pk,
            idjugador=self.jugador,
            idbingo=bingo or self.bingo,
            idpartida=None,
            codigocarton=codigo,
            matriznumeros=serializar_matriz_carton_bingo(matriz_carton_prueba()),
            indicevictoria=None,
            preciopagado=precio,
            fechacompra=timezone.now(),
            estadocarton=estado,
        )

    def _carton_historico(self, pk, codigo, precio):
        return Carton(
            idcarton=pk,
            idjugador=self.jugador,
            idbingo=self.bingo,
            idpartida=self.partidas[0],
            codigocarton=codigo,
            matriznumeros=serializar_matriz_carton_bingo(matriz_carton_prueba()),
            indicevictoria=None,
            preciopagado=precio,
            fechacompra=timezone.now(),
            estadocarton="Vendido",
        )

    def _participacion(self, pk, carton, partida):
        return CartonPartidaBingo(
            idcartonpartidabingo=pk,
            idcarton=carton,
            idpartida=partida,
            idbingo=partida.idbingo,
            estado_participacion=CartonPartidaBingo.ESTADO_PENDIENTE,
            indicevictoria=None,
            es_asignacion_original=False,
            origen_asignacion=CartonPartidaBingo.ORIGEN_APLICACION,
            fechacreacion=timezone.now(),
            fechavalidacion=None,
        )

    def _preliquidacion(self):
        return construir_preliquidacion_financiera_bingo(
            self.bingo,
            cartones=self.cartones,
            partidas=self.partidas,
        )

    def test_tres_cartones_de_cinco_no_se_multiplican_por_tres_rondas(self):
        datos = self._preliquidacion()

        self.assertEqual(datos["cartones_vendidos_unicos"], 3)
        self.assertEqual(datos["recaudacion_registrada"], Decimal("15.00"))
        self.assertNotEqual(datos["recaudacion_registrada"], Decimal("45.00"))

    def test_solo_rondas_finalizadas_cuentan_como_premio_efectivo(self):
        datos = self._preliquidacion()

        self.assertEqual(
            datos["premios_efectivo_finalizados"],
            Decimal("25.00"),
        )

    def test_premios_materiales_y_gastos_quedan_pendientes(self):
        datos = self._preliquidacion()

        self.assertEqual(
            [premio["descripcion"] for premio in datos["premios_materiales"]],
            ["Canasta", "Bono"],
        )
        self.assertTrue(
            all(
                premio["costo"] == "Costo pendiente de registrar"
                for premio in datos["premios_materiales"]
            )
        )
        self.assertEqual(datos["gastos_operativos"], "Pendientes de registrar")
        self.assertIn(
            "Los costos de premios materiales todavía no están registrados.",
            datos["alertas"],
        )
        self.assertIn(
            "Los gastos operativos todavía no están registrados.",
            datos["alertas"],
        )

    def test_resultado_provisional_resta_solo_premios_en_efectivo_finalizados(self):
        datos = self._preliquidacion()

        self.assertEqual(datos["resultado_provisional"], Decimal("-10.00"))

    def test_estado_de_rondas_detecta_pendientes_y_terminales(self):
        datos = self._preliquidacion()

        self.assertEqual(
            datos["estado_rondas"],
            {
                "total": 4,
                "finalizadas": 1,
                "canceladas": 1,
                "pendientes": 2,
            },
        )
        self.assertIn(
            "El Bingo todavía no está listo para un cierre financiero: "
            "existen rondas pendientes.",
            datos["alertas"],
        )


@override_settings(
    STORAGES={
        "default": {
            "BACKEND": "django.core.files.storage.FileSystemStorage",
        },
        "staticfiles": {
            "BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage",
        },
    }
)
class PreliquidacionFinancieraVistaTests(SimpleTestCase):
    def setUp(self):
        self.factory = RequestFactory()
        self.bingo = Bingo(
            idbingo=130,
            titulobingo="Bingo vista preliquidacion",
            fechaprogramadabingo=timezone.now(),
            tipobingo="Virtual",
            preciocarton=Decimal("5.00"),
            premiomayor=Decimal("50.00"),
            descripcionpremiomayor="Premio",
            estadobingo="Programado",
        )
        self.jugador = Jugador(idjugador=60, aliasjugador="jugador_vista")
        ahora = timezone.now()
        self.partida = Partidabingo(
            idpartidabingo=1301,
            idbingo=self.bingo,
            nombreronda="Ronda vista",
            valorefectivo=Decimal("25.00"),
            premiomaterial="Canasta",
            estadopartida=ESTADO_PARTIDA_FINALIZADA,
            bolascantadas="[]",
            ultimabola=0,
            haydesempate=False,
            horainicio=ahora,
            horafin=ahora,
        )
        self.carton = Carton(
            idcarton=13001,
            idjugador=self.jugador,
            idbingo=self.bingo,
            idpartida=None,
            codigocarton="B130-C-01",
            matriznumeros=serializar_matriz_carton_bingo(matriz_carton_prueba()),
            indicevictoria=None,
            preciopagado=Decimal("5.00"),
            fechacompra=ahora,
            estadocarton="Vendido",
        )
        self.carton.total_participaciones = 1
        self.datos = construir_preliquidacion_financiera_bingo(
            self.bingo,
            cartones=[self.carton],
            partidas=[self.partida],
        )

    def _request(self, path, user=None, method="get"):
        request = getattr(self.factory, method)(path)
        request.user = user if user is not None else self._admin()
        request.session = {}
        return request

    def _admin(self):
        return User(username="admin_preliq", is_staff=True)

    def _jugador(self):
        return User(username="jugador_preliq", is_staff=False, is_superuser=False)

    def test_administrador_puede_acceder_a_preliquidacion(self):
        with (
            patch("apps.bingos.views.get_object_or_404", return_value=self.bingo),
            patch(
                "apps.bingos.views.construir_preliquidacion_financiera_bingo",
                return_value=self.datos,
            ) as construir,
        ):
            response = preliquidacion_financiera(
                self._request("/bingos/130/preliquidacion/"),
                self.bingo.pk,
            )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Preliquidación financiera")
        self.assertContains(
            response,
            "Resultado provisional antes de costos materiales y gastos",
        )
        self.assertContains(response, "Pendientes de registrar")
        self.assertNotContains(response, "Utilidad neta")
        construir.assert_called_once_with(self.bingo)

    def test_jugador_no_puede_acceder_a_preliquidacion(self):
        with self.assertRaises(PermissionDenied):
            preliquidacion_financiera(
                self._request(
                    "/bingos/130/preliquidacion/",
                    user=self._jugador(),
                ),
                self.bingo.pk,
            )

    def test_visitante_no_puede_acceder_a_preliquidacion(self):
        response = preliquidacion_financiera(
            self._request(
                "/bingos/130/preliquidacion/",
                user=AnonymousUser(),
            ),
            self.bingo.pk,
        )

        self.assertEqual(response.status_code, 302)
        self.assertIn("/login/", response["Location"])

    def test_preliquidacion_acepta_solo_get(self):
        with (
            patch("apps.bingos.views.get_object_or_404", return_value=self.bingo),
            patch(
                "apps.bingos.views.construir_preliquidacion_financiera_bingo",
                return_value=self.datos,
            ),
        ):
            response = preliquidacion_financiera(
                self._request(
                    "/bingos/130/preliquidacion/",
                    method="post",
                ),
                self.bingo.pk,
            )

        self.assertEqual(response.status_code, 405)

    def test_ruta_de_preliquidacion_resuelve_la_vista(self):
        self.assertIs(
            resolve("/bingos/130/preliquidacion/").func,
            preliquidacion_financiera,
        )
        self.assertEqual(
            reverse(
                "bingos:preliquidacion_financiera",
                kwargs={"idbingo": self.bingo.pk},
            ),
            "/bingos/130/preliquidacion/",
        )

    def test_detalle_de_bingo_muestra_enlace_a_preliquidacion(self):
        with (
            patch("apps.bingos.views.get_object_or_404", return_value=self.bingo),
            patch(
                "apps.bingos.views.Partidabingo.objects.filter",
                return_value=_CartonesConsultaQuerySet([self.partida]),
            ),
            patch(
                "apps.bingos.views.Carton.objects.filter",
                return_value=_CartonesConsultaQuerySet([self.carton]),
            ),
        ):
            response = bingo_detalle(
                self._request("/bingos/130/"),
                self.bingo.pk,
            )

        self.assertContains(response, "Ver preliquidación financiera")
        self.assertContains(
            response,
            reverse(
                "bingos:preliquidacion_financiera",
                kwargs={"idbingo": self.bingo.pk},
            ),
        )

    def test_detalle_de_bingo_muestra_enlace_a_finanzas(self):
        with (
            patch("apps.bingos.views.get_object_or_404", return_value=self.bingo),
            patch(
                "apps.bingos.views.Partidabingo.objects.filter",
                return_value=_CartonesConsultaQuerySet([self.partida]),
            ),
            patch(
                "apps.bingos.views.Carton.objects.filter",
                return_value=_CartonesConsultaQuerySet([self.carton]),
            ),
        ):
            response = bingo_detalle(
                self._request("/bingos/130/"),
                self.bingo.pk,
            )

        self.assertContains(response, "Gestión financiera")
        self.assertContains(
            response,
            reverse(
                "bingos:bingo_finanzas",
                kwargs={"idbingo": self.bingo.pk},
            ),
        )


@override_settings(
    STORAGES={
        "default": {
            "BACKEND": "django.core.files.storage.FileSystemStorage",
        },
        "staticfiles": {
            "BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage",
        },
    }
)
class BingoFinanzasVistaTests(SimpleTestCase):
    def setUp(self):
        self.factory = RequestFactory()
        self.ahora = timezone.now()
        self.admin = User(username="admin_finanzas", is_staff=True)
        self.usuario_registro = User(username="registro_finanzas", is_staff=True)
        self.bingo = Bingo(
            idbingo=140,
            titulobingo="Bingo finanzas",
            fechaprogramadabingo=self.ahora,
            tipobingo="Virtual",
            preciocarton=Decimal("5.00"),
            premiomayor=Decimal("50.00"),
            descripcionpremiomayor="Premio",
            estadobingo="Programado",
        )
        self.partida = Partidabingo(
            idpartidabingo=1401,
            idbingo=self.bingo,
            nombreronda="Ronda material",
            valorefectivo=Decimal("10.00"),
            premiomaterial="Canasta",
            estadopartida=ESTADO_PARTIDA_FINALIZADA,
            horainicio=self.ahora,
            horafin=self.ahora,
        )
        self.gasto = BingoGastoOperativo(
            idbingogastooperativo=1,
            idbingo=self.bingo,
            concepto="Sonido",
            monto=Decimal("12.00"),
            fechagasto=self.ahora,
            estado=BingoGastoOperativo.ESTADO_REGISTRADO,
            observacion="Equipo local",
            idusuarioregistro=self.usuario_registro,
            fechacreacion=self.ahora,
        )
        self.gasto_anulado = BingoGastoOperativo(
            idbingogastooperativo=2,
            idbingo=self.bingo,
            concepto="Transporte duplicado",
            monto=Decimal("3.00"),
            fechagasto=self.ahora,
            estado=BingoGastoOperativo.ESTADO_ANULADO,
            observacion="",
            idusuarioregistro=self.usuario_registro,
            fechacreacion=self.ahora,
            motivoanulacion="Factura duplicada",
        )
        self.costo = BingoPremioMaterialCosto(
            idbingopremiomaterialcosto=1,
            idbingo=self.bingo,
            idpartidabingo=self.partida.pk,
            descripcionpremio="Canasta familiar",
            monto=Decimal("8.00"),
            estado=BingoPremioMaterialCosto.ESTADO_REGISTRADO,
            observacion="Compra local",
            idusuarioregistro=self.usuario_registro,
            fechacreacion=self.ahora,
        )
        self.costo_anulado = BingoPremioMaterialCosto(
            idbingopremiomaterialcosto=2,
            idbingo=self.bingo,
            idpartidabingo=self.partida.pk,
            descripcionpremio="Canasta duplicada",
            monto=Decimal("7.00"),
            estado=BingoPremioMaterialCosto.ESTADO_ANULADO,
            observacion="",
            idusuarioregistro=self.usuario_registro,
            fechacreacion=self.ahora,
            motivoanulacion="Costo duplicado",
        )

    def _request(self, user=None):
        request = self.factory.get(
            reverse("bingos:bingo_finanzas", kwargs={"idbingo": self.bingo.pk})
        )
        request.user = user if user is not None else self.admin
        request.session = {}
        return request

    def _resumen(self, cierre=None, bloqueos=None, pendientes=None):
        return {
            "cartones_vendidos_unicos": 3,
            "recaudacion_registrada": Decimal("15.00"),
            "premios_efectivo_finalizados": Decimal("10.00"),
            "costos_premios_materiales": Decimal("8.00"),
            "gastos_operativos": Decimal("12.00"),
            "resultado_provisional": Decimal("5.00"),
            "utilidad_bruta": Decimal("-3.00"),
            "utilidad_neta": Decimal("-15.00"),
            "total_rondas": 1,
            "rondas_finalizadas": 1,
            "rondas_canceladas": 0,
            "rondas_pendientes": 0,
            "premios_materiales_pendientes_de_costo": pendientes or [],
            "bloqueos_de_cierre": bloqueos or [],
            "cierre_existente": cierre,
            "cierre_esta_cerrado": (
                cierre is not None
                and cierre.estado == BingoCierreFinanciero.ESTADO_CERRADO
            ),
        }

    def _cierre_cerrado(self):
        return BingoCierreFinanciero(
            idbingocierrefinanciero=1,
            idbingo=self.bingo,
            estado=BingoCierreFinanciero.ESTADO_CERRADO,
            cartonesvendidosunicos=3,
            recaudacionregistrada=Decimal("15.00"),
            premiosefectivofinalizados=Decimal("10.00"),
            costospremiosmateriales=Decimal("8.00"),
            gastosoperativos=Decimal("12.00"),
            resultadoprovisional=Decimal("5.00"),
            utilidadbruta=Decimal("-3.00"),
            utilidadneta=Decimal("-15.00"),
            totalrondas=1,
            rondasfinalizadas=1,
            rondascanceladas=0,
            rondaspendientes=0,
            fechacalculo=self.ahora,
            fechacierre=self.ahora,
            idusuariocierre=self.admin,
            observacioncierre="Cierre final",
            fechacreacion=self.ahora,
        )

    def _render_finanzas(
        self,
        resumen=None,
        gastos=None,
        costos=None,
        partidas=None,
        user=None,
    ):
        resumen = resumen or self._resumen()
        gastos = _CartonesConsultaQuerySet(
            gastos if gastos is not None else [self.gasto, self.gasto_anulado]
        )
        costos = _CartonesConsultaQuerySet(
            costos if costos is not None else [self.costo, self.costo_anulado]
        )
        partidas = _CartonesConsultaQuerySet(
            partidas if partidas is not None else [self.partida]
        )
        with (
            patch("apps.bingos.views.get_object_or_404", return_value=self.bingo),
            patch(
                "apps.bingos.views.construir_resumen_financiero_real_bingo",
                return_value=resumen,
            ) as construir,
            patch(
                "apps.bingos.views.Partidabingo.objects.filter",
                return_value=partidas,
            ) as partidas_filter,
            patch(
                "apps.bingos.views.BingoGastoOperativo.objects.filter",
                return_value=gastos,
            ) as gastos_filter,
            patch(
                "apps.bingos.views.BingoPremioMaterialCosto.objects.filter",
                return_value=costos,
            ) as costos_filter,
        ):
            response = bingo_finanzas(self._request(user=user), self.bingo.pk)

        return response, construir, partidas_filter, gastos_filter, costos_filter

    def test_administrador_puede_abrir_panel_financiero(self):
        response, construir, _partidas, _gastos, _costos = self._render_finanzas()

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Gestión financiera")
        self.assertContains(response, "Resumen financiero")
        self.assertContains(response, "Cierre financiero abierto")
        construir.assert_called_once_with(self.bingo)

    def test_usuario_no_administrador_no_puede_acceder(self):
        with self.assertRaises(PermissionDenied):
            bingo_finanzas(
                self._request(
                    user=User(
                        username="jugador_finanzas",
                        is_staff=False,
                        is_superuser=False,
                    )
                ),
                self.bingo.pk,
            )

    def test_ruta_financiera_resuelve_vista(self):
        self.assertIs(resolve("/bingos/140/finanzas/").func, bingo_finanzas)
        self.assertEqual(
            reverse("bingos:bingo_finanzas", kwargs={"idbingo": self.bingo.pk}),
            "/bingos/140/finanzas/",
        )

    def test_vista_muestra_gastos_y_costos_del_bingo_correcto(self):
        response, _construir, partidas_filter, gastos_filter, costos_filter = (
            self._render_finanzas()
        )

        self.assertContains(response, "Sonido")
        self.assertContains(response, "Transporte duplicado")
        self.assertContains(response, "Factura duplicada")
        self.assertContains(response, "Canasta familiar")
        self.assertContains(response, "Canasta duplicada")
        self.assertContains(response, "Costo duplicado")
        self.assertContains(response, "Ronda material")
        self.assertEqual(partidas_filter.call_count, 2)
        partidas_filter.assert_any_call(idbingo=self.bingo)
        partidas_filter.assert_any_call(
            idbingo=self.bingo,
            premiomaterial__isnull=False,
            premiomaterial__gt="",
        )
        gastos_filter.assert_called_once_with(idbingo=self.bingo)
        costos_filter.assert_called_once_with(idbingo=self.bingo)

    def test_vista_muestra_bloqueos_y_premios_pendientes(self):
        resumen = self._resumen(
            bloqueos=[
                "No se puede cerrar el Bingo porque existen premios materiales finalizados sin costo registrado."
            ],
            pendientes=[
                {
                    "partida": self.partida,
                    "idpartidabingo": self.partida.pk,
                    "ronda": self.partida.nombreronda,
                    "descripcion": "Canasta",
                }
            ],
        )

        response, _construir, _partidas, _gastos, _costos = self._render_finanzas(
            resumen=resumen
        )

        self.assertContains(response, "Bloqueos de cierre")
        self.assertContains(response, "premios materiales finalizados")
        self.assertContains(response, "Premios materiales pendientes de costo")
        self.assertContains(response, "Canasta")

    def test_vista_muestra_cierre_cerrado_y_snapshot(self):
        cierre = self._cierre_cerrado()
        resumen = self._resumen(
            cierre=cierre,
            bloqueos=[
                "No se puede cerrar el Bingo porque ya existe un cierre Cerrado."
            ],
        )

        response, _construir, _partidas, _gastos, _costos = self._render_finanzas(
            resumen=resumen
        )

        self.assertContains(response, "Cierre financiero cerrado")
        self.assertContains(response, "Snapshot de cierre")
        self.assertContains(response, "Cierre final")
        self.assertContains(response, "irreversible")
        self.assertContains(response, "$15,00")
        self.assertContains(response, "$-15,00")

    def test_vista_no_crea_ni_modifica_datos(self):
        with (
            patch("apps.bingos.views.Carton.objects.create") as carton_create,
            patch(
                "apps.bingos.views.CartonPartidaBingo.objects.create"
            ) as participacion_create,
            patch(
                "apps.bingos.views.BingoGastoOperativo.objects.create"
            ) as gasto_create,
            patch(
                "apps.bingos.views.BingoPremioMaterialCosto.objects.create"
            ) as costo_create,
            patch(
                "apps.bingos.views.BingoCierreFinanciero.objects.create"
            ) as cierre_create,
            patch.object(BingoGastoOperativo, "save") as gasto_save,
            patch.object(BingoPremioMaterialCosto, "save") as costo_save,
            patch.object(BingoCierreFinanciero, "save") as cierre_save,
        ):
            response, _construir, _partidas, _gastos, _costos = (
                self._render_finanzas()
            )

        self.assertEqual(response.status_code, 200)
        carton_create.assert_not_called()
        participacion_create.assert_not_called()
        gasto_create.assert_not_called()
        costo_create.assert_not_called()
        cierre_create.assert_not_called()
        gasto_save.assert_not_called()
        costo_save.assert_not_called()
        cierre_save.assert_not_called()


@override_settings(
    STORAGES={
        "default": {
            "BACKEND": "django.core.files.storage.FileSystemStorage",
        },
        "staticfiles": {
            "BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage",
        },
    }
)
class BingoFinanzasGastosOperativosVistaIntegrationTests(TestCase):
    databases = {"default"}

    BOOTSTRAP_TABLES = (
        "tiposocio",
        "socio",
        "metodopago",
        "bingo",
        "jugador",
        "prestamo",
        "pago_prestamo",
        "solicitud_socio",
        "solicitud_pago_prestamo",
        "partidabingo",
        "carton",
        "carton_partida_bingo",
        "bingo_gasto_operativo",
        "bingo_premio_material_costo",
        "bingo_cierre_financiero",
    )
    PROTECTED_DATABASES = {"bingo", "bingo_ensayo_hibridos"}

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls._created_bootstrap_tables = False
        cls._bootstrap_finance_schema()

    @classmethod
    def tearDownClass(cls):
        try:
            if getattr(cls, "_created_bootstrap_tables", False):
                cls._drop_finance_schema()
        finally:
            super().tearDownClass()

    @classmethod
    def _bootstrap_finance_schema(cls):
        cls._validate_test_database()
        existing_tables = set(connection.introspection.table_names())
        required_tables = set(cls.BOOTSTRAP_TABLES)
        existing_required_tables = existing_tables & required_tables
        if existing_required_tables == required_tables:
            return
        if existing_required_tables:
            raise ImproperlyConfigured(
                "El esquema financiero de pruebas está incompleto: "
                + ", ".join(sorted(existing_required_tables))
            )

        sql_path = Path(settings.BASE_DIR) / "sql/test_schema_bingos_financiero.sql"
        statements = connection.ops.prepare_sql_script(
            sql_path.read_text(encoding="utf-8")
        )
        with connection.cursor() as cursor:
            for statement in statements:
                if statement.strip():
                    cursor.execute(statement)
        cls._created_bootstrap_tables = True

    @classmethod
    def _drop_finance_schema(cls):
        with connection.cursor() as cursor:
            for table_name in reversed(cls.BOOTSTRAP_TABLES):
                quoted_table = connection.ops.quote_name(table_name)
                cursor.execute(f"DROP TABLE IF EXISTS {quoted_table} CASCADE")

    @classmethod
    def _validate_test_database(cls):
        with connection.cursor() as cursor:
            cursor.execute("SELECT current_database()")
            database_name = str(cursor.fetchone()[0]).strip()
        if database_name in cls.PROTECTED_DATABASES:
            raise ImproperlyConfigured(
                f"No se pueden preparar pruebas financieras sobre {database_name!r}."
            )
        if not database_name.startswith("test_"):
            raise ImproperlyConfigured(
                "Las pruebas financieras de vista requieren una base de prueba "
                f"con prefijo 'test_'; base actual: {database_name!r}."
            )

    def setUp(self):
        self.ahora = timezone.now()
        self.admin = User.objects.create_user(
            username="admin_gastos_finanzas",
            password="test",
            is_staff=True,
        )
        self.no_admin = User.objects.create_user(
            username="usuario_gastos_finanzas",
            password="test",
            is_staff=False,
        )
        self.bingo = self._crear_bingo(700, "Bingo gastos abierto")
        self.url_panel = reverse(
            "bingos:bingo_finanzas",
            kwargs={"idbingo": self.bingo.pk},
        )
        self.url_registrar = reverse(
            "bingos:bingo_finanzas_gasto_registrar",
            kwargs={"idbingo": self.bingo.pk},
        )
        self.url_costo_registrar = reverse(
            "bingos:bingo_finanzas_costo_registrar",
            kwargs={"idbingo": self.bingo.pk},
        )
        self.url_cerrar = reverse(
            "bingos:bingo_finanzas_cerrar",
            kwargs={"idbingo": self.bingo.pk},
        )

    def _crear_bingo(self, pk, titulo):
        return Bingo.objects.create(
            idbingo=pk,
            titulobingo=titulo,
            fechaprogramadabingo=self.ahora,
            tipobingo="Virtual",
            preciocarton=Decimal("5.00"),
            premiomayor=Decimal("50.00"),
            descripcionpremiomayor="Premio mayor",
            estadobingo="Programado",
        )

    def _crear_gasto(self, concepto="Sonido", monto=Decimal("12.00"), bingo=None):
        return registrar_gasto_operativo_bingo(
            bingo or self.bingo,
            self.admin,
            concepto,
            monto,
        )

    def _crear_partida(
        self,
        pk=1700,
        bingo=None,
        premio_material="Canasta familiar",
        estado=ESTADO_PARTIDA_FINALIZADA,
    ):
        return Partidabingo.objects.create(
            idpartidabingo=pk,
            idbingo=bingo or self.bingo,
            nombreronda=f"Ronda {pk}",
            valorefectivo=Decimal("0.00"),
            premiomaterial=premio_material,
            estadopartida=estado,
            patronganador="carton_lleno",
            bolascantadas="[]",
            ultimabola=0,
            haydesempate=False,
            horainicio=self.ahora,
            horafin=self.ahora if estado == ESTADO_PARTIDA_FINALIZADA else None,
        )

    def _crear_costo(
        self,
        partida=None,
        monto=Decimal("12.00"),
        observacion="Compra local",
        bingo=None,
    ):
        bingo = bingo or self.bingo
        partida = partida or self._crear_partida()
        return registrar_costo_premio_material_bingo(
            bingo,
            partida,
            self.admin,
            partida.premiomaterial,
            monto,
            observacion,
        )

    def _crear_cierre_cerrado(self, bingo=None):
        return BingoCierreFinanciero.objects.create(
            idbingo=bingo or self.bingo,
            estado=BingoCierreFinanciero.ESTADO_CERRADO,
            cartonesvendidosunicos=0,
            recaudacionregistrada=Decimal("0.00"),
            premiosefectivofinalizados=Decimal("0.00"),
            costospremiosmateriales=Decimal("0.00"),
            gastosoperativos=Decimal("0.00"),
            resultadoprovisional=Decimal("0.00"),
            utilidadbruta=Decimal("0.00"),
            utilidadneta=Decimal("0.00"),
            totalrondas=0,
            rondasfinalizadas=0,
            rondascanceladas=0,
            rondaspendientes=0,
            fechacalculo=self.ahora,
            fechacierre=self.ahora,
            idusuariocierre=self.admin,
            observacioncierre="Cierre de prueba",
            fechacreacion=self.ahora,
        )

    def _url_anular(self, gasto, bingo=None):
        return reverse(
            "bingos:bingo_finanzas_gasto_anular",
            kwargs={
                "idbingo": (bingo or self.bingo).pk,
                "idgasto": gasto.pk,
            },
        )

    def _url_anular_costo(self, costo, bingo=None):
        return reverse(
            "bingos:bingo_finanzas_costo_anular",
            kwargs={
                "idbingo": (bingo or self.bingo).pk,
                "idcosto": costo.pk,
            },
        )

    def test_administrador_puede_ver_formulario_de_gasto_en_bingo_abierto(self):
        self.client.force_login(self.admin)

        response = self.client.get(self.url_panel)

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Registrar gasto operativo")
        self.assertContains(response, "Registrar gasto")
        self.assertContains(response, "Cierre financiero abierto")
        self.assertIsInstance(response.context["gasto_form"], GastoOperativoBingoForm)
        self.assertFalse(response.context["cierre_esta_cerrado"])

    def test_administrador_puede_ver_formulario_de_costo_en_bingo_abierto(self):
        self._crear_partida(pk=1701)
        self.client.force_login(self.admin)

        response = self.client.get(self.url_panel)

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Registrar costo de premio material")
        self.assertContains(response, "Registrar costo material")
        self.assertContains(
            response,
            "Un costo de $0,00 requiere una observación que indique que el premio fue donado.",
        )
        self.assertIsInstance(
            response.context["costo_form"],
            CostoPremioMaterialBingoForm,
        )
        self.assertFalse(response.context["cierre_esta_cerrado"])

    def test_administrador_ve_seccion_de_cierre_en_bingo_abierto(self):
        self.client.force_login(self.admin)

        response = self.client.get(self.url_panel)

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Cierre financiero")
        self.assertContains(response, "Cerrar financieramente Bingo")
        self.assertContains(response, "Este cierre es irreversible en esta versión.")
        self.assertIsInstance(
            response.context["cierre_form"],
            CierreFinancieroBingoForm,
        )
        self.assertFalse(response.context["cierre_esta_cerrado"])

    def test_selector_de_partidas_de_costo_limita_bingo_y_premio_material(self):
        partida_valida = self._crear_partida(
            pk=1701,
            premio_material="Canasta familiar",
        )
        partida_sin_material = self._crear_partida(pk=1702, premio_material="")
        otro_bingo = self._crear_bingo(701, "Bingo selector ajeno")
        partida_otro_bingo = self._crear_partida(
            pk=1703,
            bingo=otro_bingo,
            premio_material="Bono ajeno",
        )
        self.client.force_login(self.admin)

        response = self.client.get(self.url_panel)
        queryset = response.context["costo_form"].fields["partida"].queryset

        self.assertIn(partida_valida, list(queryset))
        self.assertNotIn(partida_sin_material, list(queryset))
        self.assertNotIn(partida_otro_bingo, list(queryset))
        self.assertContains(response, "Ronda 1701 - Canasta familiar")

    def test_usuario_no_administrativo_no_puede_acceder_a_rutas_nuevas(self):
        gasto = self._crear_gasto()
        self.client.force_login(self.no_admin)

        response_registrar = self.client.post(
            self.url_registrar,
            {
                "concepto": "Transporte",
                "monto": "4.00",
                "fechagasto": "",
                "observacion": "",
            },
        )
        response_anular = self.client.post(
            self._url_anular(gasto),
            {"motivo": "No autorizado"},
        )

        self.assertEqual(response_registrar.status_code, 403)
        self.assertEqual(response_anular.status_code, 403)
        self.assertEqual(BingoGastoOperativo.objects.count(), 1)
        gasto.refresh_from_db()
        self.assertEqual(gasto.estado, BingoGastoOperativo.ESTADO_REGISTRADO)

    def test_usuario_no_administrativo_no_puede_acceder_a_rutas_de_costos(self):
        partida = self._crear_partida(pk=1701)
        costo = self._crear_costo(partida=partida)
        self.client.force_login(self.no_admin)

        response_registrar = self.client.post(
            self.url_costo_registrar,
            {
                "partida": str(partida.pk),
                "descripcion_premio": "Canasta familiar",
                "monto": "9.00",
                "observacion": "Compra local",
            },
        )
        response_anular = self.client.post(
            self._url_anular_costo(costo),
            {"motivo": "No autorizado"},
        )

        self.assertEqual(response_registrar.status_code, 403)
        self.assertEqual(response_anular.status_code, 403)
        self.assertEqual(BingoPremioMaterialCosto.objects.count(), 1)
        costo.refresh_from_db()
        self.assertEqual(costo.estado, BingoPremioMaterialCosto.ESTADO_REGISTRADO)

    def test_usuario_no_administrativo_no_puede_acceder_a_ruta_de_cierre(self):
        self.client.force_login(self.no_admin)

        response = self.client.post(
            self.url_cerrar,
            {"observacion_cierre": "No autorizado"},
        )

        self.assertEqual(response.status_code, 403)
        self.assertEqual(BingoCierreFinanciero.objects.count(), 0)

    def test_si_hay_bloqueos_no_muestra_boton_activo_de_cierre(self):
        self._crear_partida(
            pk=1704,
            premio_material="",
            estado=ESTADO_PARTIDA_EN_ESPERA,
        )
        self.client.force_login(self.admin)

        response = self.client.get(self.url_panel)

        self.assertContains(
            response,
            "No se puede cerrar hasta resolver los bloqueos financieros.",
        )
        self.assertContains(response, "existen rondas pendientes")
        self.assertNotContains(response, "Cerrar financieramente Bingo</button>")

    def test_post_valido_de_registro_crea_gasto_y_muestra_exito(self):
        self.client.force_login(self.admin)

        response = self.client.post(
            self.url_registrar,
            {
                "concepto": "  Sonido local  ",
                "monto": "12.50",
                "fechagasto": "",
                "observacion": "  Consola principal  ",
            },
            follow=True,
        )

        self.assertRedirects(response, self.url_panel)
        self.assertEqual(BingoGastoOperativo.objects.count(), 1)
        gasto = BingoGastoOperativo.objects.get()
        self.assertEqual(gasto.estado, BingoGastoOperativo.ESTADO_REGISTRADO)
        self.assertEqual(gasto.concepto, "Sonido local")
        self.assertEqual(gasto.monto, Decimal("12.50"))
        self.assertEqual(gasto.observacion, "Consola principal")
        self.assertContains(response, "Gasto operativo registrado correctamente.")
        self.assertContains(response, "Sonido local")

    def test_post_valido_de_costo_crea_registrado_y_muestra_exito(self):
        partida = self._crear_partida(pk=1701)
        self.client.force_login(self.admin)

        response = self.client.post(
            self.url_costo_registrar,
            {
                "partida": str(partida.pk),
                "descripcion_premio": "  Canasta familiar  ",
                "monto": "18.50",
                "observacion": "  Compra local  ",
            },
            follow=True,
        )

        self.assertRedirects(response, self.url_panel)
        self.assertEqual(BingoPremioMaterialCosto.objects.count(), 1)
        costo = BingoPremioMaterialCosto.objects.get()
        self.assertEqual(costo.estado, BingoPremioMaterialCosto.ESTADO_REGISTRADO)
        self.assertEqual(costo.idbingo, self.bingo)
        self.assertEqual(costo.idpartidabingo, partida.pk)
        self.assertEqual(costo.idusuarioregistro, self.admin)
        self.assertEqual(costo.descripcionpremio, "Canasta familiar")
        self.assertEqual(costo.monto, Decimal("18.50"))
        self.assertEqual(costo.observacion, "Compra local")
        self.assertContains(
            response,
            "Costo de premio material registrado correctamente.",
        )
        self.assertContains(response, "Canasta familiar")

    def test_post_costo_cero_con_observacion_se_registra_correctamente(self):
        partida = self._crear_partida(pk=1701, premio_material="Bono donado")
        self.client.force_login(self.admin)

        response = self.client.post(
            self.url_costo_registrar,
            {
                "partida": str(partida.pk),
                "descripcion_premio": "Bono donado",
                "monto": "0.00",
                "observacion": "Premio donado por auspiciante",
            },
            follow=True,
        )

        self.assertRedirects(response, self.url_panel)
        costo = BingoPremioMaterialCosto.objects.get()
        self.assertEqual(costo.monto, Decimal("0.00"))
        self.assertEqual(costo.estado, BingoPremioMaterialCosto.ESTADO_REGISTRADO)
        self.assertContains(
            response,
            "Costo de premio material registrado correctamente.",
        )

    def test_post_costo_cero_sin_observacion_muestra_error_de_dominio(self):
        partida = self._crear_partida(pk=1701, premio_material="Bono donado")
        self.client.force_login(self.admin)

        response = self.client.post(
            self.url_costo_registrar,
            {
                "partida": str(partida.pk),
                "descripcion_premio": "Bono donado",
                "monto": "0.00",
                "observacion": "",
            },
            follow=True,
        )

        self.assertRedirects(response, self.url_panel)
        self.assertEqual(BingoPremioMaterialCosto.objects.count(), 0)
        self.assertContains(
            response,
            "Un costo de cero debe incluir una observación que indique la donación.",
        )

    def test_post_costo_con_partida_de_otro_bingo_no_crea(self):
        otro_bingo = self._crear_bingo(701, "Bingo costo ajeno")
        partida_ajena = self._crear_partida(
            pk=1701,
            bingo=otro_bingo,
            premio_material="Bono ajeno",
        )
        self.client.force_login(self.admin)

        response = self.client.post(
            self.url_costo_registrar,
            {
                "partida": str(partida_ajena.pk),
                "descripcion_premio": "Bono ajeno",
                "monto": "8.00",
                "observacion": "Compra ajena",
            },
            follow=True,
        )

        self.assertRedirects(response, self.url_panel)
        self.assertEqual(BingoPremioMaterialCosto.objects.count(), 0)
        self.assertContains(
            response,
            "No se pudo registrar el costo de premio material. Revise los datos ingresados.",
        )

    def test_segundo_costo_registrado_para_misma_partida_se_bloquea(self):
        partida = self._crear_partida(pk=1701)
        self._crear_costo(partida=partida)
        self.client.force_login(self.admin)

        response = self.client.post(
            self.url_costo_registrar,
            {
                "partida": str(partida.pk),
                "descripcion_premio": "Canasta familiar",
                "monto": "20.00",
                "observacion": "Segundo registro",
            },
            follow=True,
        )

        self.assertRedirects(response, self.url_panel)
        self.assertEqual(BingoPremioMaterialCosto.objects.count(), 1)
        self.assertContains(
            response,
            "Ya existe un costo de premio material registrado para esta partida.",
        )

    def test_cierre_exitoso_crea_cierre_y_deja_panel_solo_lectura(self):
        self._crear_partida(pk=1705, premio_material="")
        self._crear_gasto(monto=Decimal("12.00"))
        self.client.force_login(self.admin)

        response = self.client.post(
            self.url_cerrar,
            {"observacion_cierre": "  cierre validado  "},
            follow=True,
        )

        self.assertRedirects(response, self.url_panel)
        self.assertEqual(BingoCierreFinanciero.objects.count(), 1)
        cierre = BingoCierreFinanciero.objects.get()
        self.assertEqual(cierre.estado, BingoCierreFinanciero.ESTADO_CERRADO)
        self.assertEqual(cierre.idbingo, self.bingo)
        self.assertEqual(cierre.idusuariocierre, self.admin)
        self.assertEqual(cierre.observacioncierre, "cierre validado")
        self.assertEqual(cierre.gastosoperativos, Decimal("12.00"))
        self.assertContains(
            response,
            "Bingo cerrado financieramente correctamente.",
        )
        self.assertContains(response, "Cierre financiero cerrado")
        self.assertContains(
            response,
            "El cierre financiero ya fue realizado y es solo de consulta.",
        )
        self.assertNotContains(response, "Registrar gasto operativo")
        self.assertNotContains(response, "Registrar costo de premio material")
        self.assertNotContains(response, "Cerrar financieramente Bingo</button>")

    def test_cierre_bloqueado_por_ronda_pendiente_muestra_error_de_dominio(self):
        self._crear_partida(
            pk=1706,
            premio_material="",
            estado=ESTADO_PARTIDA_EN_ESPERA,
        )
        self.client.force_login(self.admin)

        response = self.client.post(
            self.url_cerrar,
            {"observacion_cierre": "Intento con bloqueo"},
            follow=True,
        )

        self.assertRedirects(response, self.url_panel)
        self.assertFalse(
            BingoCierreFinanciero.objects.filter(
                estado=BingoCierreFinanciero.ESTADO_CERRADO,
            ).exists()
        )
        self.assertContains(
            response,
            "No se puede cerrar financieramente el Bingo:",
        )
        self.assertContains(response, "rondas pendientes")

    def test_segundo_cierre_no_modifica_snapshot_y_muestra_error(self):
        self._crear_partida(pk=1707, premio_material="")
        self._crear_gasto(monto=Decimal("9.00"))
        self.client.force_login(self.admin)
        self.client.post(
            self.url_cerrar,
            {"observacion_cierre": "Cierre original"},
            follow=True,
        )
        cierre = BingoCierreFinanciero.objects.get()
        campos_snapshot = (
            "recaudacionregistrada",
            "gastosoperativos",
            "utilidadneta",
            "totalrondas",
            "fechacalculo",
            "fechacierre",
            "observacioncierre",
        )
        snapshot_original = {
            campo: getattr(cierre, campo) for campo in campos_snapshot
        }

        response = self.client.post(
            self.url_cerrar,
            {"observacion_cierre": "Intento posterior"},
            follow=True,
        )

        self.assertRedirects(response, self.url_panel)
        self.assertEqual(BingoCierreFinanciero.objects.count(), 1)
        cierre.refresh_from_db()
        for campo, valor in snapshot_original.items():
            self.assertEqual(getattr(cierre, campo), valor)
        self.assertContains(response, "El cierre financiero ya está cerrado.")

    def test_cierre_ignora_totales_manipulados_y_usa_snapshot_real(self):
        self._crear_partida(pk=1708, premio_material="")
        self._crear_gasto(monto=Decimal("7.50"))
        self.client.force_login(self.admin)

        response = self.client.post(
            self.url_cerrar,
            {
                "observacion_cierre": "Cierre con datos manipulados",
                "recaudacionregistrada": "999.99",
                "costospremiosmateriales": "888.88",
                "utilidadneta": "777.77",
            },
            follow=True,
        )

        self.assertRedirects(response, self.url_panel)
        cierre = BingoCierreFinanciero.objects.get()
        self.assertEqual(cierre.recaudacionregistrada, Decimal("0.00"))
        self.assertEqual(cierre.costospremiosmateriales, Decimal("0.00"))
        self.assertEqual(cierre.gastosoperativos, Decimal("7.50"))
        self.assertEqual(cierre.utilidadneta, Decimal("-7.50"))
        self.assertNotEqual(cierre.utilidadneta, Decimal("777.77"))

    def test_post_invalido_no_crea_gasto_y_muestra_error(self):
        self.client.force_login(self.admin)

        response = self.client.post(
            self.url_registrar,
            {
                "concepto": "Transporte",
                "monto": "valor-invalido",
                "fechagasto": "",
                "observacion": "",
            },
            follow=True,
        )

        self.assertRedirects(response, self.url_panel)
        self.assertEqual(BingoGastoOperativo.objects.count(), 0)
        self.assertContains(
            response,
            "No se pudo registrar el gasto operativo. Revise los datos ingresados.",
        )

    def test_anulacion_valida_conserva_gasto_y_registra_motivo(self):
        gasto = self._crear_gasto()
        self.client.force_login(self.admin)

        response = self.client.post(
            self._url_anular(gasto),
            {"motivo": "  Factura duplicada  "},
            follow=True,
        )

        self.assertRedirects(response, self.url_panel)
        self.assertEqual(BingoGastoOperativo.objects.count(), 1)
        gasto.refresh_from_db()
        self.assertEqual(gasto.estado, BingoGastoOperativo.ESTADO_ANULADO)
        self.assertEqual(gasto.motivoanulacion, "Factura duplicada")
        self.assertContains(response, "Gasto operativo anulado correctamente.")

    def test_anulacion_valida_de_costo_conserva_registro_y_motivo(self):
        partida = self._crear_partida(pk=1701)
        costo = self._crear_costo(partida=partida)
        self.client.force_login(self.admin)

        response = self.client.post(
            self._url_anular_costo(costo),
            {"motivo": "  Factura duplicada  "},
            follow=True,
        )

        self.assertRedirects(response, self.url_panel)
        self.assertEqual(BingoPremioMaterialCosto.objects.count(), 1)
        costo.refresh_from_db()
        self.assertEqual(costo.estado, BingoPremioMaterialCosto.ESTADO_ANULADO)
        self.assertEqual(costo.motivoanulacion, "Factura duplicada")
        self.assertContains(
            response,
            "Costo de premio material anulado correctamente.",
        )

    def test_anulacion_de_gasto_de_otro_bingo_devuelve_404_y_no_modifica(self):
        otro_bingo = self._crear_bingo(701, "Bingo gastos ajeno")
        gasto_ajeno = self._crear_gasto(
            concepto="Gasto ajeno",
            monto=Decimal("8.00"),
            bingo=otro_bingo,
        )
        self.client.force_login(self.admin)

        response = self.client.post(
            self._url_anular(gasto_ajeno, bingo=self.bingo),
            {"motivo": "No corresponde"},
        )

        self.assertEqual(response.status_code, 404)
        gasto_ajeno.refresh_from_db()
        self.assertEqual(gasto_ajeno.estado, BingoGastoOperativo.ESTADO_REGISTRADO)
        self.assertEqual(gasto_ajeno.motivoanulacion, None)

    def test_anulacion_de_costo_de_otro_bingo_devuelve_404_y_no_modifica(self):
        otro_bingo = self._crear_bingo(701, "Bingo costo ajeno")
        partida_ajena = self._crear_partida(
            pk=1701,
            bingo=otro_bingo,
            premio_material="Bono ajeno",
        )
        costo_ajeno = self._crear_costo(
            partida=partida_ajena,
            bingo=otro_bingo,
        )
        self.client.force_login(self.admin)

        response = self.client.post(
            self._url_anular_costo(costo_ajeno, bingo=self.bingo),
            {"motivo": "No corresponde"},
        )

        self.assertEqual(response.status_code, 404)
        costo_ajeno.refresh_from_db()
        self.assertEqual(costo_ajeno.estado, BingoPremioMaterialCosto.ESTADO_REGISTRADO)
        self.assertEqual(costo_ajeno.motivoanulacion, None)

    def test_despues_de_anular_permite_registrar_reemplazo_misma_partida(self):
        partida = self._crear_partida(pk=1701)
        costo = self._crear_costo(partida=partida)
        self.client.force_login(self.admin)
        self.client.post(
            self._url_anular_costo(costo),
            {"motivo": "Factura duplicada"},
            follow=True,
        )

        response = self.client.post(
            self.url_costo_registrar,
            {
                "partida": str(partida.pk),
                "descripcion_premio": "Canasta familiar",
                "monto": "19.00",
                "observacion": "Reemplazo",
            },
            follow=True,
        )

        self.assertRedirects(response, self.url_panel)
        self.assertEqual(BingoPremioMaterialCosto.objects.count(), 2)
        self.assertEqual(
            BingoPremioMaterialCosto.objects.filter(
                estado=BingoPremioMaterialCosto.ESTADO_REGISTRADO,
            ).count(),
            1,
        )
        self.assertContains(
            response,
            "Costo de premio material registrado correctamente.",
        )

    def test_bingo_cerrado_oculta_formularios_y_bloquea_posts(self):
        gasto = self._crear_gasto()
        self._crear_cierre_cerrado()
        self.client.force_login(self.admin)

        response_get = self.client.get(self.url_panel)
        self.assertNotContains(response_get, "Registrar gasto operativo")
        self.assertNotContains(response_get, "Anular</button>")
        self.assertContains(
            response_get,
            "El cierre financiero está cerrado. Los gastos se muestran solo para consulta.",
        )

        response_registrar = self.client.post(
            self.url_registrar,
            {
                "concepto": "Transporte",
                "monto": "4.00",
                "fechagasto": "",
                "observacion": "",
            },
            follow=True,
        )
        self.assertEqual(BingoGastoOperativo.objects.count(), 1)
        self.assertContains(
            response_registrar,
            "No se pueden registrar gastos porque el cierre financiero está cerrado.",
        )

        response_anular = self.client.post(
            self._url_anular(gasto),
            {"motivo": "Factura duplicada"},
            follow=True,
        )
        gasto.refresh_from_db()
        self.assertEqual(gasto.estado, BingoGastoOperativo.ESTADO_REGISTRADO)
        self.assertEqual(gasto.motivoanulacion, None)
        self.assertContains(
            response_anular,
            "No se pueden anular gastos porque el cierre financiero está cerrado.",
        )

    def test_bingo_cerrado_oculta_formulario_costos_y_bloquea_posts(self):
        partida = self._crear_partida(pk=1701)
        costo = self._crear_costo(partida=partida)
        self._crear_cierre_cerrado()
        self.client.force_login(self.admin)

        response_get = self.client.get(self.url_panel)
        self.assertNotContains(response_get, "Registrar costo de premio material")
        self.assertNotContains(response_get, "Registrar costo material")
        self.assertNotContains(response_get, "Anular</button>")
        self.assertContains(
            response_get,
            "El cierre financiero está cerrado. Los costos de premios se muestran solo para consulta.",
        )

        response_registrar = self.client.post(
            self.url_costo_registrar,
            {
                "partida": str(partida.pk),
                "descripcion_premio": "Canasta familiar",
                "monto": "22.00",
                "observacion": "Compra posterior",
            },
            follow=True,
        )
        self.assertEqual(BingoPremioMaterialCosto.objects.count(), 1)
        self.assertContains(
            response_registrar,
            "No se pueden registrar costos porque el cierre financiero está cerrado.",
        )

        response_anular = self.client.post(
            self._url_anular_costo(costo),
            {"motivo": "Factura duplicada"},
            follow=True,
        )
        costo.refresh_from_db()
        self.assertEqual(costo.estado, BingoPremioMaterialCosto.ESTADO_REGISTRADO)
        self.assertEqual(costo.motivoanulacion, None)
        self.assertContains(
            response_anular,
            "No se pueden anular costos porque el cierre financiero está cerrado.",
        )

    def test_despues_del_cierre_oculta_formularios_y_muestra_solo_consulta(self):
        gasto = self._crear_gasto()
        partida = self._crear_partida(pk=1709)
        costo = self._crear_costo(partida=partida)
        self.client.force_login(self.admin)
        self.client.post(
            self.url_cerrar,
            {"observacion_cierre": "Cierre definitivo"},
            follow=True,
        )

        response = self.client.get(self.url_panel)

        self.assertEqual(response.status_code, 200)
        gasto.refresh_from_db()
        costo.refresh_from_db()
        self.assertEqual(gasto.estado, BingoGastoOperativo.ESTADO_REGISTRADO)
        self.assertEqual(costo.estado, BingoPremioMaterialCosto.ESTADO_REGISTRADO)
        self.assertNotContains(response, "Registrar gasto operativo")
        self.assertNotContains(response, "Registrar costo de premio material")
        self.assertNotContains(response, "Cerrar financieramente Bingo</button>")
        self.assertNotContains(response, "Anular</button>")
        self.assertContains(
            response,
            "El cierre financiero está cerrado. Los gastos se muestran solo para consulta.",
        )
        self.assertContains(
            response,
            "El cierre financiero está cerrado. Los costos de premios se muestran solo para consulta.",
        )
        self.assertContains(
            response,
            "El cierre financiero ya fue realizado y es solo de consulta.",
        )

    def test_operacion_impacta_resumen_real(self):
        self.client.force_login(self.admin)

        response_registro = self.client.post(
            self.url_registrar,
            {
                "concepto": "Transporte",
                "monto": "20.00",
                "fechagasto": "",
                "observacion": "",
            },
            follow=True,
        )
        gasto = BingoGastoOperativo.objects.get()
        resumen_registrado = construir_resumen_financiero_real_bingo(self.bingo)

        self.assertContains(response_registro, "Transporte")
        self.assertEqual(
            resumen_registrado["gastos_operativos"],
            Decimal("20.00"),
        )

        self.client.post(
            self._url_anular(gasto),
            {"motivo": "Registro duplicado"},
            follow=True,
        )
        resumen_anulado = construir_resumen_financiero_real_bingo(self.bingo)

        gasto.refresh_from_db()
        self.assertEqual(gasto.estado, BingoGastoOperativo.ESTADO_ANULADO)
        self.assertEqual(resumen_anulado["gastos_operativos"], Decimal("0.00"))

    def test_operacion_de_costo_impacta_resumen_real(self):
        partida = self._crear_partida(pk=1701)
        self.client.force_login(self.admin)

        response_registro = self.client.post(
            self.url_costo_registrar,
            {
                "partida": str(partida.pk),
                "descripcion_premio": "Canasta familiar",
                "monto": "15.00",
                "observacion": "Compra local",
            },
            follow=True,
        )
        costo = BingoPremioMaterialCosto.objects.get()
        resumen_registrado = construir_resumen_financiero_real_bingo(self.bingo)

        self.assertContains(response_registro, "Canasta familiar")
        self.assertEqual(
            resumen_registrado["costos_premios_materiales"],
            Decimal("15.00"),
        )

        self.client.post(
            self._url_anular_costo(costo),
            {"motivo": "Registro duplicado"},
            follow=True,
        )
        resumen_anulado = construir_resumen_financiero_real_bingo(self.bingo)

        costo.refresh_from_db()
        self.assertEqual(costo.estado, BingoPremioMaterialCosto.ESTADO_ANULADO)
        self.assertEqual(
            resumen_anulado["costos_premios_materiales"],
            Decimal("0.00"),
        )

    def test_rutas_post_de_gastos_resuelven_vistas(self):
        self.assertIs(
            resolve("/bingos/700/finanzas/gastos/registrar/").func,
            bingo_finanzas_gasto_registrar,
        )
        self.assertIs(
            resolve("/bingos/700/finanzas/gastos/5/anular/").func,
            bingo_finanzas_gasto_anular,
        )

    def test_rutas_post_de_costos_resuelven_vistas(self):
        self.assertIs(
            resolve("/bingos/700/finanzas/costos/registrar/").func,
            bingo_finanzas_costo_registrar,
        )
        self.assertIs(
            resolve("/bingos/700/finanzas/costos/5/anular/").func,
            bingo_finanzas_costo_anular,
        )

    def test_ruta_post_de_cierre_resuelve_vista(self):
        self.assertIs(
            resolve("/bingos/700/finanzas/cerrar/").func,
            bingo_finanzas_cerrar,
        )
