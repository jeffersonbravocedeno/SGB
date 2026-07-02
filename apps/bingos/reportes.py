import re
from collections import Counter, defaultdict
from decimal import Decimal
from io import BytesIO

from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from .services import (
    ESTADO_PARTIDA_FINALIZADA,
    estado_partida_mostrar,
    formatear_bola_bingo,
    parsear_bolas_cantadas,
    preparar_datos_bolas_partida,
)


PDF_CONTENT_TYPE = "application/pdf"
XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

MONEDA_FORMAT = '"$"#,##0.00'
FECHA_FORMAT = "dd/mm/yyyy hh:mm"

TIPO_HISTORICO = "historico"
TIPO_HIBRIDO = "hibrido"
TIPO_HISTORICO_ETIQUETA = "Histórico por partida"
TIPO_HIBRIDO_ETIQUETA = "Cartón de Bingo"


class ReporteHibridoError(Exception):
    """Indica una relación inconsistente al construir un reporte híbrido."""


def nombre_archivo_seguro(prefijo, identificador, extension):
    base = f"{prefijo}_{identificador}"
    base = re.sub(r"[^A-Za-z0-9_-]+", "_", str(base)).strip("_")
    return f"{base}.{extension}"


def construir_filas_reporte_partida(
    partida,
    cartones_historicos=None,
    participaciones_hibridas=None,
):
    """Normaliza históricos y participaciones de una ronda sin escribir datos."""
    idpartida = partida.pk
    filas_historicas = []
    for carton in list(cartones_historicos or []):
        if getattr(carton, "idpartida_id", None) != idpartida:
            continue
        filas_historicas.append(
            _fila_reporte_partida(
                partida,
                carton,
                tipo=TIPO_HISTORICO,
            )
        )

    filas_hibridas = []
    participaciones_vistas = set()
    for participacion in list(participaciones_hibridas or []):
        if getattr(participacion, "idpartida_id", None) != idpartida:
            continue
        carton = participacion.idcarton
        _validar_participacion_hibrida(partida, carton, participacion)
        clave = (carton.pk, participacion.idpartida_id)
        if clave in participaciones_vistas:
            raise ReporteHibridoError(
                "Existe más de una participación del mismo cartón en la ronda."
            )
        participaciones_vistas.add(clave)
        filas_hibridas.append(
            _fila_reporte_partida(
                partida,
                carton,
                tipo=TIPO_HIBRIDO,
                participacion=participacion,
            )
        )

    clave_orden = lambda fila: (str(fila["codigo_carton"]), fila["idcarton"])
    filas_historicas.sort(key=clave_orden)
    filas_hibridas.sort(key=clave_orden)
    return filas_historicas + filas_hibridas


def construir_resumenes_cartones_bingo(
    bingo,
    cartones,
    participaciones_hibridas=None,
):
    """Devuelve una fila de inventario por maestro, nunca por participación."""
    idbingo = bingo.pk
    maestros = []
    maestros_por_id = {}
    for carton in list(cartones or []):
        if carton.pk in maestros_por_id:
            continue
        idbingo_carton = _idbingo_carton(carton)
        if idbingo_carton not in (None, idbingo):
            raise ReporteHibridoError(
                "Un cartón del resumen no pertenece al Bingo solicitado."
            )
        maestros.append(carton)
        maestros_por_id[carton.pk] = carton

    participaciones_por_carton = defaultdict(list)
    claves_vistas = set()
    for participacion in list(participaciones_hibridas or []):
        carton = participacion.idcarton
        if carton.pk not in maestros_por_id:
            raise ReporteHibridoError(
                "Una participación no corresponde a los maestros del Bingo."
            )
        _validar_participacion_hibrida(
            participacion.idpartida,
            carton,
            participacion,
            bingo=bingo,
        )
        clave = (carton.pk, participacion.idpartida_id)
        if clave in claves_vistas:
            raise ReporteHibridoError(
                "Existe más de una participación del mismo cartón en una ronda."
            )
        claves_vistas.add(clave)
        participaciones_por_carton[carton.pk].append(participacion)

    resumenes = []
    for carton in maestros:
        es_hibrido = getattr(carton, "idpartida_id", None) is None
        participaciones = participaciones_por_carton.get(carton.pk, [])
        if not es_hibrido and participaciones:
            raise ReporteHibridoError(
                "Un cartón histórico no puede reportarse como maestro híbrido."
            )
        estados = Counter(
            participacion.estado_participacion
            for participacion in participaciones
        )
        resumenes.append(
            {
                "tipo": TIPO_HIBRIDO if es_hibrido else TIPO_HISTORICO,
                "tipo_etiqueta": (
                    TIPO_HIBRIDO_ETIQUETA
                    if es_hibrido
                    else TIPO_HISTORICO_ETIQUETA
                ),
                "carton": carton,
                "idcarton": carton.pk,
                "codigo_carton": carton.codigocarton,
                "jugador": carton.idjugador,
                "estado_carton": carton.estadocarton,
                "fecha_compra": carton.fechacompra,
                "precio_pagado": carton.preciopagado,
                "matriznumeros": carton.matriznumeros,
                "total_participaciones": (
                    len(participaciones) if es_hibrido else 1
                ),
                "rondas_ganadas": sum(
                    cantidad
                    for estado, cantidad in estados.items()
                    if _estado_normalizado(estado) == "ganador"
                ),
                "rondas_pendientes_activas": sum(
                    cantidad
                    for estado, cantidad in estados.items()
                    if _estado_normalizado(estado) in {"pendiente", "en juego"}
                ),
                "estados_participacion": _resumen_estados(estados),
            }
        )
    return resumenes


def _fila_reporte_partida(partida, carton, tipo, participacion=None):
    es_hibrido = tipo == TIPO_HIBRIDO
    return {
        "tipo": tipo,
        "tipo_etiqueta": (
            TIPO_HIBRIDO_ETIQUETA if es_hibrido else TIPO_HISTORICO_ETIQUETA
        ),
        "carton": carton,
        "participacion": participacion,
        "codigo_carton": carton.codigocarton,
        "idcarton": carton.pk,
        "jugador": carton.idjugador,
        "bingo": carton.idbingo if es_hibrido else partida.idbingo,
        "partida": partida,
        "estado_carton": carton.estadocarton,
        "estado_participacion": (
            participacion.estado_participacion if participacion else None
        ),
        "indicevictoria": (
            participacion.indicevictoria
            if participacion
            else carton.indicevictoria
        ),
        "fecha_validacion": (
            participacion.fechavalidacion if participacion else None
        ),
        "matriznumeros": carton.matriznumeros,
        "es_ganador": (
            _estado_normalizado(participacion.estado_participacion) == "ganador"
            if participacion
            else _es_carton_ganador_historico(partida, carton)
        ),
        "precio_pagado": carton.preciopagado,
        "fecha_compra": carton.fechacompra,
    }


def _validar_participacion_hibrida(
    partida,
    carton,
    participacion,
    bingo=None,
):
    idbingo = (bingo or partida.idbingo).pk
    if (
        getattr(carton, "idpartida_id", None) is not None
        or participacion.idcarton_id != carton.pk
        or participacion.idpartida_id != partida.pk
        or participacion.idbingo_id != idbingo
        or getattr(partida, "idbingo_id", None) != idbingo
        or getattr(participacion.idpartida, "idbingo_id", None) != idbingo
        or _idbingo_carton(carton) != idbingo
    ):
        raise ReporteHibridoError(
            "La participación, el cartón, la ronda y el Bingo no coinciden."
        )


def _idbingo_carton(carton):
    idbingo = getattr(carton, "idbingo_id", None)
    if idbingo is not None:
        return idbingo
    partida = getattr(carton, "idpartida", None)
    return getattr(partida, "idbingo_id", None)


def _es_carton_ganador_historico(partida, carton):
    if estado_partida_mostrar(partida.estadopartida) != ESTADO_PARTIDA_FINALIZADA:
        return False
    id_ganador = _id_jugador(getattr(partida, "idjugadorganador", None))
    return id_ganador is not None and _id_jugador(carton.idjugador) == id_ganador


def _resumen_estados(estados):
    if not estados:
        return "-"
    return ", ".join(
        f"{estado}: {cantidad}"
        for estado, cantidad in sorted(estados.items())
    )


def construir_datos_reporte_partida(
    partida,
    cartones=None,
    generado_en=None,
    participaciones_hibridas=None,
    filas=None,
):
    filas = list(
        filas
        if filas is not None
        else construir_filas_reporte_partida(
            partida,
            cartones_historicos=cartones,
            participaciones_hibridas=participaciones_hibridas,
        )
    )
    datos_bolas = preparar_datos_bolas_partida(partida)
    bolas_extraidas = datos_bolas["bolas_extraidas"]
    estado = estado_partida_mostrar(partida.estadopartida)
    finalizada = estado == ESTADO_PARTIDA_FINALIZADA
    ganador = _alias_ganador(partida) if finalizada else None
    fila_ganadora = _fila_ganadora(partida, filas) if finalizada else None

    return {
        "generado_en": generado_en or timezone.localtime(timezone.now()),
        "bingo": partida.idbingo,
        "partida": partida,
        "estado": estado,
        "finalizada": finalizada,
        "bolas_extraidas": bolas_extraidas,
        "bolas_extraidas_codigos": [_formatear_bola(numero) for numero in bolas_extraidas],
        "cantidad_extraida": len(bolas_extraidas),
        "cantidad_restante": 75 - len(bolas_extraidas),
        "ultima_bola": datos_bolas["ultima_bola_codigo"],
        "hubo_desempate": bool(partida.haydesempate),
        "balota_mayor_desempate": _formatear_bola(partida.bolamayordesempate),
        "ganador": ganador,
        "carton_ganador": (
            fila_ganadora["codigo_carton"] if fila_ganadora else None
        ),
        "filas_cartones": filas,
        "mensaje_resultado": (
            "Partida finalizada."
            if finalizada
            else "La partida aún no está finalizada."
        ),
    }


def generar_pdf_reporte_partida(
    partida,
    cartones=None,
    participaciones_hibridas=None,
    filas=None,
):
    datos = construir_datos_reporte_partida(
        partida,
        cartones=cartones,
        participaciones_hibridas=participaciones_hibridas,
        filas=filas,
    )
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        title=f"Reporte de partida {partida.idpartidabingo}",
        pageCompression=0,
        rightMargin=36,
        leftMargin=36,
        topMargin=36,
        bottomMargin=36,
    )
    styles = getSampleStyleSheet()
    story = [
        Paragraph("SIAB / Sistema Integral de Administración de Bingos", styles["Title"]),
        Paragraph("Reporte de partida", styles["Heading1"]),
        Paragraph(f"Generado: {_formatear_fecha(datos['generado_en'])}", styles["Normal"]),
        Spacer(1, 12),
    ]

    bingo = datos["bingo"]
    story.append(Paragraph("Datos del Bingo", styles["Heading2"]))
    story.append(
        _tabla_pdf(
            [
                ("ID de Bingo", bingo.idbingo),
                ("Título", bingo.titulobingo),
                ("Fecha programada", _formatear_fecha(bingo.fechaprogramadabingo)),
                ("Tipo", bingo.tipobingo),
                ("Precio del cartón", _formatear_moneda(bingo.preciocarton)),
                (
                    "Premio mayor",
                    _premio_bingo(bingo),
                ),
            ]
        )
    )
    story.append(Spacer(1, 10))

    partida = datos["partida"]
    story.append(Paragraph("Datos de la partida", styles["Heading2"]))
    story.append(
        _tabla_pdf(
            [
                ("ID de partida", partida.idpartidabingo),
                ("Ronda", partida.nombreronda),
                ("Estado", datos["estado"]),
                ("Inicio", _formatear_fecha(partida.horainicio)),
                ("Finalización", _formatear_fecha(partida.horafin)),
                ("Bolas extraídas", datos["cantidad_extraida"]),
                ("Bolas restantes", datos["cantidad_restante"]),
                ("Última bola", datos["ultima_bola"] or "-"),
                ("Hubo desempate", _si_no(datos["hubo_desempate"])),
                ("Balota mayor de desempate", datos["balota_mayor_desempate"] or "-"),
            ]
        )
    )
    story.append(Spacer(1, 10))

    story.append(Paragraph("Bolas extraídas", styles["Heading2"]))
    bolas = ", ".join(datos["bolas_extraidas_codigos"]) or "Sin bolas extraídas."
    story.append(Paragraph(bolas, styles["Normal"]))
    story.append(Spacer(1, 10))

    story.append(Paragraph("Cartones de la partida", styles["Heading2"]))
    if datos["filas_cartones"]:
        story.append(_tabla_cartones_pdf(datos["filas_cartones"]))
    else:
        story.append(Paragraph("No hay cartones en esta partida.", styles["Normal"]))
    story.append(Spacer(1, 10))

    story.append(Paragraph("Resultado", styles["Heading2"]))
    resultado = [(datos["mensaje_resultado"], "")]
    if datos["finalizada"]:
        resultado.extend(
            [
                ("Ganador", datos["ganador"] or "Sin ganador registrado"),
                ("Resuelta por desempate", _si_no(datos["hubo_desempate"])),
            ]
        )
        if datos["carton_ganador"]:
            resultado.append(("Cartón ganador de la ronda", datos["carton_ganador"]))
    story.append(_tabla_pdf(resultado))

    doc.build(story)
    return buffer.getvalue()


def generar_excel_cartones_partida(
    partida,
    cartones,
    participaciones_hibridas=None,
    filas=None,
):
    filas = list(
        filas
        if filas is not None
        else construir_filas_reporte_partida(
            partida,
            cartones_historicos=cartones,
            participaciones_hibridas=participaciones_hibridas,
        )
    )
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Cartones"

    headers = [
        "Tipo de registro",
        "ID de cartón",
        "Código de cartón",
        "Jugador",
        "Estado del cartón",
        "Estado de participación",
        "Fecha de compra",
        "Precio pagado",
        "Índice de victoria",
        "Fecha de validación",
        "Ganador de la ronda",
        "ID de partida",
        "Nombre de ronda",
        "Estado de partida",
    ]
    worksheet.append(headers)

    total_recaudado = Decimal("0.00")
    estados = Counter()
    for fila in filas:
        precio = _decimal(fila["precio_pagado"])
        total_recaudado += precio
        estados[_estado_normalizado(fila["estado_carton"])] += 1
        worksheet.append(
            [
                fila["tipo_etiqueta"],
                fila["idcarton"],
                fila["codigo_carton"],
                _nombre_jugador(fila["jugador"]),
                fila["estado_carton"],
                fila["estado_participacion"] or "-",
                _fecha_para_excel(fila["fecha_compra"]),
                float(precio),
                fila["indicevictoria"],
                _fecha_para_excel(fila["fecha_validacion"]),
                _si_no(fila["es_ganador"]),
                partida.idpartidabingo,
                partida.nombreronda,
                estado_partida_mostrar(partida.estadopartida),
            ]
        )

    resumen_row = worksheet.max_row + 2
    resumen = [
        ("Total de cartones", len(filas)),
        ("Total recaudado", float(total_recaudado)),
        ("Cartones vendidos", estados.get("vendido", 0)),
        ("Cartones anulados", _cantidad_estados(estados, {"anulado", "anulada", "cancelado", "cancelada"})),
        ("Cartones disponibles", estados.get("disponible", 0)),
    ]
    for offset, (label, value) in enumerate(resumen):
        row = resumen_row + offset
        worksheet.cell(row=row, column=1, value=label)
        worksheet.cell(row=row, column=2, value=value)
    worksheet.cell(row=resumen_row + 1, column=2).number_format = MONEDA_FORMAT

    _aplicar_formato_basico_excel(worksheet)
    _aplicar_formato_columnas(
        worksheet,
        money_columns={8},
        date_columns={7, 10},
    )
    return _workbook_bytes(workbook)


def generar_excel_resumen_bingo(
    bingo,
    partidas,
    cartones,
    participaciones_hibridas=None,
):
    partidas = list(partidas or [])
    cartones = list(cartones or [])
    participaciones_hibridas = list(participaciones_hibridas or [])
    resumenes_maestros = construir_resumenes_cartones_bingo(
        bingo,
        cartones,
        participaciones_hibridas,
    )
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Resumen de partidas"

    headers = [
        "ID de Bingo",
        "Título del Bingo",
        "ID de partida",
        "Ronda",
        "Estado de partida",
        "Fecha programada",
        "Hora de inicio",
        "Hora de finalización",
        "Total de cartones",
        "Cartones históricos",
        "Participaciones híbridas",
        "Recaudación total",
        "Cantidad de bolas extraídas",
        "Ganador",
        "Hubo desempate",
        "Balota mayor de desempate",
    ]
    worksheet.append(headers)

    cartones_por_partida = defaultdict(list)
    for carton in cartones:
        if getattr(carton, "idpartida_id", None) is not None:
            cartones_por_partida[carton.idpartida_id].append(carton)
    participaciones_por_partida = defaultdict(list)
    for participacion in participaciones_hibridas:
        participaciones_por_partida[participacion.idpartida_id].append(
            participacion
        )

    total_cartones = len(resumenes_maestros)
    total_recaudado = sum(
        (_decimal(resumen["precio_pagado"]) for resumen in resumenes_maestros),
        Decimal("0.00"),
    )
    finalizadas = 0
    en_curso = 0
    con_desempate = 0

    for partida in partidas:
        cartones_historicos = cartones_por_partida.get(
            partida.idpartidabingo,
            [],
        )
        participaciones_partida = participaciones_por_partida.get(
            partida.idpartidabingo,
            [],
        )
        filas_partida = construir_filas_reporte_partida(
            partida,
            cartones_historicos,
            participaciones_partida,
        )
        recaudacion = sum(
            (_decimal(fila["precio_pagado"]) for fila in filas_partida),
            Decimal("0.00"),
        )
        bolas = parsear_bolas_cantadas(partida.bolascantadas)
        estado = estado_partida_mostrar(partida.estadopartida)
        finalizadas += 1 if estado == ESTADO_PARTIDA_FINALIZADA else 0
        en_curso += 1 if estado == "En curso" else 0
        con_desempate += 1 if partida.haydesempate else 0
        worksheet.append(
            [
                bingo.idbingo,
                bingo.titulobingo,
                partida.idpartidabingo,
                partida.nombreronda,
                estado,
                _fecha_para_excel(bingo.fechaprogramadabingo),
                _fecha_para_excel(partida.horainicio),
                _fecha_para_excel(partida.horafin),
                len(filas_partida),
                len(cartones_historicos),
                len(participaciones_partida),
                float(recaudacion),
                len(bolas),
                _alias_ganador(partida) or "-",
                _si_no(partida.haydesempate),
                _formatear_bola(partida.bolamayordesempate) or "-",
            ]
        )

    resumen_row = worksheet.max_row + 2
    resumen = [
        ("Total de partidas", len(partidas)),
        ("Partidas finalizadas", finalizadas),
        ("Partidas en curso", en_curso),
        ("Total de cartones", total_cartones),
        ("Recaudación total", float(total_recaudado)),
        ("Total de partidas con desempate", con_desempate),
    ]
    for offset, (label, value) in enumerate(resumen):
        row = resumen_row + offset
        worksheet.cell(row=row, column=1, value=label)
        worksheet.cell(row=row, column=2, value=value)
    worksheet.cell(row=resumen_row + 4, column=2).number_format = MONEDA_FORMAT

    _aplicar_formato_basico_excel(worksheet)
    _aplicar_formato_columnas(
        worksheet,
        money_columns={12},
        date_columns={6, 7, 8},
    )

    inventario = workbook.create_sheet("Cartones del Bingo")
    inventario.append(
        [
            "Tipo de registro",
            "ID de cartón",
            "Código de cartón",
            "Jugador",
            "Estado del cartón",
            "Fecha de compra",
            "Precio pagado",
            "Número de participaciones",
            "Rondas ganadas",
            "Rondas pendientes o activas",
            "Estados de participación",
        ]
    )
    for resumen in resumenes_maestros:
        inventario.append(
            [
                resumen["tipo_etiqueta"],
                resumen["idcarton"],
                resumen["codigo_carton"],
                _nombre_jugador(resumen["jugador"]),
                resumen["estado_carton"],
                _fecha_para_excel(resumen["fecha_compra"]),
                float(_decimal(resumen["precio_pagado"])),
                resumen["total_participaciones"],
                resumen["rondas_ganadas"],
                resumen["rondas_pendientes_activas"],
                resumen["estados_participacion"],
            ]
        )
    _aplicar_formato_basico_excel(inventario)
    _aplicar_formato_columnas(
        inventario,
        money_columns={7},
        date_columns={6},
    )
    return _workbook_bytes(workbook)


def _tabla_pdf(rows):
    table = Table([[str(label), str(value)] for label, value in rows], colWidths=[170, 340])
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#eef3f8")),
                ("TEXTCOLOR", (0, 0), (0, -1), colors.HexColor("#172033")),
                ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#d9e0ea")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]
        )
    )
    return table


def _tabla_cartones_pdf(filas):
    rows = [
        [
            "Tipo",
            "Código",
            "Jugador",
            "Estado cartón",
            "Participación",
            "Índice",
            "Validación",
            "Ganó ronda",
        ]
    ]
    for fila in filas:
        rows.append(
            [
                fila["tipo_etiqueta"],
                fila["codigo_carton"],
                _nombre_jugador(fila["jugador"]),
                fila["estado_carton"],
                fila["estado_participacion"] or "-",
                fila["indicevictoria"] if fila["indicevictoria"] is not None else "-",
                _formatear_fecha(fila["fecha_validacion"]),
                _si_no(fila["es_ganador"]),
            ]
        )
    table = Table(
        rows,
        colWidths=[58, 78, 70, 58, 65, 38, 83, 50],
        repeatRows=1,
    )
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#172033")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 6.5),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#d9e0ea")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 3),
                ("RIGHTPADDING", (0, 0), (-1, -1), 3),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    return table


def _aplicar_formato_basico_excel(worksheet):
    worksheet.freeze_panes = "A2"
    for cell in worksheet[1]:
        cell.font = Font(bold=True)


def _aplicar_formato_columnas(worksheet, money_columns=None, date_columns=None):
    money_columns = money_columns or set()
    date_columns = date_columns or set()
    for column_cells in worksheet.columns:
        column_letter = column_cells[0].column_letter
        max_length = max(
            len(str(cell.value)) if cell.value not in (None, "") else 0
            for cell in column_cells
        )
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 42)
        for cell in column_cells[1:]:
            if cell.column in money_columns:
                cell.number_format = MONEDA_FORMAT
            if cell.column in date_columns:
                cell.number_format = FECHA_FORMAT


def _workbook_bytes(workbook):
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _carton_ganador(partida, cartones):
    id_ganador = getattr(partida, "idjugadorganador_id", None)
    if id_ganador is None and getattr(partida, "idjugadorganador", None):
        id_ganador = partida.idjugadorganador.pk
    if id_ganador is None:
        return None
    for carton in cartones:
        id_jugador = getattr(carton, "idjugador_id", None)
        if id_jugador is None and getattr(carton, "idjugador", None):
            id_jugador = carton.idjugador.pk
        if id_jugador == id_ganador:
            return carton
    return None


def _fila_ganadora(partida, filas):
    id_ganador = _id_jugador(getattr(partida, "idjugadorganador", None))
    if id_ganador is None:
        return None
    for fila in filas:
        if fila["es_ganador"] and _id_jugador(fila["jugador"]) == id_ganador:
            return fila
    for fila in filas:
        if _id_jugador(fila["jugador"]) == id_ganador:
            return fila
    return None


def _id_jugador(jugador):
    if jugador is None:
        return None
    return getattr(jugador, "pk", None)


def _alias_ganador(partida):
    jugador = getattr(partida, "idjugadorganador", None)
    if jugador is None:
        return None
    return getattr(jugador, "aliasjugador", None) or str(jugador)


def _nombre_jugador(jugador):
    if jugador is None:
        return "Sin jugador"
    return getattr(jugador, "aliasjugador", None) or str(jugador)


def _formatear_bola(numero):
    if numero in (None, "", 0, "0"):
        return None
    try:
        return formatear_bola_bingo(int(numero))
    except (TypeError, ValueError):
        return str(numero)


def _premio_bingo(bingo):
    partes = []
    if bingo.premiomayor not in (None, ""):
        partes.append(_formatear_moneda(bingo.premiomayor))
    if bingo.descripcionpremiomayor:
        partes.append(str(bingo.descripcionpremiomayor))
    return " - ".join(partes) or "-"


def _formatear_fecha(value):
    if value in (None, ""):
        return "-"
    if timezone.is_aware(value):
        value = timezone.localtime(value)
    return value.strftime("%d/%m/%Y %H:%M")


def _fecha_para_excel(value):
    if value in (None, ""):
        return None
    if timezone.is_aware(value):
        return timezone.localtime(value).replace(tzinfo=None)
    return value


def _formatear_moneda(value):
    return f"${_decimal(value):.2f}"


def _decimal(value):
    if value in (None, ""):
        return Decimal("0.00")
    return Decimal(str(value))


def _estado_normalizado(value):
    return str(value or "").strip().lower()


def _cantidad_estados(counter, estados):
    return sum(counter.get(estado, 0) for estado in estados)


def _si_no(value):
    return "Sí" if value else "No"
